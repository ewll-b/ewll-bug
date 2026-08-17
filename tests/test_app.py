import io
import json
import sqlite3
import tempfile
import unittest
from unittest import mock
from pathlib import Path

import openpyxl

from app import create_app


class FakeSMTP:
    sent_messages = []
    login_args = []

    def __init__(self, *args, **kwargs) -> None:
        self.args = args
        self.kwargs = kwargs

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        return None

    def ehlo(self) -> None:
        return None

    def starttls(self, context=None) -> None:
        return None

    def login(self, username: str, password: str) -> None:
        self.__class__.login_args.append((username, password))

    def send_message(self, message) -> None:
        self.__class__.sent_messages.append(message)


class FakeGroupReportResponse:
    captured_requests = []

    def __init__(self, request_obj) -> None:
        self.request_obj = request_obj

    def __enter__(self):
        self.__class__.captured_requests.append(self.request_obj)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        return None

    def read(self) -> bytes:
        url = getattr(self.request_obj, "full_url", "")
        if "tenant_access_token/internal" in url:
            return b'{"code":0,"tenant_access_token":"mock-tenant-token"}'
        if "/im/v1/images" in url:
            return b'{"code":0,"data":{"image_key":"img_mock_daily_report"}}'
        return b'{"code":0,"msg":"success"}'


def fake_group_report_urlopen(request_obj, timeout=20):
    return FakeGroupReportResponse(request_obj)


class BugPlatformTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        db_path = Path(self.temp_dir.name) / "test.db"
        upload_dir = Path(self.temp_dir.name) / "uploads"
        self.app = create_app(
            {
                "TESTING": True,
                "SECRET_KEY": "test",
                "DATABASE": str(db_path),
                "UPLOAD_FOLDER": str(upload_dir),
            }
        )
        self.client = self.app.test_client()
        FakeSMTP.sent_messages = []
        FakeSMTP.login_args = []

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def login_as(self, username: str = "lit", password: str = "123456") -> None:
        response = self.client.post(
            "/login",
            data={"username": username, "password": password},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

    def build_excel_file(self) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["用例编号", "标题", "执行结果"])
        sheet.append(["TC-999", "上传的示例用例", "通过"])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_standard_case_excel_file(self) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "标准模板"
        sheet.append(["测试环境：测试环境", "", "", "", "", "", "", "", ""])
        sheet.append(["测试编号", "所属模块", "测试步骤", "预期结果", "实际结果", "", "", "优先级", "备注"])
        sheet.append(["", "", "", "", "ios", "Android", "h5", "", ""])
        sheet.append(["2.7.0-TC-001", "商品负反馈样式优化", "前置条件：用户点击进入商城首页\n1、长按商品", "1、商品样式为蒙层渐变", "pass", "pass", "", "P0", ""])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_case_excel_file_with_sparse_tail_rows(self) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "2.7.0一批"
        sheet.append(["测试环境：测试环境", "", "", "", "", "", "", "", ""])
        sheet.append(["测试编号", "所属模块", "测试步骤", "预期结果", "实际结果", "", "", "优先级", "备注"])
        sheet.append(["", "", "", "", "ios", "Android", "h5", "", ""])
        sheet.append(["2.7.0-TC-015", "编辑模特交互", "步骤015", "预期015", "failed", "pass", "pass", "P1", "震动没有ios"])
        sheet.append(["2.7.0-TC-016", "", "步骤016", "预期016", "pass", "pass", "pass", "P1", ""])
        sheet.append(["2.7.0-TC-017", "", "步骤017", "预期017", "pass", "pass", "pass", "P2", ""])
        sheet.append(["2.7.0-TC-018", "", "步骤018", "预期018", "pass", "pass", "pass", "P1", ""])
        sheet.append(["2.7.0-TC-019", "", "", "", "", "", "", "", ""])
        sheet.append(["2.7.0-TC-020", "", "", "", "", "", "", "", ""])
        sheet.append(["2.7.0-TC-021", "", "", "", "", "", "", "", ""])
        sheet.append(["2.7.0-TC-022", "", "", "", "", "", "", "", ""])
        sheet.append(["2.7.0-TC-023", "", "", "", "", "", "", "", ""])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_case_excel_file_with_image_only_step(self) -> io.BytesIO:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as OpenpyxlImage

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "图片用例"
        sheet.append(["测试环境：测试环境", "", "", "", "", "", "", "", ""])
        sheet.append(["测试编号", "所属模块", "测试步骤", "预期结果", "实际结果", "", "", "优先级", "备注"])
        sheet.append(["", "", "", "", "ios", "Android", "h5", "", ""])
        sheet.append(["2.7.0-TC-IMG", "", "", "", "", "", "", "P1", ""])

        image_stream = io.BytesIO()
        PILImage.new("RGB", (12, 12), "#ff3333").save(image_stream, format="PNG")
        image_stream.seek(0)
        sheet.add_image(OpenpyxlImage(image_stream), "C4")

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def enable_mail_notifications(self) -> None:
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.execute(
                """
                UPDATE mail_settings
                SET enabled = 1, host = ?, port = ?, security = ?, username = ?, password = ?, from_email = ?, sender_name = ?
                WHERE id = 1
                """,
                ("smtp.test.local", 587, "none", "notice@test.local", "secret", "notice@test.local", "Bug Platform"),
            )
            conn.commit()

    def test_login_required(self) -> None:
        response = self.client.get("/bugs", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])

    def test_login_redirect_returns_to_original_detail_page(self) -> None:
        response = self.client.get("/bugs/1?tab=detail", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])
        self.assertIn("next=", response.headers["Location"])
        self.assertIn("%3Ftab%3Ddetail", response.headers["Location"])

        login_response = self.client.post(
            response.headers["Location"],
            data={"username": "lit", "password": "123456"},
            follow_redirects=False,
        )
        self.assertEqual(login_response.status_code, 302)
        self.assertEqual(login_response.headers["Location"], "/bugs/1?tab=detail")

    def test_login_rejects_external_next_redirects(self) -> None:
        response = self.client.post(
            "/login?next=https://third-party.example/path",
            data={"username": "lit", "password": "123456"},
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs")

    def test_bug_list_loads_after_login(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/bugs")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Bug列表".encode("utf-8"), response.data)

    def test_bug_list_pagination_controls_filtered_results(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            creator = conn.execute("SELECT id FROM users WHERE username = 'lit'").fetchone()
            assignee = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.assertIsNotNone(creator)
            self.assertIsNotNone(assignee)
            rows = []
            for index in range(23):
                rows.append(
                    (
                        f"PG-{index + 1:03d}",
                        f"分页测试 Bug {index + 1:02d}",
                        "page-test",
                        "APP",
                        "Android",
                        "高",
                        "高",
                        "open",
                        assignee["id"],
                        creator["id"],
                        "李婷",
                        "测试环境",
                        "用于验证 Bug 列表分页",
                        "正常分页",
                        "分页异常",
                        f"2026-07-14 11:{index:02d}:00",
                    )
                )
            conn.executemany(
                """
                INSERT INTO bugs (
                    bug_no, title, version, module, platform, severity, priority, status,
                    assignee_id, creator_id, reporter, environment, description,
                    expected_result, actual_result, project_id, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                [(*row, row[-1]) for row in rows],
            )
            conn.commit()

        response = self.client.get("/bugs?version=page-test&page=2")
        self.assertEqual(response.status_code, 200)
        self.assertIn("共 23 条，当前显示 21 - 23".encode("utf-8"), response.data)
        self.assertIn("第 2 / 2 页".encode("utf-8"), response.data)
        self.assertIn('href="/bugs?version=page-test&amp;page=1"'.encode("utf-8"), response.data)
        self.assertIn('aria-current="page">2'.encode("utf-8"), response.data)
        self.assertIn('class="pager-action"'.encode("utf-8"), response.data)
        self.assertIn('class="page-number active" aria-current="page">2'.encode("utf-8"), response.data)
        self.assertNotIn('class="page-jump-form"'.encode("utf-8"), response.data)
        self.assertNotIn('btn btn-secondary mini page-number'.encode("utf-8"), response.data)

    def test_bug_list_supports_multi_select_filters(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            creator = conn.execute("SELECT id FROM users WHERE username = 'lit'").fetchone()
            assignee = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            rows = [
                ("MF-001", "多选筛选 A", "multi-a", "WEB", "WEB", "高", "高", "open"),
                ("MF-002", "多选筛选 B", "multi-b", "APP", "Android", "高", "高", "pending_verification"),
                ("MF-003", "多选筛选 C", "multi-c", "Backend", "后端", "高", "高", "closed"),
            ]
            conn.executemany(
                """
                INSERT INTO bugs (
                    bug_no, title, version, module, platform, severity, priority, status,
                    assignee_id, creator_id, reporter, environment, description,
                    expected_result, actual_result, project_id, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                [
                    (
                        bug_no,
                        title,
                        version,
                        module,
                        platform,
                        severity,
                        priority,
                        status,
                        assignee["id"],
                        creator["id"],
                        "李婷",
                        "测试环境",
                        "验证多选筛选",
                        "筛选正确",
                        "待确认",
                        "2026-07-15 10:00:00",
                        "2026-07-15 10:00:00",
                    )
                    for bug_no, title, version, module, platform, severity, priority, status in rows
                ],
            )
            conn.commit()

        response = self.client.get("/bugs?version=multi-a&version=multi-b&status=open&status=pending_verification")

        self.assertEqual(response.status_code, 200)
        self.assertIn("共 2 条".encode("utf-8"), response.data)
        self.assertIn("版本(2)".encode("utf-8"), response.data)
        self.assertIn("状态(2)".encode("utf-8"), response.data)
        self.assertIn("多选筛选 A".encode("utf-8"), response.data)
        self.assertIn("多选筛选 B".encode("utf-8"), response.data)
        self.assertNotIn("多选筛选 C".encode("utf-8"), response.data)

    def test_bug_list_summary_uses_all_active_filters(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            creator = conn.execute("SELECT id FROM users WHERE username = 'lit'").fetchone()
            assignee = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            rows = [
                ("FS-001", "筛选统计 WEB 打开", "summary-filter", "WEB", "open", "2026-07-15 10:00:00"),
                ("FS-002", "筛选统计 WEB 待验证", "summary-filter", "WEB", "pending_verification", "2026-07-16 10:00:00"),
                ("FS-003", "筛选统计 Android 关闭", "summary-filter", "Android", "closed", "2026-07-17 10:00:00"),
            ]
            conn.executemany(
                """
                INSERT INTO bugs (
                    bug_no, title, version, module, platform, severity, priority, status,
                    assignee_id, creator_id, reporter, environment, description,
                    expected_result, actual_result, project_id, created_at, updated_at
                )
                VALUES (?, ?, ?, 'APP', ?, '高', '高', ?, ?, ?, '李婷', '测试环境', '验证筛选统计', '统计正确', '待确认', 1, ?, ?)
                """,
                [
                    (
                        bug_no,
                        title,
                        version,
                        platform,
                        status,
                        assignee["id"],
                        creator["id"],
                        created_at,
                        created_at,
                    )
                    for bug_no, title, version, platform, status, created_at in rows
                ],
            )
            conn.commit()

        response = self.client.get("/bugs?version=summary-filter&platform=WEB")

        self.assertEqual(response.status_code, 200)
        self.assertIn("共 2 条".encode("utf-8"), response.data)
        self.assertIn("总数：2".encode("utf-8"), response.data)
        self.assertIn("待处理：1".encode("utf-8"), response.data)
        self.assertIn("待验证：1".encode("utf-8"), response.data)
        self.assertIn("已关闭：0".encode("utf-8"), response.data)
        self.assertIn("筛选统计 WEB 打开".encode("utf-8"), response.data)
        self.assertIn("筛选统计 WEB 待验证".encode("utf-8"), response.data)
        self.assertNotIn("筛选统计 Android 关闭".encode("utf-8"), response.data)

    def test_bug_create_form_prefills_selected_version(self) -> None:
        self.login_as("lit", "123456")

        list_response = self.client.get("/bugs?version=2.8.0")
        self.assertEqual(list_response.status_code, 200)
        self.assertIn('<input type="text" name="version" value="2.8.0"'.encode("utf-8"), list_response.data)

        new_response = self.client.get("/bugs/new?version=2.8.0")
        self.assertEqual(new_response.status_code, 200)
        self.assertIn('<input type="text" name="version" value="2.8.0"'.encode("utf-8"), new_response.data)

    def test_bug_create_form_includes_web_platform(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.get("/bugs/new")

        self.assertEqual(response.status_code, 200)
        self.assertIn('data-filter-value="WEB"'.encode("utf-8"), response.data)

    def test_bug_assignee_choices_exclude_admin(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            admin_user = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
        self.assertIsNotNone(admin_user)

        response = self.client.get("/bugs")

        self.assertEqual(response.status_code, 200)
        self.assertIn("周越 · APP 开发".encode("utf-8"), response.data)
        self.assertNotIn("Admin · 系统管理员".encode("utf-8"), response.data)
        self.assertNotIn('data-filter-label="Admin"'.encode("utf-8"), response.data)
        self.assertNotIn(
            f'name="assignee_id" value="{admin_user["id"]}"'.encode("utf-8"),
            response.data,
        )
        self.assertNotIn(
            f'data-filter-target="bug-form-assignee" data-filter-value="{admin_user["id"]}"'.encode("utf-8"),
            response.data,
        )

    def test_create_bug_rejects_admin_assignee(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            admin_user = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
        self.assertIsNotNone(admin_user)

        response = self.client.post(
            "/bugs/new?format=json",
            data={
                "title": "不能指派 admin 的 Bug",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "高",
                "assignee_id": str(admin_user["id"]),
                "environment": "测试环境",
                "description": "验证管理员不能成为缺陷处理人",
                "expected_result": "提交失败",
                "actual_result": "待验证",
            },
            headers={"Accept": "application/json"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["message"], "缺陷处理人不能选择 admin 账号。")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            bug_count = conn.execute(
                "SELECT COUNT(*) FROM bugs WHERE title = ?",
                ("不能指派 admin 的 Bug",),
            ).fetchone()[0]
        self.assertEqual(bug_count, 0)

    def test_bug_detail_edit_persists_changes_and_bumps_sync_token(self) -> None:
        self.login_as("admin", "admin123")
        with self.client.session_transaction() as session_state:
            before_token = session_state.get("bug_sync_token")

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            bug = conn.execute("SELECT id, requirement_id, case_id FROM bugs ORDER BY id LIMIT 1").fetchone()
            assignee = conn.execute("SELECT id FROM users WHERE username <> 'admin' ORDER BY id LIMIT 1").fetchone()
            requirement = conn.execute("SELECT id FROM requirements ORDER BY id LIMIT 1").fetchone()
            case_item = conn.execute("SELECT id FROM test_cases ORDER BY id LIMIT 1").fetchone()

        self.assertIsNotNone(bug)
        self.assertIsNotNone(assignee)

        response = self.client.post(
            f"/bugs/{bug['id']}/edit",
            data={
                "title": "详情页编辑后的标题",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "最高",
                "assignee_id": str(assignee["id"]),
                "requirement_id": str((bug["requirement_id"] or (requirement["id"] if requirement else "")) or ""),
                "case_id": str((bug["case_id"] or (case_item["id"] if case_item else "")) or ""),
                "environment": "测试环境-详情页编辑",
                "description": "详情页编辑后的问题描述",
                "expected_result": "详情页编辑后的期望结果",
                "actual_result": "详情页编辑后的实际结果",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Bug 信息已更新。".encode("utf-8"), response.data)
        self.assertIn("详情页编辑后的标题".encode("utf-8"), response.data)
        self.assertIn("严重级别".encode("utf-8"), response.data)
        self.assertIn("最高".encode("utf-8"), response.data)
        self.assertNotIn("优先级".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            updated_bug = conn.execute(
                """
                SELECT title, version, module, platform, severity, priority, assignee_id, environment, description,
                       expected_result, actual_result
                FROM bugs
                WHERE id = ?
                """,
                (bug["id"],),
            ).fetchone()

        self.assertEqual(updated_bug["title"], "详情页编辑后的标题")
        self.assertEqual(updated_bug["version"], "2.9.0")
        self.assertEqual(updated_bug["module"], "APP")
        self.assertEqual(updated_bug["platform"], "Android")
        self.assertEqual(updated_bug["severity"], "最高")
        self.assertEqual(updated_bug["priority"], "最高")
        self.assertEqual(updated_bug["assignee_id"], assignee["id"])
        self.assertEqual(updated_bug["environment"], "测试环境-详情页编辑")
        self.assertEqual(updated_bug["description"], "详情页编辑后的问题描述")
        self.assertEqual(updated_bug["expected_result"], "详情页编辑后的期望结果")
        self.assertEqual(updated_bug["actual_result"], "详情页编辑后的实际结果")

        with self.client.session_transaction() as session_state:
            after_token = session_state.get("bug_sync_token")
        self.assertNotEqual(before_token, after_token)

    def test_admin_can_update_severity_for_bug_created_by_others(self) -> None:
        self.login_as("admin", "admin123")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            admin_user = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
            bug = conn.execute(
                """
                SELECT *
                FROM bugs
                WHERE creator_id <> ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (admin_user["id"],),
            ).fetchone()

        self.assertIsNotNone(bug)
        next_severity = "建议" if bug["severity"] != "建议" else "高"

        response = self.client.post(
            f"/bugs/{bug['id']}/edit",
            data={
                "title": bug["title"],
                "version": bug["version"],
                "module": bug["module"],
                "platform": bug["platform"],
                "severity": next_severity,
                "assignee_id": str(bug["assignee_id"]),
                "requirement_id": str(bug["requirement_id"] or ""),
                "case_id": str(bug["case_id"] or ""),
                "environment": bug["environment"] or "",
                "description": bug["description"],
                "expected_result": bug["expected_result"] or "",
                "actual_result": bug["actual_result"] or "",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Bug 信息已更新。".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            updated_bug = conn.execute("SELECT severity FROM bugs WHERE id = ?", (bug["id"],)).fetchone()
        self.assertEqual(updated_bug["severity"], next_severity)

    def test_admin_user_role_and_account_type_are_saved_independently(self) -> None:
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/admin",
            data={
                "entity": "user",
                "action": "create",
                "next": "/admin/users",
                "name": "管理员测试账号",
                "account_type": "admin",
                "role_code": "backend_developer",
                "username": "admin_backend",
                "password": "123456",
                "email": "admin_backend@alvinsclub.ai",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("账号已创建".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            created_user = conn.execute(
                "SELECT account_type, role_code, role FROM users WHERE username = ?",
                ("admin_backend",),
            ).fetchone()

        self.assertEqual(created_user["account_type"], "admin")
        self.assertEqual(created_user["role_code"], "backend_developer")
        self.assertEqual(created_user["role"], "后端开发")

    def test_my_todos_page_loads(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            assignee = conn.execute("SELECT id FROM users WHERE username = 'lit'").fetchone()
            conn.execute(
                """
                INSERT INTO bugs (
                    bug_no, title, project_id, version, module, platform, severity, priority, status,
                    assignee_id, creator_id, reporter, environment, description,
                    expected_result, actual_result, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "TODO-001",
                    "待办端字段测试",
                    1,
                    "2.9.0",
                    "WEB",
                    "WEB",
                    "高",
                    "高",
                    "open",
                    assignee["id"],
                    assignee["id"],
                    "李婷",
                    "测试环境",
                    "验证我的待办展示端字段",
                    "显示 WEB",
                    "未显示",
                    "2026-07-28 10:00:00",
                    "2026-07-28 10:00:00",
                ),
            )
            conn.commit()
        response = self.client.get("/todos")
        self.assertEqual(response.status_code, 200)
        self.assertIn("我的待办".encode("utf-8"), response.data)
        self.assertIn('<th class="bug-col-platform">端</th>'.encode("utf-8"), response.data)
        self.assertIn("platform-chip".encode("utf-8"), response.data)
        self.assertIn("WEB".encode("utf-8"), response.data)

    def test_my_todos_status_change_to_closed_stays_on_todos(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.post(
            "/bugs/3/update",
            data={
                "action": "change_status",
                "status": "closed",
                "redirect_to": "/todos",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/todos")

    def test_case_library_loads(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/cases")
        self.assertEqual(response.status_code, 200)
        self.assertIn("用例库".encode("utf-8"), response.data)
        self.assertIn("打开在线文档".encode("utf-8"), response.data)

    def test_case_library_can_move_document_between_folders(self) -> None:
        self.login_as("admin", "admin123")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            document = conn.execute(
                """
                SELECT id, project_id, version, folder_name, doc_name
                FROM test_cases
                WHERE project_id = 1
                ORDER BY id
                LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(document)

        response = self.client.post(
            "/cases/manage",
            data={
                "action": "move_document",
                "document_id": str(document["id"]),
                "folder_name": "回归用例",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("folder=", response.headers["Location"])

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            moved_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = ?
                    AND COALESCE(version, '') = COALESCE(?, '')
                    AND folder_name = '回归用例'
                    AND COALESCE(doc_name, '') = COALESCE(?, '')
                """,
                (document["project_id"], document["version"], document["doc_name"]),
            ).fetchone()["count"]
            old_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = ?
                    AND COALESCE(version, '') = COALESCE(?, '')
                    AND folder_name = ?
                    AND COALESCE(doc_name, '') = COALESCE(?, '')
                """,
                (document["project_id"], document["version"], document["folder_name"], document["doc_name"]),
            ).fetchone()["count"]

        self.assertGreater(moved_count, 0)
        self.assertEqual(old_count, 0)

        page = self.client.get("/cases?folder=%E5%9B%9E%E5%BD%92%E7%94%A8%E4%BE%8B")
        self.assertEqual(page.status_code, 200)
        self.assertIn("当前文件夹：回归用例".encode("utf-8"), page.data)
        self.assertIn('data-folder-drop-zone'.encode("utf-8"), page.data)
        self.assertIn('draggable="true"'.encode("utf-8"), page.data)

    def test_case_library_create_folder_without_document(self) -> None:
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/cases/manage",
            data={"action": "create_folder", "folder_name": "空文件夹"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("当前文件夹：空文件夹".encode("utf-8"), response.data)
        self.assertIn("当前文件夹暂无在线文档".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            folder = conn.execute(
                """
                SELECT id
                FROM case_folders
                WHERE project_id = 1 AND name = ?
                """,
                ("空文件夹",),
            ).fetchone()
            auto_document_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = 1
                    AND folder_name = ?
                    AND doc_name = ?
                """,
                ("空文件夹", "空文件夹-在线文档"),
            ).fetchone()["count"]

        self.assertIsNotNone(folder)
        self.assertEqual(auto_document_count, 0)

    def test_case_document_loads_in_new_page(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/cases/1")
        self.assertEqual(response.status_code, 200)
        self.assertIn("在线文档".encode("utf-8"), response.data)
        self.assertIn("测试编号".encode("utf-8"), response.data)
        self.assertIn("data-autosave-url".encode("utf-8"), response.data)
        self.assertIn("已实时保存".encode("utf-8"), response.data)

    def test_case_document_autosave_updates_single_field(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            before = conn.execute(
                "SELECT case_no, expected_result FROM test_cases WHERE id = 1"
            ).fetchone()

        response = self.client.post(
            "/cases/1/autosave",
            data={
                "case_id": "1",
                "field": "steps",
                "value": "实时保存步骤",
            },
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["ok"], True)
        self.assertEqual(response.json["field"], "steps")

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                "SELECT case_no, steps, expected_result FROM test_cases WHERE id = 1"
            ).fetchone()

        self.assertEqual(row["steps"], "实时保存步骤")
        self.assertEqual(row["case_no"], before["case_no"])
        self.assertEqual(row["expected_result"], before["expected_result"])

    def test_case_document_autosave_updates_result_status(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/cases/1/autosave",
            data={
                "case_id": "1",
                "field": "ios_result",
                "value": "failed",
            },
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["ok"], True)
        self.assertEqual(response.json["execute_status"], "失败")

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT ios_result, execute_status FROM test_cases WHERE id = 1").fetchone()

        self.assertEqual(row["ios_result"], "failed")
        self.assertEqual(row["execute_status"], "失败")

    def test_case_document_autosave_persists_dynamic_cell(self) -> None:
        self.login_as("lit", "123456")
        self.client.post(
            "/cases/1/update",
            data={
                "document_action": "add_column",
                "new_column_name": "实时备注",
            },
            follow_redirects=True,
        )
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            column = conn.execute(
                """
                SELECT id
                FROM case_document_columns
                WHERE project_id = 3 AND version = '2.6.0' AND folder_name = '测试用例' AND doc_name = '2.6.0-首页优化测试用例'
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()

        response = self.client.post(
            "/cases/1/autosave",
            data={
                "case_id": "1",
                "field": f"dynamic_{column['id']}",
                "value": "动态列实时保存",
            },
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["ok"], True)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            cell = conn.execute(
                "SELECT cell_value FROM case_document_cells WHERE column_id = ? AND case_id = ?",
                (column["id"], 1),
            ).fetchone()

        self.assertIsNotNone(cell)
        self.assertEqual(cell["cell_value"], "动态列实时保存")

    def test_case_document_update_syncs_status(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/cases/1/update",
            data={
                "case_no_1": "2.6.0-HOME-TC-001A",
                "priority_level_1": "P0",
                "module_name_1": "首页推荐新模块",
                "steps_1": "1. 打开首页\\n2. 点击新入口",
                "expected_result_1": "成功跳转到新页面",
                "ios_result_1": "failed",
                "android_result_1": "",
                "h5_result_1": "",
                "remark_1": "iOS 失败",
                "executor_1": "李婷",
                "ios_result_2": "pass",
                "android_result_2": "",
                "h5_result_2": "",
                "remark_2": "",
                "ios_result_3": "block",
                "android_result_3": "",
                "h5_result_3": "",
                "remark_3": "阻塞",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("在线文档已保存".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT case_no, module_name, steps, expected_result, execute_status, remark, executor FROM test_cases WHERE id = 1").fetchone()
        self.assertEqual(row["case_no"], "2.6.0-HOME-TC-001A")
        self.assertEqual(row["module_name"], "首页推荐新模块")
        self.assertEqual(row["steps"], "1. 打开首页\\n2. 点击新入口")
        self.assertEqual(row["expected_result"], "成功跳转到新页面")
        self.assertEqual(row["execute_status"], "失败")
        self.assertEqual(row["remark"], "iOS 失败")
        self.assertEqual(row["executor"], "李婷")

    def test_case_document_update_allows_non_tester_roles(self) -> None:
        self.login_as("zhouyue", "123456")
        response = self.client.post(
            "/cases/1/update",
            data={
                "case_no_1": "2.6.0-HOME-TC-001B",
                "priority_level_1": "P1",
                "module_name_1": "首页推荐",
                "steps_1": "1. 打开首页",
                "expected_result_1": "页面正常展示",
                "ios_result_1": "pass",
                "android_result_1": "",
                "h5_result_1": "",
                "remark_1": "开发补充验证",
                "executor_1": "周越",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("在线文档已保存".encode("utf-8"), response.data)

    def test_case_document_can_add_row(self) -> None:
        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            before_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = 3 AND version = '2.6.0' AND folder_name = '测试用例' AND doc_name = '2.6.0-首页优化测试用例'
                """
            ).fetchone()["count"]

        response = self.client.post(
            "/cases/1/update",
            data={"document_action": "add_row"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已新增一行".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            after_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = 3 AND version = '2.6.0' AND folder_name = '测试用例' AND doc_name = '2.6.0-首页优化测试用例'
                """
            ).fetchone()["count"]
            new_case = conn.execute(
                """
                SELECT case_no, execute_status
                FROM test_cases
                WHERE project_id = 3 AND version = '2.6.0' AND folder_name = '测试用例' AND doc_name = '2.6.0-首页优化测试用例'
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()

        self.assertEqual(after_count, before_count + 1)
        self.assertEqual(new_case["case_no"], "2.6.0-HOME-TC-004")
        self.assertEqual(new_case["execute_status"], "未测")

    def test_case_document_can_add_column_and_persist_cell_value(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/cases/1/update",
            data={
                "document_action": "add_column",
                "new_column_name": "复测备注",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已新增一列".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            column = conn.execute(
                """
                SELECT id, column_name
                FROM case_document_columns
                WHERE project_id = 3 AND version = '2.6.0' AND folder_name = '测试用例' AND doc_name = '2.6.0-首页优化测试用例'
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()

        self.assertIsNotNone(column)
        self.assertEqual(column["column_name"], "复测备注")

        response = self.client.post(
            "/cases/1/update",
            data={f"dynamic_{column['id']}_1": "需要重点回归"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("在线文档已保存".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            cell = conn.execute(
                """
                SELECT cell_value
                FROM case_document_cells
                WHERE column_id = ? AND case_id = ?
                """,
                (column["id"], 1),
            ).fetchone()

        self.assertIsNotNone(cell)
        self.assertEqual(cell["cell_value"], "需要重点回归")

        page = self.client.get("/cases/1")
        self.assertEqual(page.status_code, 200)
        self.assertIn("复测备注".encode("utf-8"), page.data)
        self.assertIn("需要重点回归".encode("utf-8"), page.data)

    def test_case_document_delete_case_item_unlinks_bug_and_keeps_document_open(self) -> None:
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/cases/1/items/1/delete",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("用例已删除".encode("utf-8"), response.data)
        self.assertNotIn("2.6.0-HOME-TC-001".encode("utf-8"), response.data)
        self.assertIn("2.6.0-HOME-TC-002".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            deleted_case = conn.execute("SELECT id FROM test_cases WHERE id = 1").fetchone()
            bug = conn.execute("SELECT case_id FROM bugs WHERE id = 1").fetchone()
        self.assertIsNone(deleted_case)
        self.assertIsNone(bug["case_id"])

    def test_case_upload(self) -> None:
        self.login_as("lit", "123456")
        excel = self.build_excel_file()
        response = self.client.post(
            "/cases/upload",
            data={"excel_file": (excel, "cases.xlsx"), "folder_name": "测试组"},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已同步".encode("utf-8"), response.data)

    def test_case_upload_uses_selected_folder(self) -> None:
        self.login_as("lit", "123456")
        excel = self.build_standard_case_excel_file()
        response = self.client.post(
            "/cases/upload",
            data={"excel_file": (excel, "folder_cases.xlsx"), "folder_name": "回归上传"},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("folder=%E5%9B%9E%E5%BD%92%E4%B8%8A%E4%BC%A0", response.headers["Location"])

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT folder_name, doc_name
                FROM test_cases
                WHERE case_no = ?
                """,
                ("2.7.0-TC-001",),
            ).fetchone()
            excel_import_count = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE folder_name = 'Excel导入'
                    AND doc_name = ?
                """,
                ("folder_cases",),
            ).fetchone()["count"]

        self.assertIsNotNone(row)
        self.assertEqual(row["folder_name"], "回归上传")
        self.assertEqual(row["doc_name"], "folder_cases")
        self.assertEqual(excel_import_count, 0)

    def test_case_upload_standard_template_keeps_columns_aligned(self) -> None:
        self.login_as("lit", "123456")
        excel = self.build_standard_case_excel_file()
        response = self.client.post(
            "/cases/upload",
            data={"excel_file": (excel, "standard_cases.xlsx"), "folder_name": "测试组"},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已同步".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT case_no, title, priority_level, module_name, steps, expected_result, ios_result, android_result, h5_result FROM test_cases WHERE case_no = ?", ("2.7.0-TC-001",)).fetchone()

        self.assertEqual(row["title"], "商品负反馈样式优化")
        self.assertEqual(row["priority_level"], "P0")
        self.assertEqual(row["module_name"], "商品负反馈样式优化")
        self.assertEqual(row["steps"], "前置条件：用户点击进入商城首页\n1、长按商品")
        self.assertEqual(row["expected_result"], "1、商品样式为蒙层渐变")
        self.assertEqual(row["ios_result"], "pass")
        self.assertEqual(row["android_result"], "pass")
        self.assertEqual(row["h5_result"], "")

    def test_case_upload_imports_row_with_only_excel_image_content(self) -> None:
        self.login_as("lit", "123456")
        excel = self.build_case_excel_file_with_image_only_step()
        response = self.client.post(
            "/cases/upload",
            data={"excel_file": (excel, "image_cases.xlsx"), "folder_name": "测试组"},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已同步".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                "SELECT case_no, title, steps FROM test_cases WHERE case_no = ?",
                ("2.7.0-TC-IMG",),
            ).fetchone()

        self.assertIsNotNone(row)
        self.assertEqual(row["title"], "2.7.0-TC-IMG")
        self.assertIn("原Excel含图片", row["steps"])

    def test_case_upload_skips_sparse_rows_with_only_case_number(self) -> None:
        self.login_as("lit", "123456")
        excel = self.build_case_excel_file_with_sparse_tail_rows()
        response = self.client.post(
            "/cases/upload",
            data={"excel_file": (excel, "sparse_tail_cases.xlsx"), "folder_name": "测试组"},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已同步".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT case_no, module_name, steps, expected_result
                FROM test_cases
                WHERE doc_name = ?
                ORDER BY case_no
                """,
                ("sparse_tail_cases",),
            ).fetchall()

        self.assertEqual([row["case_no"] for row in rows], ["2.7.0-TC-015", "2.7.0-TC-016", "2.7.0-TC-017", "2.7.0-TC-018"])
        self.assertEqual(rows[1]["module_name"], "编辑模特交互")
        self.assertEqual(rows[1]["steps"], "步骤016")
        self.assertEqual(rows[1]["expected_result"], "预期016")

    def test_chart_endpoint(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/reports/testing/chart.png")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "image/svg+xml")

    def test_project_switch(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post("/switch-project", data={"project_id": "2"}, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("智能客服系统".encode("utf-8"), response.data)

    def test_create_bug_uses_logged_in_user_as_creator(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "新建测试 Bug",
                "version": "2.9.0",
                "module": "WEB",
                "platform": "iOS",
                "severity": "中",
                "assignee_id": "2",
                "environment": "测试环境",
                "description": "步骤 1，步骤 2，出现错误",
                "expected_result": "正常",
                "actual_result": "异常",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("新建测试 Bug".encode("utf-8"), response.data)
        self.assertIn("李婷".encode("utf-8"), response.data)

    def test_create_bug_with_attachment_and_open_it(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "带附件的 Bug",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "高",
                "assignee_id": "2",
                "environment": "iOS",
                "description": "附带日志和截图",
                "expected_result": "正常执行",
                "actual_result": "发生异常",
                "attachments": [
                    (io.BytesIO(b"log content"), "app.log"),
                ],
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("app.log".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            bug = conn.execute("SELECT id FROM bugs WHERE title = ?", ("带附件的 Bug",)).fetchone()
            attachment = conn.execute("SELECT id FROM bug_attachments WHERE bug_id = ?", (bug["id"],)).fetchone()

        detail = self.client.get(f"/bugs/{bug['id']}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn("app.log".encode("utf-8"), detail.data)

        attachment_response = self.client.get(f"/attachments/{attachment['id']}")
        self.assertEqual(attachment_response.status_code, 200)
        self.assertIn(attachment_response.mimetype, {"text/plain", "application/octet-stream"})

    def test_create_bug_with_inline_image_source_shows_in_field(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "带字段截图的 Bug",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "高",
                "assignee_id": "2",
                "environment": "iOS",
                "description": "这里应该展示截图",
                "expected_result": "正常执行",
                "actual_result": "发生异常",
                "inline_image_sources": "description",
                "inline_images": [
                    (io.BytesIO(b"fake png"), "screen.png"),
                ],
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            bug = conn.execute("SELECT id FROM bugs WHERE title = ?", ("带字段截图的 Bug",)).fetchone()
            attachment = conn.execute(
                "SELECT source_field FROM bug_attachments WHERE bug_id = ?",
                (bug["id"],),
            ).fetchone()

        self.assertEqual(attachment["source_field"], "description")
        detail = self.client.get(f"/bugs/{bug['id']}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn(b"inline-attachment-gallery", detail.data)
        self.assertIn(b"data-image-preview-src", detail.data)
        self.assertIn("暂无附件".encode("utf-8"), detail.data)
        self.assertNotIn("图片已展示在对应字段".encode("utf-8"), detail.data)
        self.assertNotIn(b"attachment-card compact", detail.data)

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_create_bug_sends_project_group_notification(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            cursor = conn.execute(
                """
                INSERT INTO requirements (
                    project_id, code, title, version, status, priority, description,
                    acceptance_criteria, creator_id, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    1,
                    "REQ-GROUP-001",
                    "项目群通知需求",
                    "2.9.0",
                    "pending",
                    "中",
                    "用于验证群通知卡片的需求",
                    "通知卡片可打开需求页",
                    1,
                    "2026-07-29 10:00:00",
                    "2026-07-29 10:00:00",
                ),
            )
            requirement_id = cursor.lastrowid
            conn.execute(
                """
                UPDATE projects
                SET bug_notify_enabled = 1,
                    bug_notify_webhook = ?,
                    bug_notify_secret = '',
                    bug_notify_base_url = ?
                WHERE id = 1
                """,
                (
                    "https://open.feishu.cn/open-apis/bot/v2/hook/project-webhook",
                    "http://bug.test.local",
                ),
            )
            conn.commit()

        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "需要发送到项目群的Bug",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "中",
                "assignee_id": "2",
                "requirement_id": str(requirement_id),
                "environment": "测试环境",
                "description": "创建后需要同步到项目群",
                "expected_result": "功能正常",
                "actual_result": "发生异常",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("新建Bug群通知已发送".encode("utf-8"), response.data)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 1)

        payload = json.loads(FakeGroupReportResponse.captured_requests[0].data.decode("utf-8"))
        payload_text = json.dumps(payload, ensure_ascii=False)
        self.assertEqual(payload["msg_type"], "interactive")
        self.assertEqual(payload["card"]["header"]["template"], "red")
        self.assertEqual(payload["card"]["header"]["title"]["content"], "新建 Bug")
        self.assertIn("需要发送到项目群的Bug", payload_text)
        self.assertIn("项目：** 零售增长平台", payload_text)
        self.assertIn("负责人：** 周越", payload_text)
        self.assertIn("查看 Bug 详情", payload_text)
        self.assertIn("进入需求页", payload_text)
        self.assertIn("AI处理", payload_text)
        self.assertIn("http://bug.test.local/bugs/", payload_text)
        self.assertIn("http://bug.test.local/requirements/", payload_text)
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            created_bug_id = conn.execute("SELECT id FROM bugs WHERE title = ?", ("需要发送到项目群的Bug",)).fetchone()[0]
        action_element = next(element for element in payload["card"]["elements"] if element.get("tag") == "action")
        buttons = {button["text"]["content"]: button for button in action_element["actions"]}
        detail_url = f"http://bug.test.local/bugs/{created_bug_id}"
        self.assertEqual(buttons["查看 Bug 详情"]["type"], "primary")
        self.assertEqual(buttons["查看 Bug 详情"]["url"], detail_url)
        self.assertEqual(
            buttons["查看 Bug 详情"]["multi_url"],
            {
                "url": detail_url,
                "pc_url": detail_url,
                "ios_url": detail_url,
                "android_url": detail_url,
            },
        )
        self.assertEqual(buttons["进入需求页"]["url"], f"http://bug.test.local/requirements/{requirement_id}")
        self.assertEqual(buttons["AI处理"]["url"], f"{detail_url}?tab=process")
        self.assertEqual(FakeGroupReportResponse.captured_requests[0].full_url, "https://open.feishu.cn/open-apis/bot/v2/hook/project-webhook")

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_create_bug_uses_platform_group_notification_before_default(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.execute(
                """
                UPDATE projects
                SET bug_notify_enabled = 1,
                    bug_notify_webhook = ?,
                    bug_notify_secret = '',
                    bug_notify_base_url = ?
                WHERE id = 1
                """,
                (
                    "https://open.feishu.cn/open-apis/bot/v2/hook/default-webhook",
                    "http://bug.test.local",
                ),
            )
            conn.execute(
                """
                INSERT INTO project_bug_notify_rules (
                    project_id, module, enabled, webhook_url, secret, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    1,
                    "H5",
                    1,
                    "https://open.feishu.cn/open-apis/bot/v2/hook/h5-webhook",
                    "",
                    "2026-07-13 10:00:00",
                    "2026-07-13 10:00:00",
                ),
            )
            conn.commit()

        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "H5页面按钮不可点击",
                "version": "2.9.0",
                "platform": "H5",
                "severity": "高",
                "assignee_id": "3",
                "environment": "测试环境",
                "description": "H5 页面操作异常，需要同步到 H5 群",
                "expected_result": "按钮可点击",
                "actual_result": "点击无响应",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("H5群".encode("utf-8"), response.data)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 1)
        self.assertEqual(FakeGroupReportResponse.captured_requests[0].full_url, "https://open.feishu.cn/open-apis/bot/v2/hook/h5-webhook")

        payload = json.loads(FakeGroupReportResponse.captured_requests[0].data.decode("utf-8"))
        payload_text = json.dumps(payload, ensure_ascii=False)
        self.assertEqual(payload["msg_type"], "interactive")
        self.assertIn("H5页面按钮不可点击", payload_text)
        self.assertIn("端：** H5", payload_text)
        self.assertNotIn("模块：", payload_text)

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_create_bug_without_project_group_config_does_not_send_group_message(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/new",
            data={
                "title": "未配置项目群通知的Bug",
                "version": "2.9.0",
                "module": "APP",
                "platform": "Android",
                "severity": "中",
                "assignee_id": "2",
                "environment": "测试环境",
                "description": "普通问题",
                "expected_result": "功能正常",
                "actual_result": "展示异常",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 0)
        self.assertNotIn("新建Bug群通知未发送".encode("utf-8"), response.data)

    @mock.patch("app.smtplib.SMTP", FakeSMTP)
    def test_mail_send_test_is_cancelled(self) -> None:
        self.enable_mail_notifications()
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/admin",
            data={
                "entity": "mail",
                "action": "send_test",
                "next": "/admin/mail",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("邮件发送已取消".encode("utf-8"), response.data)
        self.assertEqual(len(FakeSMTP.sent_messages), 0)

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_manual_group_report_send_includes_manual_note(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        self.login_as("admin", "admin123")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.executemany(
                """
                INSERT INTO bugs (
                    bug_no, title, project_id, version, module, platform, severity, priority, status,
                    assignee_id, creator_id, previous_assignee_id, reporter, environment, description,
                    expected_result, actual_result, resolution_note, created_at, updated_at
                )
                VALUES (?, ?, 1, '9.9.0', ?, ?, '高', '高', ?, 2, 1, 2, '李婷', '测试环境',
                        '报告统计用缺陷', '符合预期', '实际异常', '', '2026-07-14 10:00:00', '2026-07-14 10:00:00')
                """,
                [
                    ("901", "Android 未关闭 Bug", "APP", "Android", "open"),
                    ("902", "IOS 未关闭 Bug", "APP", "iOS", "in_progress"),
                    ("903", "H5 待验证 Bug", "H5", "H5", "pending_verification"),
                    ("904", "双端已关闭 Bug", "APP", "双端", "closed"),
                ],
            )
            conn.execute("UPDATE bugs SET severity = '最高' WHERE bug_no = '901'")
            conn.commit()
        save_response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "update",
                "next": "/admin/report-notify",
                "enabled": "1",
                "send_time": "18:00",
                "message_format": "card",
                "webhook_url": "https://open.feishu.cn/open-apis/bot/v2/hook/mock-webhook",
                "secret": "",
                "project_id": "1",
                "version": "9.9.0",
                "base_url": "http://127.0.0.1:5050",
                "manual_note": "今天重点关注支付回归\n客户端提测较晚，晚点补最终结论",
                "tracking_progress": "60%",
            },
            follow_redirects=True,
        )
        self.assertEqual(save_response.status_code, 200)
        self.assertIn("群测试报告通知设置已保存".encode("utf-8"), save_response.data)
        self.assertIn("今天重点关注支付回归".encode("utf-8"), save_response.data)
        self.assertIn("60%".encode("utf-8"), save_response.data)

        response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "send_test",
                "next": "/admin/report-notify",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("测试发送成功".encode("utf-8"), response.data)
        self.assertIn("已附带补充信息".encode("utf-8"), response.data)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 1)

        payload = json.loads(FakeGroupReportResponse.captured_requests[0].data.decode("utf-8"))
        payload_text = json.dumps(payload, ensure_ascii=False)
        self.assertEqual(payload["msg_type"], "interactive")
        self.assertEqual(payload["card"]["schema"], "2.0")
        self.assertNotIn("width_mode", payload["card"]["config"])
        self.assertEqual(payload["card"]["config"]["summary"]["content"], "QA 测试日报")
        self.assertNotIn("fallback", payload["card"])
        self.assertNotIn("rgba(", payload_text)
        self.assertIn("indigo-900", payload_text)
        self.assertIn("turquoise-50", payload_text)
        self.assertIn("interactive_container", payload_text)
        self.assertIn('"tag": "column_set"', payload_text)
        self.assertIn('"tag": "column"', payload_text)
        self.assertIn('"flex_mode": "flow"', payload_text)
        self.assertIn('"flex_mode": "bisect"', payload_text)
        self.assertIn("12px 18px 12px 18px", payload_text)
        self.assertIn("10px 14px 10px 14px", payload_text)
        self.assertIn("6px 16px 0px 16px", payload_text)
        self.assertNotIn('"padding": "20px"', payload_text)
        self.assertNotIn('"padding": "18px"', payload_text)
        self.assertNotIn("16px 24px 16px 24px", payload_text)
        self.assertNotIn("14px 24px 0px 24px", payload_text)
        self.assertIn("<text_tag color='green'>有风险</text_tag>", payload_text)
        self.assertNotIn("单列卡片", payload_text)
        self.assertIn("QA 测试日报", payload_text)
        self.assertIn("测试日报", payload_text)
        self.assertIn("零售增长平台 9.9.0", payload_text)
        self.assertIn("发布前收尾验证", payload_text)
        self.assertIn("今日结论", payload_text)
        self.assertIn("整体进度", payload_text)
        self.assertIn("用例执行", payload_text)
        self.assertIn("缺陷修复", payload_text)
        self.assertIn("打开缺陷", payload_text)
        self.assertNotIn("进入发版前收尾阶段", payload_text)
        self.assertNotIn("测试进度", payload_text)
        self.assertNotIn("用例执行进度与整体风险状态", payload_text)
        self.assertIn("缺陷闭环", payload_text)
        self.assertIn("发现 · 修复 · 回归 · 打开缺陷状态一屏看清", payload_text)
        self.assertIn("发现 4 → 修复 2 → 回归 1 → 打开 2", payload_text)
        self.assertIn("Android 1", payload_text)
        self.assertIn("IOS 1", payload_text)
        self.assertNotIn("H5 1", payload_text)
        self.assertNotIn("双端 1", payload_text)
        self.assertIn("严重 Bug / 风险备注", payload_text)
        self.assertIn("严重 Bug 汇总：1 个待处理", payload_text)
        self.assertIn("Android 未关闭 Bug", payload_text)
        self.assertIn("已打开 / 周越", payload_text)
        self.assertIn("风险备注", payload_text)
        self.assertNotIn("手动补充", payload_text)
        self.assertNotIn("暂无未关闭严重 Bug", payload_text)
        self.assertIn("建议日报结构：结论先行 / 关键指标 / 缺陷闭环 / 风险备注", payload_text)
        self.assertIn("QA Daily Report", payload_text)
        self.assertNotIn("明日/下一步", payload_text)
        self.assertNotIn("下一步", payload_text)
        self.assertNotIn("完成剩余用例", payload_text)
        self.assertIn("今天重点关注支付回归", payload_text)
        self.assertIn("客户端提测较晚，晚点补最终结论", payload_text)
        self.assertIn("埋点进度：60%", payload_text)

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_manual_group_report_can_send_image_message(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        self.login_as("admin", "admin123")
        save_response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "update",
                "next": "/admin/report-notify",
                "enabled": "1",
                "send_time": "18:00",
                "message_format": "image",
                "webhook_url": "https://open.feishu.cn/open-apis/bot/v2/hook/mock-webhook",
                "secret": "",
                "lark_app_id": "cli_mock",
                "lark_app_secret": "mock-secret",
                "project_id": "1",
                "version": "2.9.0",
                "base_url": "http://127.0.0.1:5050",
                "manual_note": "图片日报备注",
                "tracking_progress": "60%",
            },
            follow_redirects=True,
        )
        self.assertEqual(save_response.status_code, 200)

        response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "send_test",
                "next": "/admin/report-notify",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("图片消息测试报告".encode("utf-8"), response.data)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 3)

        token_request = FakeGroupReportResponse.captured_requests[0]
        token_payload = json.loads(token_request.data.decode("utf-8"))
        self.assertEqual(token_request.full_url, "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal")
        self.assertEqual(token_payload["app_id"], "cli_mock")
        self.assertEqual(token_payload["app_secret"], "mock-secret")

        upload_request = FakeGroupReportResponse.captured_requests[1]
        self.assertEqual(upload_request.full_url, "https://open.feishu.cn/open-apis/im/v1/images")
        self.assertIn(b'image_type"\r\n\r\nmessage', upload_request.data)
        self.assertIn(b"\x89PNG\r\n\x1a\n", upload_request.data)

        webhook_request = FakeGroupReportResponse.captured_requests[2]
        self.assertEqual(webhook_request.full_url, "https://open.feishu.cn/open-apis/bot/v2/hook/mock-webhook")
        payload = json.loads(webhook_request.data.decode("utf-8"))
        self.assertEqual(payload["msg_type"], "image")
        self.assertEqual(payload["content"]["image_key"], "img_mock_daily_report")

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_group_report_direct_image_requires_lark_app_credentials(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        self.login_as("admin", "admin123")
        save_response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "update",
                "next": "/admin/report-notify",
                "enabled": "1",
                "send_time": "18:00",
                "message_format": "image",
                "webhook_url": "https://open.feishu.cn/open-apis/bot/v2/hook/mock-webhook",
                "secret": "",
                "lark_app_id": "",
                "lark_app_secret": "",
                "project_id": "1",
                "version": "2.9.0",
                "base_url": "",
                "manual_note": "直接展示日报备注",
                "tracking_progress": "60%",
            },
            follow_redirects=True,
        )
        self.assertEqual(save_response.status_code, 200)
        self.assertIn("选择图片消息时，请填写飞书应用 App ID 和 App Secret。".encode("utf-8"), save_response.data)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 0)

    @mock.patch("app.urllib_request.urlopen", side_effect=fake_group_report_urlopen)
    def test_group_report_manual_note_without_severe_bug_skips_empty_risk_line(self, _mocked_urlopen) -> None:
        FakeGroupReportResponse.captured_requests = []
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "update",
                "next": "/admin/report-notify",
                "enabled": "1",
                "send_time": "18:00",
                "message_format": "card",
                "webhook_url": "https://open.feishu.cn/open-apis/bot/v2/hook/mock-webhook",
                "secret": "",
                "project_id": "1",
                "version": "2.9.0",
                "base_url": "http://127.0.0.1:5050",
                "manual_note": "123345",
                "tracking_progress": "",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/admin",
            data={
                "entity": "report_notify",
                "action": "send_test",
                "next": "/admin/report-notify",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(FakeGroupReportResponse.captured_requests), 1)

        payload = json.loads(FakeGroupReportResponse.captured_requests[0].data.decode("utf-8"))
        payload_text = json.dumps(payload, ensure_ascii=False)
        self.assertIn("严重 Bug 汇总：0 个待处理", payload_text)
        self.assertIn("风险备注", payload_text)
        self.assertNotIn("手动补充", payload_text)
        self.assertIn("123345", payload_text)
        self.assertNotIn("暂无未关闭严重 Bug", payload_text)

    def test_add_bug_comment_displays_in_bug_activity(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/1/comments",
            data={
                "content": "这里补充一条活动评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("评论已发布".encode("utf-8"), response.data)
        self.assertIn("这里补充一条活动评论".encode("utf-8"), response.data)
        self.assertIn("李婷".encode("utf-8"), response.data)

    def test_reply_bug_comment_displays_in_bug_comment_list(self) -> None:
        self.login_as("lit", "123456")
        first_response = self.client.post(
            "/bugs/1/comments",
            data={
                "content": "第一条评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        self.assertEqual(first_response.status_code, 200)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            parent_comment = conn.execute(
                "SELECT id FROM bug_comments WHERE bug_id = ? ORDER BY id DESC LIMIT 1",
                (1,),
            ).fetchone()

        reply_response = self.client.post(
            "/bugs/1/comments",
            data={
                "parent_id": str(parent_comment["id"]),
                "content": "这是回复内容",
                "redirect_to": f"/bugs/1?tab=detail#comment-{parent_comment['id']}",
            },
            follow_redirects=True,
        )
        self.assertEqual(reply_response.status_code, 200)
        self.assertIn("回复已发布".encode("utf-8"), reply_response.data)
        self.assertIn("这是回复内容".encode("utf-8"), reply_response.data)
        self.assertIn("回复 李婷".encode("utf-8"), reply_response.data)

    def test_bug_comments_show_newest_first(self) -> None:
        self.login_as("lit", "123456")
        self.client.post(
            "/bugs/1/comments",
            data={
                "content": "较早的评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        response = self.client.post(
            "/bugs/1/comments",
            data={
                "content": "最新的评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        body = response.data.decode("utf-8")
        self.assertLess(body.index("最新的评论"), body.index("较早的评论"))

    def test_delete_bug_comment_removes_comment_and_replies(self) -> None:
        self.login_as("lit", "123456")
        self.client.post(
            "/bugs/1/comments",
            data={
                "content": "待删除主评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            parent_comment = conn.execute(
                "SELECT id FROM bug_comments WHERE bug_id = ? ORDER BY id DESC LIMIT 1",
                (1,),
            ).fetchone()

        self.client.post(
            "/bugs/1/comments",
            data={
                "parent_id": str(parent_comment["id"]),
                "content": "待删除回复",
                "redirect_to": f"/bugs/1?tab=detail#comment-{parent_comment['id']}",
            },
            follow_redirects=True,
        )

        response = self.client.post(
            f"/bugs/1/comments/{parent_comment['id']}/delete",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("评论及回复已删除".encode("utf-8"), response.data)
        self.assertNotIn("待删除主评论".encode("utf-8"), response.data)
        self.assertNotIn("待删除回复".encode("utf-8"), response.data)

    def test_bug_comment_mention_creates_unread_notification(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.post(
            "/bugs/1/comments",
            data={
                "content": "请 @周越 看一下这个问题",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            notification = conn.execute(
                """
                SELECT notifications.*, users.name AS recipient_name
                FROM notifications
                JOIN users ON notifications.user_id = users.id
                WHERE users.username = ? AND notifications.category = ?
                """,
                ("zhouyue", "comment_mention"),
            ).fetchone()

        self.assertIsNotNone(notification)
        self.assertEqual(notification["is_read"], 0)
        self.assertIsNotNone(notification["comment_id"])
        self.assertIn("提到了你", notification["title"])
        self.assertIn("请 @周越 看一下这个问题", notification["body"])

        self.login_as("zhouyue", "123456")
        page = self.client.get("/notifications")
        self.assertEqual(page.status_code, 200)
        self.assertIn("消息中心".encode("utf-8"), page.data)
        self.assertIn("@提及".encode("utf-8"), page.data)
        self.assertIn("请 @周越 看一下这个问题".encode("utf-8"), page.data)

    def test_bug_comment_mention_picker_includes_all_system_users(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.get("/bugs/1?tab=detail")

        self.assertEqual(response.status_code, 200)
        self.assertIn("data-mention-picker".encode("utf-8"), response.data)
        self.assertIn("data-mention-option".encode("utf-8"), response.data)
        self.assertIn('data-mention-insert="Admin"'.encode("utf-8"), response.data)
        self.assertIn("提示所有成员".encode("utf-8"), response.data)

    def test_mentioned_name_text_changes_after_recipient_reads(self) -> None:
        self.login_as("lit", "123456")
        self.client.post(
            "/bugs/1/comments",
            data={
                "content": "@周越 帮忙确认",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )

        unread_page = self.client.get("/bugs/1?tab=detail")
        self.assertEqual(unread_page.status_code, 200)
        self.assertIn('<p class="bug-comment-content"><span class="mention-text'.encode("utf-8"), unread_page.data)
        self.assertIn("mention-text is-unread".encode("utf-8"), unread_page.data)
        self.assertIn('data-mention-read-state="unread"'.encode("utf-8"), unread_page.data)
        self.assertNotIn("mention-status-dot".encode("utf-8"), unread_page.data)

        self.login_as("zhouyue", "123456")
        read_response = self.client.post("/bugs/1/comments/read")
        self.assertEqual(read_response.status_code, 200)
        self.assertEqual(read_response.json["ok"], True)

        read_page = self.client.get("/bugs/1?tab=detail")
        self.assertEqual(read_page.status_code, 200)
        self.assertIn("mention-text is-read".encode("utf-8"), read_page.data)
        self.assertIn('data-mention-read-state="read"'.encode("utf-8"), read_page.data)
        self.assertNotIn("mention-status-dot".encode("utf-8"), read_page.data)

    def test_comment_on_created_bug_notifies_creator(self) -> None:
        self.login_as("zhouyue", "123456")
        response = self.client.post(
            "/bugs/1/comments",
            data={
                "content": "我补充一个处理进展",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            notification = conn.execute(
                """
                SELECT notifications.*, users.name AS recipient_name
                FROM notifications
                JOIN users ON notifications.user_id = users.id
                WHERE users.username = ? AND notifications.category = ?
                """,
                ("lit", "bug_comment"),
            ).fetchone()

        self.assertIsNotNone(notification)
        self.assertEqual(notification["is_read"], 0)
        self.assertIn("评论了你创建的 Bug", notification["title"])
        self.assertIn("我补充一个处理进展", notification["body"])

    def test_comment_unread_dot_disappears_after_mark_read(self) -> None:
        self.login_as("zhouyue", "123456")
        self.client.post(
            "/bugs/1/comments",
            data={
                "content": "这里有一条给创建人的未读评论",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )

        self.login_as("lit", "123456")
        detail = self.client.get("/bugs/1?tab=detail")
        self.assertEqual(detail.status_code, 200)
        self.assertIn('id="bug-comments" data-comment-unread-root'.encode("utf-8"), detail.data)

        read_response = self.client.post("/bugs/1/comments/read")
        self.assertEqual(read_response.status_code, 200)
        self.assertEqual(read_response.json["ok"], True)
        self.assertEqual(read_response.json["marked_count"], 1)
        self.assertEqual(read_response.json["unread_count"], 0)

        refreshed = self.client.get("/bugs/1?tab=detail")
        self.assertEqual(refreshed.status_code, 200)
        self.assertNotIn('id="bug-comments" data-comment-unread-root'.encode("utf-8"), refreshed.data)

    def test_open_notification_marks_it_read_and_redirects_to_comment(self) -> None:
        self.login_as("zhouyue", "123456")
        self.client.post(
            "/bugs/1/comments",
            data={
                "content": "打开消息后应当变成已读",
                "redirect_to": "/bugs/1?tab=detail#bug-comments",
            },
            follow_redirects=True,
        )

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            notification = conn.execute(
                """
                SELECT notifications.*
                FROM notifications
                JOIN users ON notifications.user_id = users.id
                WHERE users.username = ?
                ORDER BY notifications.id DESC
                LIMIT 1
                """,
                ("lit",),
            ).fetchone()

        self.assertIsNotNone(notification)
        self.login_as("lit", "123456")
        response = self.client.get(f"/notifications/{notification['id']}/open", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/bugs/1?tab=detail#comment-", response.headers["Location"])

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            updated = conn.execute("SELECT is_read FROM notifications WHERE id = ?", (notification["id"],)).fetchone()

        self.assertEqual(updated["is_read"], 1)

    def test_resolve_flow_moves_bug_back_to_creator_todo(self) -> None:
        self.login_as("zhouyue", "123456")
        response = self.client.post(
            "/bugs/1/update",
            data={"action": "resolve", "resolution_note": "修复完成"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("待验证".encode("utf-8"), response.data)
        self.assertIn("李婷".encode("utf-8"), response.data)

    def test_change_status_redirects_back_to_current_page(self) -> None:
        self.login_as("zhouyue", "123456")

        response = self.client.post(
            "/bugs/1/update",
            data={
                "action": "change_status",
                "status": "in_progress",
                "redirect_to": "/bugs/1?tab=process",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs/1?tab=process")

    def test_change_status_to_closed_redirects_to_bug_list(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.post(
            "/bugs/1/update",
            data={
                "action": "change_status",
                "status": "closed",
                "redirect_to": "/bugs/1?tab=detail",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs")

    def test_close_action_redirects_to_bug_list(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.post(
            "/bugs/1/update",
            data={
                "action": "close",
                "resolution_note": "验证通过",
                "redirect_to": "/bugs/1?tab=process",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs")

    def test_bug_update_ignores_external_redirect_target(self) -> None:
        self.login_as("zhouyue", "123456")

        response = self.client.post(
            "/bugs/1/update",
            data={
                "action": "change_status",
                "status": "in_progress",
                "redirect_to": "https://example.com/elsewhere",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs/1")

    def test_bug_detail_rejects_protocol_relative_back_url(self) -> None:
        self.login_as("zhouyue", "123456")

        response = self.client.get("/bugs/1?next=//third-party.example/path")

        self.assertEqual(response.status_code, 200)
        self.assertIn('href="/bugs"'.encode("utf-8"), response.data)
        self.assertNotIn("third-party.example".encode("utf-8"), response.data)
        self.assertNotIn("next=//third-party.example".encode("utf-8"), response.data)

    def test_bug_update_rejects_protocol_relative_redirect_target(self) -> None:
        self.login_as("zhouyue", "123456")

        response = self.client.post(
            "/bugs/1/update",
            data={
                "action": "change_status",
                "status": "in_progress",
                "redirect_to": "//third-party.example/elsewhere",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/bugs/1")

    def test_editing_pending_verification_bug_keeps_return_handler_for_reject(self) -> None:
        self.login_as("zhouyue", "123456")
        resolve_response = self.client.post(
            "/bugs/1/update",
            data={"action": "resolve", "resolution_note": "修复完成"},
            follow_redirects=True,
        )
        self.assertEqual(resolve_response.status_code, 200)

        self.login_as("lit", "123456")
        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            bug = conn.execute(
                """
                SELECT id, title, version, module, platform, severity, assignee_id, requirement_id, case_id,
                       environment, description, expected_result, actual_result, previous_assignee_id
                FROM bugs
                WHERE id = 1
                """
            ).fetchone()

        edit_response = self.client.post(
            "/bugs/1/edit",
            data={
                "title": bug["title"],
                "version": bug["version"],
                "module": bug["module"],
                "platform": bug["platform"],
                "severity": "高" if bug["severity"] != "高" else "中",
                "assignee_id": str(bug["assignee_id"]),
                "requirement_id": str(bug["requirement_id"] or ""),
                "case_id": str(bug["case_id"] or ""),
                "environment": bug["environment"] or "",
                "description": bug["description"],
                "expected_result": bug["expected_result"] or "",
                "actual_result": bug["actual_result"] or "",
            },
            follow_redirects=True,
        )
        self.assertEqual(edit_response.status_code, 200)
        self.assertIn("Bug 信息已更新".encode("utf-8"), edit_response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            edited_bug = conn.execute(
                "SELECT assignee_id, previous_assignee_id, status FROM bugs WHERE id = 1"
            ).fetchone()

        self.assertEqual(edited_bug["status"], "pending_verification")
        self.assertEqual(edited_bug["assignee_id"], 1)
        self.assertEqual(edited_bug["previous_assignee_id"], 2)

        reject_response = self.client.post(
            "/bugs/1/update",
            data={"action": "reject", "resolution_note": "验证失败"},
            follow_redirects=True,
        )
        self.assertEqual(reject_response.status_code, 200)
        self.assertIn("处理中".encode("utf-8"), reject_response.data)
        self.assertIn("周越".encode("utf-8"), reject_response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            rejected_bug = conn.execute(
                "SELECT assignee_id, previous_assignee_id, status FROM bugs WHERE id = 1"
            ).fetchone()

        self.assertEqual(rejected_bug["status"], "in_progress")
        self.assertEqual(rejected_bug["assignee_id"], 2)

    def test_admin_page_requires_admin(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/admin", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("仅管理员可访问".encode("utf-8"), response.data)

    def test_profile_password_change_redirects_to_login(self) -> None:
        self.login_as("lit", "123456")

        response = self.client.post(
            "/profile",
            data={
                "name": "李婷",
                "email": "lit@alvinsclub.ai",
                "password": "new-password",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])

        with self.client.session_transaction() as session_state:
            self.assertNotIn("user_id", session_state)

        login_page = self.client.get(response.headers["Location"])
        self.assertEqual(login_page.status_code, 200)
        self.assertIn("密码已修改，请使用新密码重新登录。".encode("utf-8"), login_page.data)

        old_password_response = self.client.post(
            "/login",
            data={"username": "lit", "password": "123456"},
        )
        self.assertEqual(old_password_response.status_code, 200)
        self.assertIn("账号或密码错误。".encode("utf-8"), old_password_response.data)

        new_password_response = self.client.post(
            "/login",
            data={"username": "lit", "password": "new-password"},
            follow_redirects=True,
        )
        self.assertEqual(new_password_response.status_code, 200)
        self.assertIn("Bug列表".encode("utf-8"), new_password_response.data)

    def test_admin_page_loads_for_admin(self) -> None:
        self.login_as("admin", "admin123")
        response = self.client.get("/admin")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Admin 管理端".encode("utf-8"), response.data)

    def test_admin_can_create_admin_account(self) -> None:
        self.login_as("admin", "admin123")
        response = self.client.post(
            "/admin",
            data={
                "entity": "user",
                "action": "create",
                "next": "/admin/users",
                "name": "新管理员",
                "account_type": "admin",
                "username": "newadmin",
                "email": "newadmin@alvinsclub.ai",
                "password": "123456",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("账号已创建".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT role, role_code FROM users WHERE username = ?", ("newadmin",)).fetchone()
        self.assertEqual(row["role"], "系统管理员")
        self.assertEqual(row["role_code"], "admin")

    def test_admin_can_update_member_to_admin(self) -> None:
        self.login_as("admin", "admin123")
        self.client.post(
            "/admin",
            data={
                "entity": "user",
                "action": "create",
                "next": "/admin/users",
                "name": "待升级成员",
                "account_type": "member",
                "role_code": "tester_engineer",
                "username": "membertoadmin",
                "email": "membertoadmin@alvinsclub.ai",
                "password": "123456",
            },
            follow_redirects=True,
        )

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            created_user = conn.execute("SELECT id FROM users WHERE username = ?", ("membertoadmin",)).fetchone()

        response = self.client.post(
            "/admin",
            data={
                "entity": "user",
                "action": "update",
                "user_id": str(created_user["id"]),
                "next": f"/admin/users/{created_user['id']}",
                "name": "待升级成员",
                "account_type": "admin",
                "username": "membertoadmin",
                "email": "membertoadmin@alvinsclub.ai",
                "password": "",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("账号已更新".encode("utf-8"), response.data)

        with sqlite3.connect(self.app.config["DATABASE"]) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT role, role_code FROM users WHERE username = ?", ("membertoadmin",)).fetchone()
        self.assertEqual(row["role"], "系统管理员")
        self.assertEqual(row["role_code"], "admin")

    def test_report_export(self) -> None:
        self.login_as("lit", "123456")
        response = self.client.get("/reports/testing/export")
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response.headers.get("Content-Disposition", ""))


if __name__ == "__main__":
    unittest.main()
