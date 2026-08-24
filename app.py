from __future__ import annotations

import io
import math
import os
import re
import sqlite3
import smtplib
import ssl
import threading
import time
import uuid
import base64
import hashlib
import hmac
import json
from contextlib import closing
from datetime import datetime
from email.message import EmailMessage
from html import escape
from pathlib import Path
from typing import Any
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request
from zoneinfo import ZoneInfo

os.environ.setdefault("MPLCONFIGDIR", "/private/tmp/mpl")
os.environ.setdefault("XDG_CACHE_HOME", "/private/tmp")

import openpyxl
from flask import Flask, Response, flash, g, get_flashed_messages, jsonify, redirect, render_template, request, send_file, session, url_for
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DB_PATH = BASE_DIR / "data" / "bug_platform.db"
DEFAULT_UPLOAD_FOLDER = BASE_DIR / "uploads"
APP_TIME_ZONE = ZoneInfo("Asia/Shanghai")
PAGE_SIZE = 10
BUG_PAGE_SIZE = 20
CASE_PAGE_SIZE = 12

STATUS_OPTIONS = [
    ("open", "已打开"),
    ("in_progress", "处理中"),
    ("pending_verification", "待验证"),
    ("closed", "已关闭"),
    ("duplicate", "重复"),
    ("on_hold", "搁置"),
]
STATUS_LABELS = dict(STATUS_OPTIONS)
TODO_STATUS_CODES = ("open", "in_progress", "pending_verification")
BUG_SEVERITY_OPTIONS = ["最高", "高", "中", "低", "最低", "建议"]
MAIL_NOTIFY_SEVERITY = "最高"
BUG_PLATFORM_OPTIONS = ["Android", "iOS", "双端", "H5", "WEB", "后端", "AI"]
BUG_NOTIFY_RULE_OPTIONS = [
    {"key": "APP", "label": "APP（Android / iOS / 双端）"},
    {"key": "H5", "label": "H5"},
    {"key": "WEB", "label": "WEB"},
    {"key": "Backend", "label": "后端"},
    {"key": "AI", "label": "AI"},
]
BUG_PLATFORM_NOTIFY_KEY_MAP = {
    "Android": "APP",
    "iOS": "APP",
    "双端": "APP",
    "H5": "H5",
    "WEB": "WEB",
    "后端": "Backend",
    "AI": "AI",
}
BUG_ATTACHMENT_SOURCE_FIELDS = {
    "title",
    "version",
    "environment",
    "description",
    "expected_result",
    "actual_result",
    "attachments",
}
BUG_INLINE_ATTACHMENT_FIELDS = (
    "title",
    "version",
    "environment",
    "description",
    "expected_result",
    "actual_result",
)
REPORT_PLATFORM_LABELS = {
    "iOS": "IOS",
}
BUG_PRIORITY_OPTIONS = BUG_SEVERITY_OPTIONS.copy()
BUG_MULTI_FILTER_KEYS = ("version", "platform", "creator_id", "assignee_id", "status")
COMMENT_NOTIFICATION_CATEGORIES = ("bug_comment", "comment_mention")
NOTIFICATION_CATEGORY_LABELS = {
    "bug_comment": "评论",
    "comment_mention": "@提及",
    "severe_bug": "严重Bug",
}
BUG_PRIORITY_ICON_MAP = {
    "最高": "highest",
    "高": "high",
    "中": "medium",
    "低": "low",
    "最低": "lowest",
    "建议": "suggestion",
}
BUG_SEVERITY_FALLBACK_MAP = {
    "严重": "最高",
    "一般": "中",
    "建议": "建议",
}


def app_now() -> datetime:
    return datetime.now(APP_TIME_ZONE)


def bug_notify_key_for_platform(platform: object) -> str:
    platform_text = str(platform or "").strip()
    return BUG_PLATFORM_NOTIFY_KEY_MAP.get(platform_text, platform_text)


def bug_notify_label_for_key(rule_key: object) -> str:
    rule_key_text = str(rule_key or "").strip()
    for rule_option in BUG_NOTIFY_RULE_OPTIONS:
        if rule_option["key"] == rule_key_text:
            return str(rule_option["label"])
    return rule_key_text


REQUIREMENT_STATUS_OPTIONS = [
    ("pending", "待评估"),
    ("in_progress", "进行中"),
    ("completed", "已完成"),
    ("on_hold", "已搁置"),
]
REQUIREMENT_STATUS_LABELS = dict(REQUIREMENT_STATUS_OPTIONS)
MAIL_SECURITY_OPTIONS = [
    ("tls", "STARTTLS"),
    ("ssl", "SSL/TLS"),
    ("none", "无加密"),
]
DEFAULT_MAIL_SETTINGS = {
    "enabled": False,
    "host": "",
    "port": "587",
    "security": "tls",
    "username": "",
    "password": "",
    "from_email": "",
    "sender_name": "Alvin's Club Bug Management Platform",
    "send_time": "10:05",
    "last_sent_at": "",
    "last_sent_date": "",
    "last_result": "",
}
DEFAULT_GROUP_REPORT_SETTINGS = {
    "enabled": False,
    "webhook_url": "",
    "secret": "",
    "send_time": "18:00",
    "project_id": "",
    "version": "",
    "base_url": "",
    "manual_note": "",
    "tracking_progress": "",
    "message_format": "card",
    "lark_app_id": "",
    "lark_app_secret": "",
    "last_sent_at": "",
    "last_sent_date": "",
    "last_result": "",
}
GROUP_REPORT_MESSAGE_FORMATS = {
    "image": "飞书图片消息",
    "card": "自适应卡片消息",
}
CASE_STATUS_OPTIONS = ["未测", "通过", "失败", "受阻", "跳过"]
CASE_STATUS_COLORS = {
    "未测": "#d9d9d9",
    "通过": "#7bc67e",
    "失败": "#ef6f6c",
    "受阻": "#f6c85f",
    "跳过": "#68b5e8",
}
CASE_STATUS_CHART_LABELS = {
    "未测": "Not Run",
    "通过": "Pass",
    "失败": "Fail",
    "受阻": "Blocked",
    "跳过": "Skip",
}
PLATFORM_RESULT_OPTIONS = ["", "pass", "failed", "block", "skip"]

SVG_CHART_WIDTH = 680
SVG_CHART_HEIGHT = 320

SAMPLE_USERS = [
    ("李婷", "测试负责人", "tester"),
    ("周越", "APP 开发", "developer"),
    ("王昊", "WEB 开发", "developer"),
    ("赵航", "后端开发", "developer"),
    ("沈意", "AI 工程师", "developer"),
    ("陈默", "产品经理", "pm"),
    ("Admin", "系统管理员", "admin"),
]

ADMIN_ROLE_CODE = "admin"
ADMIN_ROLE_LABEL = "系统管理员"

ROLE_OPTIONS = [
    ("tester_engineer", "测试工程师"),
    ("app_developer", "APP开发"),
    ("h5_developer", "H5开发"),
    ("backend_developer", "后端开发"),
    ("ai_developer", "AI开发"),
    ("product", "产品"),
    ("designer", "设计"),
]

ROLE_LABELS = dict(ROLE_OPTIONS)
ACCOUNT_TYPE_OPTIONS = [
    ("member", "普通成员"),
    ("admin", "管理员"),
]
ACCOUNT_TYPE_LABELS = dict(ACCOUNT_TYPE_OPTIONS)

SAMPLE_CREDENTIALS = {
    "lit": "123456",
    "zhouyue": "123456",
    "wanghao": "123456",
    "zhaohang": "123456",
    "shenyi": "123456",
    "chenmo": "123456",
    "admin": "admin123",
}

SAMPLE_USER_PROFILES = [
    {
        "name": name,
        "role": role,
        "role_code": role_code,
        "account_type": "admin" if role_code == ADMIN_ROLE_CODE else "member",
        "username": username,
        "password": SAMPLE_CREDENTIALS[username],
        "email": f"{username}@alvinsclub.ai",
    }
    for (name, role, role_code), username in zip(SAMPLE_USERS, SAMPLE_CREDENTIALS.keys())
]

SAMPLE_PROJECTS = [
    ("零售增长平台", "增长业务相关项目"),
    ("智能客服系统", "AI 服务与管理后台"),
    ("商家工作台", "商家端 APP / WEB / API"),
]

SAMPLE_REQUIREMENTS = [
    ("商家工作台", "REQ-001", "首页升级改版"),
    ("商家工作台", "REQ-002", "退款链路容错优化"),
    ("智能客服系统", "REQ-003", "知识库检索准确率提升"),
]

SAMPLE_CASES = [
    {
        "project_name": "商家工作台",
        "folder_name": "测试用例",
        "doc_name": "2.6.0-首页优化测试用例",
        "case_no": "2.6.0-HOME-TC-001",
        "title": "首页 banner 点击跳转",
        "priority_level": "P0",
        "module_name": "首页推荐",
        "steps": "1. 打开首页\n2. 点击 banner 卡片",
        "expected_result": "正常跳转到活动页，页面数据完整展示。",
        "execute_status": "通过",
        "source_type": "在线文档",
        "ios_result": "pass",
        "android_result": "",
        "h5_result": "",
        "remark": "",
    },
    {
        "project_name": "商家工作台",
        "folder_name": "测试用例",
        "doc_name": "2.6.0-首页优化测试用例",
        "case_no": "2.6.0-HOME-TC-002",
        "title": "首页固定坑位顺序校验",
        "priority_level": "P0",
        "module_name": "首页推荐",
        "steps": "1. 打开首页\n2. 查看固定坑位顺序",
        "expected_result": "固定坑位顺序正确且位置稳定。",
        "execute_status": "失败",
        "source_type": "在线文档",
        "ios_result": "failed",
        "android_result": "",
        "h5_result": "",
        "remark": "iOS 展示顺序错误",
    },
    {
        "project_name": "商家工作台",
        "folder_name": "测试用例",
        "doc_name": "2.6.0-首页优化测试用例",
        "case_no": "2.6.0-HOME-TC-003",
        "title": "星模横滑推荐内容过滤",
        "priority_level": "P1",
        "module_name": "首页推荐",
        "steps": "1. 使用新账号进入首页\n2. 查看横滑推荐内容",
        "expected_result": "已关注星模和品牌内容不再重复推荐。",
        "execute_status": "受阻",
        "source_type": "在线文档",
        "ios_result": "block",
        "android_result": "",
        "h5_result": "",
        "remark": "测试账号画像未准备完成",
    },
    {
        "project_name": "零售增长平台",
        "folder_name": "测试组",
        "doc_name": "2.4.0_功能测试用例",
        "case_no": "2.4.0-ORDER-TC-001",
        "title": "订单列表按时间筛选",
        "priority_level": "P1",
        "module_name": "订单列表",
        "steps": "1. 进入订单列表\n2. 选择开始和结束日期",
        "expected_result": "列表仅展示时间范围内数据。",
        "execute_status": "跳过",
        "source_type": "在线文档",
        "ios_result": "skip",
        "android_result": "",
        "h5_result": "",
        "remark": "需求排期顺延",
    },
    {
        "project_name": "智能客服系统",
        "folder_name": "共享文档",
        "doc_name": "2.5.0_OOTD_AI试穿_测试用例",
        "case_no": "2.5.0-AI-TC-001",
        "title": "知识库召回最新政策",
        "priority_level": "P0",
        "module_name": "知识库检索",
        "steps": "1. 提问配送政策\n2. 查看命中答案",
        "expected_result": "优先命中最新知识库切片内容。",
        "execute_status": "通过",
        "source_type": "在线文档",
        "ios_result": "pass",
        "android_result": "pass",
        "h5_result": "",
        "remark": "",
    },
]

SAMPLE_BUGS = [
    {
        "title": "首页帖子无法 tryon",
        "project_name": "商家工作台",
        "version": "2.6.0",
        "module": "APP",
        "platform": "双端",
        "severity": "高",
        "priority": "高",
        "status": "open",
        "creator_name": "李婷",
        "assignee_name": "周越",
        "previous_assignee_name": "周越",
        "environment": "iOS 18.1 / 测试环境",
        "description": "进入首页后点击帖子 tryon 入口，无法进入试穿流程。",
        "expected_result": "应当正常进入试穿页面。",
        "actual_result": "点击无反应。",
        "resolution_note": "",
        "requirement_code": "REQ-001",
        "case_code": "TC-001",
    },
    {
        "title": "AI 问答推荐结果与知识库不一致",
        "project_name": "智能客服系统",
        "version": "2.5.0",
        "module": "AI",
        "platform": "AI",
        "severity": "低",
        "priority": "低",
        "status": "pending_verification",
        "creator_name": "李婷",
        "assignee_name": "沈意",
        "previous_assignee_name": "沈意",
        "environment": "RAG v2 / 测试知识库",
        "description": "用户问配送时效时，答案引用了过期政策。",
        "expected_result": "优先返回最新知识库答案。",
        "actual_result": "命中了过期切片内容。",
        "resolution_note": "已更新召回规则。",
        "requirement_code": "REQ-003",
        "case_code": "TC-005",
    },
    {
        "title": "后端退款接口偶发 500",
        "project_name": "商家工作台",
        "version": "2.6.0",
        "module": "Backend",
        "platform": "后端",
        "severity": "高",
        "priority": "高",
        "status": "pending_verification",
        "creator_name": "李婷",
        "assignee_name": "李婷",
        "previous_assignee_name": "赵航",
        "environment": "压测环境",
        "description": "批量退款时偶发空指针。",
        "expected_result": "接口稳定返回结果。",
        "actual_result": "部分请求返回 500。",
        "resolution_note": "后端已修复，待验证。",
        "requirement_code": "REQ-002",
        "case_code": "TC-002",
    },
]


def create_app(test_config: dict | None = None) -> Flask:
    app = Flask(__name__)
    # Let Flask honor nginx subpath prefixes such as X-Forwarded-Prefix.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_prefix=1)
    app.config.from_mapping(
        SECRET_KEY=os.environ.get("SECRET_KEY", "dev-change-me-before-deploy"),
        DATABASE=os.environ.get("DATABASE", str(DEFAULT_DB_PATH)),
        UPLOAD_FOLDER=os.environ.get("UPLOAD_FOLDER", str(DEFAULT_UPLOAD_FOLDER)),
        START_SCHEDULER=os.environ.get("START_SCHEDULER", "1").strip().lower() not in {"0", "false", "no", "off"},
        PAGE_SIZE=PAGE_SIZE,
        BUG_PAGE_SIZE=BUG_PAGE_SIZE,
        CASE_PAGE_SIZE=CASE_PAGE_SIZE,
    )

    if test_config:
        app.config.update(test_config)
    if app.config.get("TESTING") and (not test_config or "START_SCHEDULER" not in test_config):
        app.config["START_SCHEDULER"] = False

    Path(app.config["DATABASE"]).parent.mkdir(parents=True, exist_ok=True)
    Path(app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)

    def current_time() -> str:
        return app_now().strftime("%Y-%m-%d %H:%M:%S")

    def normalize_bug_severity_value(severity: object, priority: object = "") -> str:
        severity_text = str(severity or "").strip()
        priority_text = str(priority or "").strip()
        if severity_text in BUG_SEVERITY_OPTIONS:
            return severity_text
        if priority_text in BUG_SEVERITY_OPTIONS:
            return priority_text
        if severity_text in BUG_SEVERITY_FALLBACK_MAP:
            return BUG_SEVERITY_FALLBACK_MAP[severity_text]
        if priority_text in BUG_SEVERITY_FALLBACK_MAP:
            return BUG_SEVERITY_FALLBACK_MAP[priority_text]
        return severity_text or priority_text or "高"

    def get_bug_sync_token() -> str:
        token = str(session.get("bug_sync_token", "") or "").strip()
        if not token:
            token = str(time.time_ns())
            session["bug_sync_token"] = token
        return token

    def bump_bug_sync_token() -> str:
        token = str(time.time_ns())
        session["bug_sync_token"] = token
        return token

    def format_bug_no(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        if text.startswith("BUG-"):
            return text
        if text.isdigit():
            return text.zfill(3)
        return text

    def get_db() -> sqlite3.Connection:
        if "db" not in g:
            g.db = sqlite3.connect(app.config["DATABASE"])
            g.db.row_factory = sqlite3.Row
        return g.db

    def build_asset_version() -> str:
        watched_files = [BASE_DIR / "app.py"]
        watched_files.extend(path for path in (BASE_DIR / "static").rglob("*") if path.is_file())
        watched_files.extend(path for path in (BASE_DIR / "templates").rglob("*") if path.is_file())
        latest_mtime = max(int(path.stat().st_mtime) for path in watched_files)
        return str(latest_mtime)

    @app.context_processor
    def inject_asset_version() -> dict[str, str]:
        return {
            "asset_version": build_asset_version(),
            "bug_sync_token": get_bug_sync_token(),
            "role_options": ROLE_OPTIONS,
            "role_labels": ROLE_LABELS,
            "account_type_options": ACCOUNT_TYPE_OPTIONS,
            "account_type_labels": ACCOUNT_TYPE_LABELS,
        }

    @app.after_request
    def add_no_cache_headers(response: Response) -> Response:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    def init_db() -> None:
        schema = """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            role TEXT NOT NULL,
            role_code TEXT,
            account_type TEXT NOT NULL DEFAULT 'member',
            username TEXT,
            password TEXT,
            email TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            description TEXT,
            bug_notify_enabled INTEGER NOT NULL DEFAULT 0,
            bug_notify_webhook TEXT,
            bug_notify_secret TEXT,
            bug_notify_base_url TEXT,
            bug_notify_last_sent_at TEXT,
            bug_notify_last_result TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS project_bug_notify_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            module TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 0,
            webhook_url TEXT,
            secret TEXT,
            last_sent_at TEXT,
            last_result TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(project_id, module)
        );

        CREATE TABLE IF NOT EXISTS requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            code TEXT NOT NULL,
            title TEXT NOT NULL,
            version TEXT,
            status TEXT,
            priority TEXT,
            description TEXT,
            acceptance_criteria TEXT,
            requirement_doc_link TEXT,
            design_doc_link TEXT,
            creator_id INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS test_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            version TEXT,
            folder_name TEXT,
            doc_name TEXT,
            case_no TEXT NOT NULL,
            title TEXT NOT NULL,
            priority_level TEXT,
            module_name TEXT,
            steps TEXT,
            expected_result TEXT,
            actual_result TEXT,
            ios_result TEXT,
            android_result TEXT,
            h5_result TEXT,
            remark TEXT,
            executor TEXT,
            environment_info TEXT,
            device_info TEXT,
            network_info TEXT,
            source_type TEXT NOT NULL,
            doc_link TEXT,
            execute_status TEXT NOT NULL,
            creator_id INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS case_document_columns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            version TEXT,
            folder_name TEXT,
            doc_name TEXT,
            column_name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            creator_id INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS case_folders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            creator_id INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(project_id, name)
        );

        CREATE TABLE IF NOT EXISTS case_document_cells (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            column_id INTEGER NOT NULL,
            case_id INTEGER NOT NULL,
            cell_value TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_case_document_cells_unique
        ON case_document_cells(column_id, case_id);

        CREATE INDEX IF NOT EXISTS idx_case_document_columns_document
        ON case_document_columns(project_id, version, folder_name, doc_name, sort_order, id);

        CREATE INDEX IF NOT EXISTS idx_case_document_cells_case
        ON case_document_cells(case_id);

        CREATE TABLE IF NOT EXISTS bugs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bug_no TEXT,
            title TEXT NOT NULL,
            project_id INTEGER NOT NULL,
            version TEXT,
            module TEXT NOT NULL,
            platform TEXT,
            severity TEXT NOT NULL,
            priority TEXT NOT NULL,
            status TEXT NOT NULL,
            assignee_id INTEGER NOT NULL,
            creator_id INTEGER,
            previous_assignee_id INTEGER,
            reporter TEXT,
            requirement_id INTEGER,
            case_id INTEGER,
            environment TEXT,
            description TEXT NOT NULL,
            expected_result TEXT,
            actual_result TEXT,
            resolution_note TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS bug_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bug_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            detail TEXT NOT NULL,
            operator_name TEXT NOT NULL,
            environment_snapshot TEXT,
            status_snapshot TEXT,
            assignee_snapshot TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS bug_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bug_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            content_type TEXT,
            source_field TEXT NOT NULL DEFAULT 'attachments',
            file_path TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS bug_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bug_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            parent_id INTEGER,
            author_name TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            actor_id INTEGER,
            bug_id INTEGER,
            comment_id INTEGER,
            category TEXT NOT NULL,
            title TEXT NOT NULL,
            body TEXT NOT NULL,
            link_path TEXT,
            is_read INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            read_at TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_notifications_user_read
        ON notifications(user_id, is_read, created_at DESC, id DESC);

        CREATE INDEX IF NOT EXISTS idx_notifications_bug
        ON notifications(bug_id);

        CREATE TABLE IF NOT EXISTS mail_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            enabled INTEGER NOT NULL DEFAULT 0,
            host TEXT,
            port INTEGER,
            security TEXT,
            username TEXT,
            password TEXT,
            from_email TEXT,
            sender_name TEXT,
            send_time TEXT,
            last_sent_at TEXT,
            last_sent_date TEXT,
            last_result TEXT,
            report_notify_enabled INTEGER NOT NULL DEFAULT 0,
            report_notify_webhook TEXT,
            report_notify_secret TEXT,
            report_notify_send_time TEXT,
            report_notify_project_id INTEGER,
            report_notify_version TEXT,
            report_notify_base_url TEXT,
            report_notify_manual_note TEXT,
            report_notify_tracking_progress TEXT,
            report_notify_message_format TEXT,
            report_notify_lark_app_id TEXT,
            report_notify_lark_app_secret TEXT,
            report_notify_last_sent_at TEXT,
            report_notify_last_sent_date TEXT,
            report_notify_last_result TEXT
        );
        """

        with closing(sqlite3.connect(app.config["DATABASE"])) as db:
            db.executescript(schema)
            db.commit()

    def column_names(table_name: str) -> set[str]:
        rows = get_db().execute(f"PRAGMA table_info({table_name})").fetchall()
        return {row["name"] for row in rows}

    def run_migrations() -> None:
        db = get_db()
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS case_document_columns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                version TEXT,
                folder_name TEXT,
                doc_name TEXT,
                column_name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                creator_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS case_document_cells (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                column_id INTEGER NOT NULL,
                case_id INTEGER NOT NULL,
                cell_value TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_case_document_cells_unique
            ON case_document_cells(column_id, case_id);

            CREATE INDEX IF NOT EXISTS idx_case_document_columns_document
            ON case_document_columns(project_id, version, folder_name, doc_name, sort_order, id);

            CREATE INDEX IF NOT EXISTS idx_case_document_cells_case
            ON case_document_cells(case_id);

            CREATE TABLE IF NOT EXISTS case_folders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                creator_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(project_id, name)
            );

            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                actor_id INTEGER,
                bug_id INTEGER,
                comment_id INTEGER,
                category TEXT NOT NULL,
                title TEXT NOT NULL,
                body TEXT NOT NULL,
                link_path TEXT,
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                read_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_notifications_user_read
            ON notifications(user_id, is_read, created_at DESC, id DESC);

            CREATE INDEX IF NOT EXISTS idx_notifications_bug
            ON notifications(bug_id);

            CREATE TABLE IF NOT EXISTS project_bug_notify_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                module TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 0,
                webhook_url TEXT,
                secret TEXT,
                last_sent_at TEXT,
                last_result TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(project_id, module)
            );

            CREATE INDEX IF NOT EXISTS idx_project_bug_notify_rules_project
            ON project_bug_notify_rules(project_id);

            CREATE INDEX IF NOT EXISTS idx_case_folders_project_name
            ON case_folders(project_id, name);
            """
        )
        for table, column_sqls in {
            "projects": [
                ("bug_notify_enabled", "ALTER TABLE projects ADD COLUMN bug_notify_enabled INTEGER NOT NULL DEFAULT 0"),
                ("bug_notify_webhook", "ALTER TABLE projects ADD COLUMN bug_notify_webhook TEXT"),
                ("bug_notify_secret", "ALTER TABLE projects ADD COLUMN bug_notify_secret TEXT"),
                ("bug_notify_base_url", "ALTER TABLE projects ADD COLUMN bug_notify_base_url TEXT"),
                ("bug_notify_last_sent_at", "ALTER TABLE projects ADD COLUMN bug_notify_last_sent_at TEXT"),
                ("bug_notify_last_result", "ALTER TABLE projects ADD COLUMN bug_notify_last_result TEXT"),
            ],
            "users": [
                ("role_code", "ALTER TABLE users ADD COLUMN role_code TEXT"),
                ("account_type", "ALTER TABLE users ADD COLUMN account_type TEXT NOT NULL DEFAULT 'member'"),
                ("username", "ALTER TABLE users ADD COLUMN username TEXT"),
                ("password", "ALTER TABLE users ADD COLUMN password TEXT"),
                ("email", "ALTER TABLE users ADD COLUMN email TEXT"),
            ],
            "mail_settings": [
                ("enabled", "ALTER TABLE mail_settings ADD COLUMN enabled INTEGER NOT NULL DEFAULT 0"),
                ("host", "ALTER TABLE mail_settings ADD COLUMN host TEXT"),
                ("port", "ALTER TABLE mail_settings ADD COLUMN port INTEGER"),
                ("security", "ALTER TABLE mail_settings ADD COLUMN security TEXT"),
                ("username", "ALTER TABLE mail_settings ADD COLUMN username TEXT"),
                ("password", "ALTER TABLE mail_settings ADD COLUMN password TEXT"),
                ("from_email", "ALTER TABLE mail_settings ADD COLUMN from_email TEXT"),
                ("sender_name", "ALTER TABLE mail_settings ADD COLUMN sender_name TEXT"),
                ("send_time", "ALTER TABLE mail_settings ADD COLUMN send_time TEXT"),
                ("last_sent_at", "ALTER TABLE mail_settings ADD COLUMN last_sent_at TEXT"),
                ("last_sent_date", "ALTER TABLE mail_settings ADD COLUMN last_sent_date TEXT"),
                ("last_result", "ALTER TABLE mail_settings ADD COLUMN last_result TEXT"),
                ("report_notify_enabled", "ALTER TABLE mail_settings ADD COLUMN report_notify_enabled INTEGER NOT NULL DEFAULT 0"),
                ("report_notify_webhook", "ALTER TABLE mail_settings ADD COLUMN report_notify_webhook TEXT"),
                ("report_notify_secret", "ALTER TABLE mail_settings ADD COLUMN report_notify_secret TEXT"),
                ("report_notify_send_time", "ALTER TABLE mail_settings ADD COLUMN report_notify_send_time TEXT"),
                ("report_notify_project_id", "ALTER TABLE mail_settings ADD COLUMN report_notify_project_id INTEGER"),
                ("report_notify_version", "ALTER TABLE mail_settings ADD COLUMN report_notify_version TEXT"),
                ("report_notify_base_url", "ALTER TABLE mail_settings ADD COLUMN report_notify_base_url TEXT"),
                ("report_notify_manual_note", "ALTER TABLE mail_settings ADD COLUMN report_notify_manual_note TEXT"),
                ("report_notify_tracking_progress", "ALTER TABLE mail_settings ADD COLUMN report_notify_tracking_progress TEXT"),
                ("report_notify_message_format", "ALTER TABLE mail_settings ADD COLUMN report_notify_message_format TEXT"),
                ("report_notify_lark_app_id", "ALTER TABLE mail_settings ADD COLUMN report_notify_lark_app_id TEXT"),
                ("report_notify_lark_app_secret", "ALTER TABLE mail_settings ADD COLUMN report_notify_lark_app_secret TEXT"),
                ("report_notify_last_sent_at", "ALTER TABLE mail_settings ADD COLUMN report_notify_last_sent_at TEXT"),
                ("report_notify_last_sent_date", "ALTER TABLE mail_settings ADD COLUMN report_notify_last_sent_date TEXT"),
                ("report_notify_last_result", "ALTER TABLE mail_settings ADD COLUMN report_notify_last_result TEXT"),
            ],
            "test_cases": [
                ("version", "ALTER TABLE test_cases ADD COLUMN version TEXT"),
                ("folder_name", "ALTER TABLE test_cases ADD COLUMN folder_name TEXT"),
                ("doc_name", "ALTER TABLE test_cases ADD COLUMN doc_name TEXT"),
                ("priority_level", "ALTER TABLE test_cases ADD COLUMN priority_level TEXT"),
                ("module_name", "ALTER TABLE test_cases ADD COLUMN module_name TEXT"),
                ("steps", "ALTER TABLE test_cases ADD COLUMN steps TEXT"),
                ("expected_result", "ALTER TABLE test_cases ADD COLUMN expected_result TEXT"),
                ("actual_result", "ALTER TABLE test_cases ADD COLUMN actual_result TEXT"),
                ("ios_result", "ALTER TABLE test_cases ADD COLUMN ios_result TEXT"),
                ("android_result", "ALTER TABLE test_cases ADD COLUMN android_result TEXT"),
                ("h5_result", "ALTER TABLE test_cases ADD COLUMN h5_result TEXT"),
                ("remark", "ALTER TABLE test_cases ADD COLUMN remark TEXT"),
                ("executor", "ALTER TABLE test_cases ADD COLUMN executor TEXT"),
                ("environment_info", "ALTER TABLE test_cases ADD COLUMN environment_info TEXT"),
                ("device_info", "ALTER TABLE test_cases ADD COLUMN device_info TEXT"),
                ("network_info", "ALTER TABLE test_cases ADD COLUMN network_info TEXT"),
                ("creator_id", "ALTER TABLE test_cases ADD COLUMN creator_id INTEGER"),
            ],
            "bugs": [
                ("bug_no", "ALTER TABLE bugs ADD COLUMN bug_no TEXT"),
                ("version", "ALTER TABLE bugs ADD COLUMN version TEXT"),
                ("creator_id", "ALTER TABLE bugs ADD COLUMN creator_id INTEGER"),
                ("previous_assignee_id", "ALTER TABLE bugs ADD COLUMN previous_assignee_id INTEGER"),
                ("reporter", "ALTER TABLE bugs ADD COLUMN reporter TEXT"),
                ("requirement_id", "ALTER TABLE bugs ADD COLUMN requirement_id INTEGER"),
                ("case_id", "ALTER TABLE bugs ADD COLUMN case_id INTEGER"),
                ("platform", "ALTER TABLE bugs ADD COLUMN platform TEXT"),
            ],
            "bug_history": [
                ("environment_snapshot", "ALTER TABLE bug_history ADD COLUMN environment_snapshot TEXT"),
                ("status_snapshot", "ALTER TABLE bug_history ADD COLUMN status_snapshot TEXT"),
                ("assignee_snapshot", "ALTER TABLE bug_history ADD COLUMN assignee_snapshot TEXT"),
            ],
            "bug_comments": [
                ("parent_id", "ALTER TABLE bug_comments ADD COLUMN parent_id INTEGER"),
            ],
            "bug_attachments": [
                ("source_field", "ALTER TABLE bug_attachments ADD COLUMN source_field TEXT NOT NULL DEFAULT 'attachments'"),
            ],
            "notifications": [
                ("comment_id", "ALTER TABLE notifications ADD COLUMN comment_id INTEGER"),
            ],
            "requirements": [
                ("version", "ALTER TABLE requirements ADD COLUMN version TEXT"),
                ("status", "ALTER TABLE requirements ADD COLUMN status TEXT"),
                ("priority", "ALTER TABLE requirements ADD COLUMN priority TEXT"),
                ("description", "ALTER TABLE requirements ADD COLUMN description TEXT"),
                ("acceptance_criteria", "ALTER TABLE requirements ADD COLUMN acceptance_criteria TEXT"),
                ("requirement_doc_link", "ALTER TABLE requirements ADD COLUMN requirement_doc_link TEXT"),
                ("design_doc_link", "ALTER TABLE requirements ADD COLUMN design_doc_link TEXT"),
                ("creator_id", "ALTER TABLE requirements ADD COLUMN creator_id INTEGER"),
                ("updated_at", "ALTER TABLE requirements ADD COLUMN updated_at TEXT"),
            ],
        }.items():
            existing = column_names(table)
            for column_name, sql in column_sqls:
                if column_name not in existing:
                    db.execute(sql)
        if "comment_id" in column_names("notifications"):
            db.execute("CREATE INDEX IF NOT EXISTS idx_notifications_comment ON notifications(comment_id)")
        mail_settings_count = db.execute("SELECT COUNT(*) AS count FROM mail_settings").fetchone()["count"]
        if not mail_settings_count:
            db.execute(
                """
                INSERT INTO mail_settings (
                    id, enabled, host, port, security, username, password,
                    from_email, sender_name, send_time, last_sent_at, last_sent_date, last_result,
                    report_notify_enabled, report_notify_webhook, report_notify_secret, report_notify_send_time,
                    report_notify_project_id, report_notify_version, report_notify_base_url,
                    report_notify_manual_note, report_notify_tracking_progress,
                    report_notify_message_format, report_notify_lark_app_id, report_notify_lark_app_secret,
                    report_notify_last_sent_at, report_notify_last_sent_date, report_notify_last_result
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    1,
                    1 if DEFAULT_MAIL_SETTINGS["enabled"] else 0,
                    DEFAULT_MAIL_SETTINGS["host"],
                    int(DEFAULT_MAIL_SETTINGS["port"]),
                    DEFAULT_MAIL_SETTINGS["security"],
                    DEFAULT_MAIL_SETTINGS["username"],
                    DEFAULT_MAIL_SETTINGS["password"],
                    DEFAULT_MAIL_SETTINGS["from_email"],
                    DEFAULT_MAIL_SETTINGS["sender_name"],
                    DEFAULT_MAIL_SETTINGS["send_time"],
                    DEFAULT_MAIL_SETTINGS["last_sent_at"],
                    DEFAULT_MAIL_SETTINGS["last_sent_date"],
                    DEFAULT_MAIL_SETTINGS["last_result"],
                    1 if DEFAULT_GROUP_REPORT_SETTINGS["enabled"] else 0,
                    DEFAULT_GROUP_REPORT_SETTINGS["webhook_url"],
                    DEFAULT_GROUP_REPORT_SETTINGS["secret"],
                    DEFAULT_GROUP_REPORT_SETTINGS["send_time"],
                    None,
                    DEFAULT_GROUP_REPORT_SETTINGS["version"],
                    DEFAULT_GROUP_REPORT_SETTINGS["base_url"],
                    DEFAULT_GROUP_REPORT_SETTINGS["manual_note"],
                    DEFAULT_GROUP_REPORT_SETTINGS["tracking_progress"],
                    DEFAULT_GROUP_REPORT_SETTINGS["message_format"],
                    DEFAULT_GROUP_REPORT_SETTINGS["lark_app_id"],
                    DEFAULT_GROUP_REPORT_SETTINGS["lark_app_secret"],
                    DEFAULT_GROUP_REPORT_SETTINGS["last_sent_at"],
                    DEFAULT_GROUP_REPORT_SETTINGS["last_sent_date"],
                    DEFAULT_GROUP_REPORT_SETTINGS["last_result"],
                ),
            )
        if "version" in column_names("requirements"):
            db.execute(
                """
                UPDATE requirements
                SET version = CASE
                    WHEN code LIKE 'REQ-%' THEN substr(code, 5, instr(substr(code, 5), '-') - 1)
                    ELSE version
                END
                WHERE version IS NULL
                """
            )
        if "version" in column_names("test_cases"):
            db.execute(
                """
                UPDATE test_cases
                SET version = CASE
                    WHEN COALESCE(version, '') <> '' THEN version
                    WHEN instr(COALESCE(doc_name, ''), '-') > 0
                        AND substr(doc_name, 1, instr(doc_name, '-') - 1) LIKE '%.%'
                        THEN substr(doc_name, 1, instr(doc_name, '-') - 1)
                    WHEN instr(COALESCE(case_no, ''), '-') > 0
                        AND substr(case_no, 1, instr(case_no, '-') - 1) LIKE '%.%'
                        THEN substr(case_no, 1, instr(case_no, '-') - 1)
                    ELSE ''
                END
                WHERE COALESCE(version, '') = ''
                """
            )
        if "creator_id" in column_names("requirements"):
            db.execute("UPDATE requirements SET creator_id = 6 WHERE creator_id IS NULL")
        if "status" in column_names("requirements"):
            db.execute("UPDATE requirements SET status = 'pending' WHERE COALESCE(status, '') = ''")
        if "priority" in column_names("requirements"):
            db.execute("UPDATE requirements SET priority = '中' WHERE COALESCE(priority, '') = ''")
        if "description" in column_names("requirements"):
            db.execute("UPDATE requirements SET description = '' WHERE description IS NULL")
        if "acceptance_criteria" in column_names("requirements"):
            db.execute("UPDATE requirements SET acceptance_criteria = '' WHERE acceptance_criteria IS NULL")
        if "requirement_doc_link" in column_names("requirements"):
            db.execute("UPDATE requirements SET requirement_doc_link = '' WHERE requirement_doc_link IS NULL")
        if "design_doc_link" in column_names("requirements"):
            db.execute("UPDATE requirements SET design_doc_link = '' WHERE design_doc_link IS NULL")
        if "updated_at" in column_names("requirements"):
            db.execute("UPDATE requirements SET updated_at = created_at WHERE COALESCE(updated_at, '') = ''")
        if "creator_id" in column_names("test_cases"):
            db.execute("UPDATE test_cases SET creator_id = 7 WHERE creator_id IS NULL")
        if "severity" in column_names("bugs"):
            db.execute(
                """
                UPDATE bugs
                SET severity = CASE
                    WHEN COALESCE(priority, '') IN ('最高', '高', '中', '低', '最低', '建议') THEN priority
                    WHEN severity = '严重' THEN '最高'
                    WHEN severity = '一般' THEN '中'
                    WHEN severity = '建议' THEN '建议'
                    WHEN COALESCE(severity, '') = '' THEN '高'
                    ELSE severity
                END
                WHERE (
                    COALESCE(priority, '') IN ('最高', '高', '中', '低', '最低', '建议')
                    AND COALESCE(severity, '') <> COALESCE(priority, '')
                )
                OR COALESCE(severity, '') = ''
                OR severity IN ('严重', '一般')
                OR severity NOT IN ('最高', '高', '中', '低', '最低', '建议')
                """
            )
        if "status" in column_names("bugs"):
            db.execute("UPDATE bugs SET status = 'open' WHERE COALESCE(status, '') = ''")
            db.execute("UPDATE bugs SET status = 'pending_verification' WHERE status = 'resolved'")
            db.execute("UPDATE bugs SET status = 'in_progress' WHERE status = 'rejected'")
        if "priority" in column_names("bugs"):
            db.execute(
                """
                UPDATE bugs
                SET priority = COALESCE(NULLIF(severity, ''), '高')
                WHERE COALESCE(priority, '') <> COALESCE(NULLIF(severity, ''), '高')
                """
            )
        if "account_type" in column_names("users"):
            db.execute(
                """
                UPDATE users
                SET account_type = CASE
                    WHEN COALESCE(account_type, '') <> '' THEN account_type
                    WHEN COALESCE(role_code, '') = ? THEN 'admin'
                    ELSE 'member'
                END
                """,
                (ADMIN_ROLE_CODE,),
            )
        if "bug_no" in column_names("bugs"):
            bug_rows = db.execute(
                """
                SELECT id
                FROM bugs
                WHERE COALESCE(bug_no, '') = ''
                ORDER BY datetime(created_at) ASC, id ASC
                """
            ).fetchall()
            if bug_rows:
                current_max = 0
                existing_numbers = db.execute("SELECT bug_no FROM bugs WHERE COALESCE(bug_no, '') <> ''").fetchall()
                for row in existing_numbers:
                    bug_no_text = str(row["bug_no"] or "").strip()
                    if bug_no_text.isdigit():
                        current_max = max(current_max, int(bug_no_text))
                next_no = current_max + 1
                for row in bug_rows:
                    db.execute("UPDATE bugs SET bug_no = ? WHERE id = ?", (str(next_no).zfill(3), row["id"]))
                    next_no += 1
        if "platform" in column_names("bugs"):
            db.execute(
                """
                UPDATE bugs
                SET platform = CASE
                    WHEN module = 'AI' THEN 'AI'
                    WHEN module = 'Backend' THEN '后端'
                    WHEN module = 'H5' THEN 'H5'
                    WHEN module = 'APP' THEN '双端'
                    WHEN module = 'WEB' THEN 'WEB'
                    ELSE COALESCE(platform, '')
                END
                WHERE COALESCE(platform, '') = ''
                """
            )
        bug_history_columns = column_names("bug_history")
        if "environment_snapshot" in bug_history_columns:
            db.execute(
                """
                UPDATE bug_history
                SET environment_snapshot = COALESCE(
                    (SELECT COALESCE(bugs.environment, '') FROM bugs WHERE bugs.id = bug_history.bug_id),
                    ''
                )
                WHERE COALESCE(environment_snapshot, '') = ''
                """
            )
        if "status_snapshot" in bug_history_columns:
            db.execute(
                """
                UPDATE bug_history
                SET status_snapshot = COALESCE(
                    (SELECT COALESCE(bugs.status, '') FROM bugs WHERE bugs.id = bug_history.bug_id),
                    ''
                )
                WHERE COALESCE(status_snapshot, '') = ''
                """
            )
            db.execute("UPDATE bug_history SET status_snapshot = 'open' WHERE action IN ('初始化', '创建缺陷')")
            db.execute("UPDATE bug_history SET status_snapshot = 'in_progress' WHERE action = '开始处理'")
            db.execute("UPDATE bug_history SET status_snapshot = 'pending_verification' WHERE action IN ('标记已解决', '提交待验证')")
            db.execute("UPDATE bug_history SET status_snapshot = 'in_progress' WHERE action IN ('驳回缺陷', '退回处理')")
            db.execute("UPDATE bug_history SET status_snapshot = 'closed' WHERE action = '关闭缺陷'")
            db.execute("UPDATE bug_history SET status_snapshot = 'pending_verification' WHERE status_snapshot = 'resolved'")
            db.execute("UPDATE bug_history SET status_snapshot = 'in_progress' WHERE status_snapshot = 'rejected'")
        if "assignee_snapshot" in bug_history_columns:
            db.execute(
                """
                UPDATE bug_history
                SET assignee_snapshot = COALESCE(
                    (
                        SELECT COALESCE(users.name, '')
                        FROM bugs
                        LEFT JOIN users ON bugs.assignee_id = users.id
                        WHERE bugs.id = bug_history.bug_id
                    ),
                    ''
                )
                WHERE COALESCE(assignee_snapshot, '') = ''
                """
            )
        db.commit()

    def fetch_project_by_name(name: str) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM projects WHERE name = ?", (name,)).fetchone()

    def fetch_project(project_id: int) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()

    def fetch_projects() -> list[sqlite3.Row]:
        return get_db().execute("SELECT * FROM projects ORDER BY id").fetchall()

    def fetch_project_bug_notify_rule(project_id: int, module: str) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT *
            FROM project_bug_notify_rules
            WHERE project_id = ? AND module = ?
            """,
            (project_id, module),
        ).fetchone()

    def fetch_project_bug_notify_rule_options(project_id: int) -> list[dict[str, object]]:
        rows = get_db().execute(
            """
            SELECT *
            FROM project_bug_notify_rules
            WHERE project_id = ?
            """,
            (project_id,),
        ).fetchall()
        by_module = {str(row["module"]): row for row in rows}
        rule_options: list[dict[str, object]] = []
        for rule_option in BUG_NOTIFY_RULE_OPTIONS:
            module = str(rule_option["key"])
            row = by_module.get(module)
            rule_options.append(
                {
                    "module": module,
                    "label": str(rule_option["label"]),
                    "enabled": bool(row["enabled"]) if row else False,
                    "webhook_url": str(row["webhook_url"] or "") if row else "",
                    "secret": str(row["secret"] or "") if row else "",
                    "last_sent_at": str(row["last_sent_at"] or "") if row else "",
                    "last_result": str(row["last_result"] or "") if row else "",
                }
            )
        return rule_options

    def fetch_users() -> list[sqlite3.Row]:
        return get_db().execute("SELECT * FROM users ORDER BY id").fetchall()

    def is_bug_assignee_user(user: sqlite3.Row | None) -> bool:
        return user is not None and str(user["username"] or "").strip().lower() != "admin"

    def fetch_bug_assignee_users() -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT *
            FROM users
            WHERE LOWER(COALESCE(username, '')) != 'admin'
            ORDER BY id
            """
        ).fetchall()

    def fetch_user(user_id: int) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    def fetch_user_by_identity(identity: str) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT *
            FROM users
            WHERE username = ? OR name = ?
            ORDER BY id
            LIMIT 1
            """,
            (identity, identity),
        ).fetchone()

    def fetch_bug_assignee_user(user_id: object) -> sqlite3.Row | None:
        try:
            target_user_id = int(user_id)
        except (TypeError, ValueError):
            return None
        user = fetch_user(target_user_id)
        return user if is_bug_assignee_user(user) else None

    def fetch_mail_settings() -> dict[str, str]:
        row = get_db().execute("SELECT * FROM mail_settings WHERE id = 1").fetchone()
        if row is None:
            return DEFAULT_MAIL_SETTINGS.copy()
        settings = DEFAULT_MAIL_SETTINGS.copy()
        settings.update(
            {
                "enabled": bool(row["enabled"]),
                "host": row["host"] or "",
                "port": str(row["port"] or DEFAULT_MAIL_SETTINGS["port"]),
                "security": row["security"] or DEFAULT_MAIL_SETTINGS["security"],
                "username": row["username"] or "",
                "password": row["password"] or "",
                "from_email": row["from_email"] or "",
                "sender_name": row["sender_name"] or DEFAULT_MAIL_SETTINGS["sender_name"],
                "send_time": row["send_time"] or DEFAULT_MAIL_SETTINGS["send_time"],
                "last_sent_at": row["last_sent_at"] or "",
                "last_sent_date": row["last_sent_date"] or "",
                "last_result": row["last_result"] or "",
            }
        )
        return settings

    def fetch_group_report_settings() -> dict[str, str]:
        row = get_db().execute("SELECT * FROM mail_settings WHERE id = 1").fetchone()
        if row is None:
            return DEFAULT_GROUP_REPORT_SETTINGS.copy()
        message_format = row["report_notify_message_format"] or DEFAULT_GROUP_REPORT_SETTINGS["message_format"]
        if message_format not in GROUP_REPORT_MESSAGE_FORMATS:
            message_format = DEFAULT_GROUP_REPORT_SETTINGS["message_format"]
        settings = DEFAULT_GROUP_REPORT_SETTINGS.copy()
        settings.update(
            {
                "enabled": bool(row["report_notify_enabled"]),
                "webhook_url": row["report_notify_webhook"] or "",
                "secret": row["report_notify_secret"] or "",
                "send_time": row["report_notify_send_time"] or DEFAULT_GROUP_REPORT_SETTINGS["send_time"],
                "project_id": str(row["report_notify_project_id"] or ""),
                "version": row["report_notify_version"] or "",
                "base_url": row["report_notify_base_url"] or "",
                "manual_note": row["report_notify_manual_note"] or "",
                "tracking_progress": row["report_notify_tracking_progress"] or "",
                "message_format": message_format,
                "lark_app_id": row["report_notify_lark_app_id"] or "",
                "lark_app_secret": row["report_notify_lark_app_secret"] or "",
                "last_sent_at": row["report_notify_last_sent_at"] or "",
                "last_sent_date": row["report_notify_last_sent_date"] or "",
                "last_result": row["report_notify_last_result"] or "",
            }
        )
        return settings

    def update_mail_settings(form) -> None:
        enabled = 1 if form.get("enabled") == "1" else 0
        host = form.get("host", "").strip()
        port_text = form.get("port", "").strip() or DEFAULT_MAIL_SETTINGS["port"]
        security = form.get("security", "").strip() or DEFAULT_MAIL_SETTINGS["security"]
        username = form.get("username", "").strip()
        password = form.get("password", "").strip()
        from_email = form.get("from_email", "").strip()
        sender_name = form.get("sender_name", "").strip() or DEFAULT_MAIL_SETTINGS["sender_name"]
        send_time = form.get("send_time", "").strip() or DEFAULT_MAIL_SETTINGS["send_time"]
        if security not in dict(MAIL_SECURITY_OPTIONS):
            raise ValueError("请选择有效的邮箱加密方式。")
        try:
            port = int(port_text)
        except ValueError as exc:
            raise ValueError("端口号格式不正确。") from exc
        if host and (not from_email or not username):
            raise ValueError("请至少填写 SMTP 主机、登录账号和发件邮箱。")
        if len(send_time) != 5 or ":" not in send_time:
            raise ValueError("发送时间格式不正确，请使用 HH:MM。")
        db = get_db()
        db.execute(
            """
            UPDATE mail_settings
            SET enabled = ?, host = ?, port = ?, security = ?, username = ?, password = ?,
                from_email = ?, sender_name = ?, send_time = ?
            WHERE id = 1
            """,
            (enabled, host, port, security, username, password, from_email, sender_name, send_time),
        )
        db.commit()

    def update_group_report_settings(form) -> None:
        current_settings = fetch_group_report_settings()
        enabled = 1 if form.get("enabled") == "1" else 0
        # 前端不回显敏感值，未提交时必须保留数据库中的原配置。
        webhook_url = form.get("webhook_url", current_settings["webhook_url"]).strip()
        secret = form.get("secret", current_settings["secret"]).strip()
        send_time = form.get("send_time", "").strip() or DEFAULT_GROUP_REPORT_SETTINGS["send_time"]
        project_id_text = form.get("project_id", "").strip()
        version = form.get("version", "").strip()
        base_url = form.get("base_url", "").strip()
        manual_note = form.get("manual_note", "").strip()
        tracking_progress = form.get("tracking_progress", "").strip()
        message_format = form.get("message_format", DEFAULT_GROUP_REPORT_SETTINGS["message_format"]).strip()
        lark_app_id = form.get("lark_app_id", "").strip()
        lark_app_secret = form.get("lark_app_secret", current_settings["lark_app_secret"]).strip()
        if message_format not in GROUP_REPORT_MESSAGE_FORMATS:
            raise ValueError("请选择有效的群测试报告发送格式。")
        if webhook_url and not webhook_url.startswith(("https://", "http://")):
            raise ValueError("群机器人 Webhook 地址格式不正确。")
        if len(send_time) != 5 or ":" not in send_time:
            raise ValueError("发送时间格式不正确，请使用 HH:MM。")
        project_id = None
        if project_id_text:
            try:
                project_id = int(project_id_text)
            except ValueError as exc:
                raise ValueError("请选择有效项目。") from exc
            if fetch_project(project_id) is None:
                raise ValueError("所选项目不存在。")
        if enabled and not webhook_url:
            raise ValueError("开启群测试报告通知前，请先填写群机器人 Webhook。")
        if enabled and message_format == "image" and (not lark_app_id or not lark_app_secret):
            raise ValueError("选择图片消息时，请填写飞书应用 App ID 和 App Secret。")
        db = get_db()
        db.execute(
            """
            UPDATE mail_settings
            SET report_notify_enabled = ?, report_notify_webhook = ?, report_notify_secret = ?,
                report_notify_send_time = ?, report_notify_project_id = ?, report_notify_version = ?,
                report_notify_base_url = ?, report_notify_manual_note = ?,
                report_notify_tracking_progress = ?, report_notify_message_format = ?,
                report_notify_lark_app_id = ?, report_notify_lark_app_secret = ?
            WHERE id = 1
            """,
            (
                enabled,
                webhook_url,
                secret,
                send_time,
                project_id,
                version,
                base_url,
                manual_note,
                tracking_progress,
                message_format,
                lark_app_id,
                lark_app_secret,
            ),
        )
        db.commit()

    def validate_project_bug_notify_settings(
        enabled: int,
        webhook_url: str,
        base_url: str = "",
        label: str = "新建 Bug 群通知",
    ) -> None:
        if webhook_url and not webhook_url.startswith(("https://", "http://")):
            raise ValueError(f"{label} Webhook 地址格式不正确。")
        if base_url and not base_url.startswith(("https://", "http://")):
            raise ValueError("平台访问地址格式不正确，请填写 http:// 或 https:// 开头的地址。")
        if enabled and not webhook_url:
            raise ValueError(f"开启{label}前，请先填写群机器人 Webhook。")

    def parse_project_bug_notify_rule_form(form) -> list[dict[str, object]]:
        rules: list[dict[str, object]] = []
        for rule_option in BUG_NOTIFY_RULE_OPTIONS:
            module = str(rule_option["key"])
            enabled = 1 if form.get(f"bug_notify_rule_enabled_{module}") == "1" else 0
            webhook_url = form.get(f"bug_notify_rule_webhook_{module}", "").strip()
            secret = form.get(f"bug_notify_rule_secret_{module}", "").strip()
            validate_project_bug_notify_settings(
                enabled=enabled,
                webhook_url=webhook_url,
                label=f"{rule_option['label']} 群通知",
            )
            rules.append(
                {
                    "module": module,
                    "enabled": enabled,
                    "webhook_url": webhook_url,
                    "secret": secret,
                }
            )
        return rules

    def save_project_bug_notify_rules(project_id: int, rules: list[dict[str, object]]) -> None:
        db = get_db()
        now_text = current_time()
        for rule in rules:
            module = str(rule["module"])
            existing = fetch_project_bug_notify_rule(project_id, module)
            if existing is None:
                db.execute(
                    """
                    INSERT INTO project_bug_notify_rules (
                        project_id, module, enabled, webhook_url, secret, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project_id,
                        module,
                        int(rule["enabled"]),
                        str(rule["webhook_url"]),
                        str(rule["secret"]),
                        now_text,
                        now_text,
                    ),
                )
            else:
                db.execute(
                    """
                    UPDATE project_bug_notify_rules
                    SET enabled = ?, webhook_url = ?, secret = ?, updated_at = ?
                    WHERE project_id = ? AND module = ?
                    """,
                    (
                        int(rule["enabled"]),
                        str(rule["webhook_url"]),
                        str(rule["secret"]),
                        now_text,
                        project_id,
                        module,
                    ),
                )

    def resolve_user_role_from_form(form) -> tuple[str, str, str]:
        account_type = form.get("account_type", "member").strip() or "member"
        role_code = form.get("role_code", "").strip()
        if account_type == "admin" and not role_code:
            return account_type, ADMIN_ROLE_CODE, ADMIN_ROLE_LABEL
        if role_code == ADMIN_ROLE_CODE:
            return account_type, role_code, ADMIN_ROLE_LABEL
        role = ROLE_LABELS.get(role_code, "").strip()
        return account_type, role_code, role

    def update_mail_run_result(result_text: str, mark_daily_sent: bool = False) -> str:
        now_text = current_time()
        params: list[object] = [result_text]
        sql = "UPDATE mail_settings SET last_result = ?"
        if mark_daily_sent:
            sql += ", last_sent_at = ?, last_sent_date = ?"
            today_text = app_now().strftime("%Y-%m-%d")
            params.extend([now_text, today_text])
        sql += " WHERE id = 1"
        get_db().execute(sql, params)
        get_db().commit()
        return now_text

    def update_group_report_run_result(result_text: str, mark_daily_sent: bool = False) -> str:
        now_text = current_time()
        params: list[object] = [result_text]
        sql = "UPDATE mail_settings SET report_notify_last_result = ?"
        if mark_daily_sent:
            sql += ", report_notify_last_sent_at = ?, report_notify_last_sent_date = ?"
            today_text = app_now().strftime("%Y-%m-%d")
            params.extend([now_text, today_text])
        sql += " WHERE id = 1"
        get_db().execute(sql, params)
        get_db().commit()
        return now_text

    def fetch_user_todo_summary_rows(severity: str = "") -> list[sqlite3.Row]:
        db = get_db()
        join_parts = [
            "bugs.assignee_id = users.id",
            f"bugs.status IN ({','.join('?' for _ in TODO_STATUS_CODES)})",
        ]
        params: list[object] = list(TODO_STATUS_CODES)
        if severity:
            join_parts.append("COALESCE(bugs.severity, '') = ?")
            params.append(severity)
        return db.execute(
            f"""
            SELECT
                users.id,
                users.name,
                users.username,
                users.email,
                COUNT(bugs.id) AS todo_count
            FROM users
            LEFT JOIN bugs
                ON {' AND '.join(join_parts)}
            GROUP BY users.id, users.name, users.username, users.email
            ORDER BY users.id
            """,
            params,
        ).fetchall()

    def fetch_user_todo_items(user_id: int, limit: int = 5, severity: str = "") -> list[sqlite3.Row]:
        db = get_db()
        where_parts = [
            "bugs.assignee_id = ?",
            f"bugs.status IN ({','.join('?' for _ in TODO_STATUS_CODES)})",
        ]
        params: list[object] = [user_id, *TODO_STATUS_CODES]
        if severity:
            where_parts.append("COALESCE(bugs.severity, '') = ?")
            params.append(severity)
        params.append(limit)
        return db.execute(
            f"""
            SELECT
                bugs.id,
                bugs.bug_no,
                bugs.title,
                bugs.status,
                bugs.version,
                bugs.severity,
                projects.name AS project_name
            FROM bugs
            JOIN projects ON projects.id = bugs.project_id
            WHERE {' AND '.join(where_parts)}
            ORDER BY bugs.updated_at DESC, bugs.id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()

    def smtp_client(mail_settings: dict[str, str]):
        host = mail_settings["host"]
        port = int(mail_settings["port"] or 0)
        security = mail_settings["security"]
        if security == "ssl":
            client = smtplib.SMTP_SSL(host, port, context=ssl.create_default_context(), timeout=30)
        else:
            client = smtplib.SMTP(host, port, timeout=30)
        client.ehlo()
        if security == "tls":
            client.starttls(context=ssl.create_default_context())
            client.ehlo()
        return client

    def build_todo_email_html(user: sqlite3.Row, todo_count: int, todo_items: list[sqlite3.Row]) -> str:
        items_html = "".join(
            f"""
            <tr>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;">{format_bug_no(item["bug_no"] or item["id"])}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;">{escape(item["project_name"] or "-")}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;">{escape(item["title"] or "-")}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;">{escape(item["severity"] or "-")}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;">{escape(STATUS_LABELS.get(item["status"], item["status"]))}</td>
            </tr>
            """
            for item in todo_items
        )
        if not items_html:
            items_html = """
            <tr>
                <td colspan="5" style="padding:14px 12px;color:#7b8798;text-align:center;">当前没有待办事项</td>
            </tr>
            """
        return f"""
        <div style="font-family:'PingFang SC','Microsoft YaHei',sans-serif;background:#f5f7fb;padding:28px;">
            <div style="max-width:720px;margin:0 auto;background:#ffffff;border-radius:18px;border:1px solid #e4ebf5;overflow:hidden;">
                <div style="padding:22px 24px;background:linear-gradient(135deg,#eef5ff 0%,#f9fbff 100%);border-bottom:1px solid #e4ebf5;">
                    <h2 style="margin:0;font-size:22px;color:#223349;">Alvin's Club Bug Management Platform 待办汇总</h2>
                    <p style="margin:8px 0 0;color:#617289;font-size:13px;">{escape(user['name'])}，以下是你当前的缺陷待办汇总。</p>
                </div>
                <div style="padding:22px 24px;">
                    <div style="display:inline-block;padding:10px 16px;border-radius:999px;background:#edf4ff;color:#2d6fe3;font-size:14px;font-weight:700;">
                        当前待办总数：{todo_count}
                    </div>
                    <table style="width:100%;margin-top:18px;border-collapse:collapse;font-size:13px;color:#2a3a50;">
                        <thead>
                            <tr style="background:#f8fbff;">
                                <th style="padding:10px 12px;text-align:left;border-bottom:1px solid #edf2f7;">编号</th>
                                <th style="padding:10px 12px;text-align:left;border-bottom:1px solid #edf2f7;">项目</th>
                                <th style="padding:10px 12px;text-align:left;border-bottom:1px solid #edf2f7;">标题</th>
                                <th style="padding:10px 12px;text-align:left;border-bottom:1px solid #edf2f7;">严重级别</th>
                                <th style="padding:10px 12px;text-align:left;border-bottom:1px solid #edf2f7;">状态</th>
                            </tr>
                        </thead>
                        <tbody>
                            {items_html}
                        </tbody>
                    </table>
                    <p style="margin:18px 0 0;color:#7b8798;font-size:12px;">邮件由系统在工作日 10:05 自动发送，生成时间：{current_time()}</p>
                </div>
            </div>
        </div>
        """

    def send_todo_summary_emails(
        force: bool = False,
        mark_daily_sent: bool = False,
        fail_when_empty: bool = True,
    ) -> tuple[int, int, str]:
        raise ValueError("邮件发送已取消，请使用项目新建 Bug 群通知。")
        mail_settings = fetch_mail_settings()
        if not mail_settings["enabled"] and not force:
            raise ValueError("待办邮件通知未开启。")
        required_values = [mail_settings["host"], mail_settings["port"], mail_settings["username"], mail_settings["from_email"]]
        if not all(required_values):
            raise ValueError("请先在 Admin 中完整配置 SMTP、发件邮箱和账号信息。")
        users = [row for row in fetch_user_todo_summary_rows() if (row["email"] or "").strip()]
        if not users:
            raise ValueError("当前没有可接收邮件的账号，请先为账号填写邮箱。")
        deliverable_users = [row for row in users if int(row["todo_count"] or 0) > 0]
        skipped_count = len(users) - len(deliverable_users)
        if not deliverable_users:
            result_text = "今日无待办，无需发送。"
            sent_at = update_mail_run_result(result_text, mark_daily_sent=mark_daily_sent)
            if fail_when_empty:
                raise ValueError("当前没有待办，无需发送。")
            return 0, skipped_count, sent_at

        sent_count = 0
        with smtp_client(mail_settings) as client:
            if mail_settings["username"]:
                client.login(mail_settings["username"], mail_settings["password"])
            for user in deliverable_users:
                todo_items = fetch_user_todo_items(int(user["id"]))
                todo_count = int(user["todo_count"] or 0)
                message = EmailMessage()
                message["Subject"] = f"[待办汇总] {user['name']} 当前有 {todo_count} 个待办"
                message["From"] = f"{mail_settings['sender_name']} <{mail_settings['from_email']}>"
                message["To"] = user["email"]
                plain_lines = [
                    f"{user['name']}，您好：",
                    f"当前待办总数：{todo_count}",
                    "",
                ]
                for item in todo_items:
                    plain_lines.append(
                        f"- {format_bug_no(item['bug_no'] or item['id'])} | {item['project_name']} | {item['title']} | {item['severity'] or '-'} | {STATUS_LABELS.get(item['status'], item['status'])}"
                    )
                message.set_content("\n".join(plain_lines))
                message.add_alternative(build_todo_email_html(user, todo_count, todo_items), subtype="html")
                client.send_message(message)
                sent_count += 1

        sent_at = update_mail_run_result(
            f"成功发送 {sent_count} 封，未发送 {skipped_count} 个无待办账号。",
            mark_daily_sent=mark_daily_sent,
        )
        return sent_count, skipped_count, sent_at

    def build_bug_detail_absolute_url(bug_id: int, base_url: str = "") -> str:
        base_url = (base_url.strip() or request.host_url or "").rstrip("/")
        return f"{base_url}{url_for('bug_detail', bug_id=bug_id)}" if base_url else url_for("bug_detail", bug_id=bug_id)

    def build_app_absolute_url(endpoint: str, base_url: str = "", **values: object) -> str:
        base = (base_url.strip() or request.host_url or "").rstrip("/")
        path = url_for(endpoint, **values)
        return f"{base}{path}" if base else path

    def build_severe_bug_assignment_email_html(
        assignee_user: sqlite3.Row,
        bug: sqlite3.Row,
        trigger_reason: str,
        operator_name: str,
        bug_url: str,
    ) -> str:
        rows = [
            ("缺陷编号", format_bug_no(bug["bug_no"] or bug["id"])),
            ("项目", bug["project_name"] or "-"),
            ("标题", bug["title"] or "-"),
            ("当前状态", STATUS_LABELS.get(str(bug["status"] or ""), str(bug["status"] or "-"))),
            ("当前处理人", assignee_user["name"] or "-"),
            ("创建人", bug["creator_name"] or "-"),
            ("严重级别", bug["severity"] or "-"),
            ("触发动作", trigger_reason),
            ("操作人", operator_name or "-"),
        ]
        if bug["version"]:
            rows.insert(3, ("版本", bug["version"]))
        if bug["platform"]:
            rows.insert(4, ("端", bug["platform"]))
        if bug["environment"]:
            rows.append(("环境", bug["environment"]))
        rows_html = "".join(
            f"""
            <tr>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;width:92px;color:#66778e;">{escape(label)}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf2f7;color:#243446;">{escape(str(value))}</td>
            </tr>
            """
            for label, value in rows
        )
        description_html = escape(str(bug["description"] or "-")).replace("\n", "<br>")
        return f"""
        <div style="font-family:'PingFang SC','Microsoft YaHei',sans-serif;background:#f5f7fb;padding:28px;">
            <div style="max-width:760px;margin:0 auto;background:#ffffff;border-radius:18px;border:1px solid #e4ebf5;overflow:hidden;">
                <div style="padding:22px 24px;background:linear-gradient(135deg,#fff4f2 0%,#fffaf8 100%);border-bottom:1px solid #f2ddd8;">
                    <h2 style="margin:0;font-size:22px;color:#223349;">Alvin's Club Bug Management Platform 严重 Bug 通知</h2>
                    <p style="margin:8px 0 0;color:#617289;font-size:13px;">{escape(assignee_user['name'])}，你有一条新的严重 Bug 待处理，请尽快关注。</p>
                </div>
                <div style="padding:22px 24px;">
                    <div style="display:inline-block;padding:10px 16px;border-radius:999px;background:#fff1ea;color:#c0563f;font-size:14px;font-weight:700;">
                        严重级别：{escape(str(bug['severity'] or '-'))}
                    </div>
                    <table style="width:100%;margin-top:18px;border-collapse:collapse;font-size:13px;">
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                    <div style="margin-top:18px;padding:14px 16px;border-radius:14px;background:#f8fafc;border:1px solid #e8edf5;">
                        <div style="font-size:12px;color:#6f7f92;margin-bottom:8px;">问题描述</div>
                        <div style="font-size:13px;color:#243446;line-height:1.8;">{description_html}</div>
                    </div>
                    <p style="margin:18px 0 0;font-size:13px;">
                        <a href="{escape(bug_url)}" style="color:#1f63d8;text-decoration:none;">点击查看缺陷详情</a>
                    </p>
                    <p style="margin:12px 0 0;color:#7b8798;font-size:12px;">邮件发送时间：{current_time()}</p>
                </div>
            </div>
        </div>
        """

    def send_single_severe_bug_notification(
        bug: sqlite3.Row | None,
        assignee_user_id: int | None,
        trigger_reason: str,
        operator_name: str,
    ) -> tuple[bool, str]:
        if bug is None:
            return False, "未找到对应 Bug。"
        if str(bug["severity"] or "") != MAIL_NOTIFY_SEVERITY:
            return False, "当前 Bug 非严重级别。"
        if str(bug["status"] or "") not in TODO_STATUS_CODES:
            return False, "当前状态无需发送严重 Bug 待办通知。"
        if not assignee_user_id:
            return False, "当前处理人为空，无法发送通知。"

        assignee_user = fetch_user(int(assignee_user_id))
        if assignee_user is None:
            return False, "当前处理人不存在。"
        recipient_email = str(assignee_user["email"] or "").strip()
        if not recipient_email:
            return False, f"{assignee_user['name']} 未配置邮箱。"

        mail_settings = fetch_mail_settings()
        if not mail_settings["enabled"]:
            return False, "待办邮件通知未开启。"
        required_values = [mail_settings["host"], mail_settings["port"], mail_settings["username"], mail_settings["from_email"]]
        if not all(required_values):
            return False, "SMTP、发件邮箱或邮箱账号配置不完整。"

        bug_url = build_bug_detail_absolute_url(int(bug["id"]))
        message = EmailMessage()
        message["Subject"] = f"[严重Bug通知] {format_bug_no(bug['bug_no'] or bug['id'])} 已进入你的待办"
        message["From"] = f"{mail_settings['sender_name']} <{mail_settings['from_email']}>"
        message["To"] = recipient_email
        plain_lines = [
            f"{assignee_user['name']}，您好：",
            f"你有一条严重 Bug 待办，请尽快处理。",
            f"缺陷编号：{format_bug_no(bug['bug_no'] or bug['id'])}",
            f"项目：{bug['project_name'] or '-'}",
            f"标题：{bug['title'] or '-'}",
            f"状态：{STATUS_LABELS.get(str(bug['status'] or ''), str(bug['status'] or '-'))}",
            f"触发动作：{trigger_reason}",
            f"操作人：{operator_name or '-'}",
            f"详情链接：{bug_url}",
        ]
        message.set_content("\n".join(plain_lines))
        message.add_alternative(
            build_severe_bug_assignment_email_html(
                assignee_user=assignee_user,
                bug=bug,
                trigger_reason=trigger_reason,
                operator_name=operator_name,
                bug_url=bug_url,
            ),
            subtype="html",
        )

        with smtp_client(mail_settings) as client:
            if mail_settings["username"]:
                client.login(mail_settings["username"], mail_settings["password"])
            client.send_message(message)

        update_mail_run_result(
            f"已即时通知 {assignee_user['name']}：{format_bug_no(bug['bug_no'] or bug['id'])} {trigger_reason}",
            mark_daily_sent=False,
        )
        return True, f"已通知 {assignee_user['name']}（{recipient_email}）。"

    def maybe_send_severe_bug_assignment_notification(
        bug_id: int,
        assignee_user_id: int | None,
        trigger_reason: str,
        operator_name: str,
    ) -> tuple[bool, str]:
        bug = fetch_bug(bug_id)
        return create_severe_bug_assignment_message(
            bug=bug,
            assignee_user_id=assignee_user_id,
            trigger_reason=trigger_reason,
            operator_name=operator_name,
        )

    def update_project_bug_notify_result(project_id: int, result_text: str, mark_sent: bool = False) -> str:
        now_text = current_time()
        params: list[object] = [result_text]
        sql = "UPDATE projects SET bug_notify_last_result = ?"
        if mark_sent:
            sql += ", bug_notify_last_sent_at = ?"
            params.append(now_text)
        sql += " WHERE id = ?"
        params.append(project_id)
        get_db().execute(sql, params)
        get_db().commit()
        return now_text

    def update_project_bug_notify_rule_result(
        project_id: int,
        module: str,
        result_text: str,
        mark_sent: bool = False,
    ) -> str:
        now_text = current_time()
        params: list[object] = [result_text]
        sql = "UPDATE project_bug_notify_rules SET last_result = ?"
        if mark_sent:
            sql += ", last_sent_at = ?"
            params.append(now_text)
        sql += " WHERE project_id = ? AND module = ?"
        params.extend([project_id, module])
        get_db().execute(sql, params)
        get_db().commit()
        return now_text

    def build_new_bug_group_message(bug: sqlite3.Row, operator_name: str, bug_url: str) -> str:
        lines = [
            "新建 Bug 通知",
            f"项目：{bug['project_name'] or '-'}",
            f"缺陷编号：{format_bug_no(bug['bug_no'] or bug['id'])}",
            f"标题：{bug['title'] or '-'}",
            f"严重级别：{bug['severity'] or '-'}",
            f"状态：{STATUS_LABELS.get(str(bug['status'] or ''), str(bug['status'] or '-'))}",
            f"当前处理人：{bug['assignee_name'] or '-'}",
            f"创建人：{bug['creator_name'] or operator_name or '-'}",
            f"版本：{bug['version'] or '-'}",
            f"端：{bug['platform'] or '-'}",
        ]
        if bug["environment"]:
            lines.append(f"环境：{bug['environment']}")
        lines.extend(
            [
                "",
                f"问题描述：{str(bug['description'] or '-').strip()}",
            ]
        )
        if bug_url:
            lines.extend(["", f"详情链接：{bug_url}"])
        lines.extend(["", f"发送时间：{current_time()}"])
        return "\n".join(lines)

    def compact_card_text(value: object, default: str = "-") -> str:
        text = str(value or "").strip()
        return re.sub(r"\s+", " ", text) if text else default

    def clipped_card_text(value: object, limit: int = 180, default: str = "-") -> str:
        text = compact_card_text(value, default=default)
        if len(text) <= limit:
            return text
        return f"{text[:limit].rstrip()}..."

    def build_lark_link_button(label: str, target_url: str, button_type: str = "default") -> dict[str, object]:
        return {
            "tag": "button",
            "text": {"tag": "plain_text", "content": label},
            "type": button_type,
            "url": target_url,
            "multi_url": {
                "url": target_url,
                "pc_url": target_url,
                "ios_url": target_url,
                "android_url": target_url,
            },
        }

    def build_new_bug_group_card_payload(
        bug: sqlite3.Row,
        operator_name: str,
        bug_url: str,
        base_url: str = "",
        secret: str = "",
    ) -> bytes:
        status_label = STATUS_LABELS.get(str(bug["status"] or ""), str(bug["status"] or "-"))
        requirement_label = compact_card_text(bug["requirement_title"] or bug["requirement_code"] or bug["version"])
        fields = [
            ("Bug ID", format_bug_no(bug["bug_no"] or bug["id"])),
            ("项目", bug["project_name"]),
            ("需求", requirement_label),
            ("端", bug["platform"]),
            ("状态", status_label),
            ("优先级", bug["priority"]),
            ("严重程度", bug["severity"]),
            ("负责人", bug["assignee_name"]),
            ("提交人", bug["creator_name"] or operator_name),
        ]
        actions: list[dict[str, object]] = []
        if bug_url:
            actions.append(
                build_lark_link_button("查看 Bug 详情", bug_url, button_type="primary")
            )
        if bug["requirement_id"]:
            actions.append(
                build_lark_link_button(
                    "进入需求页",
                    build_app_absolute_url(
                        "requirement_detail",
                        base_url,
                        requirement_id=int(bug["requirement_id"]),
                    ),
                )
            )
        actions.append(
            build_lark_link_button(
                "AI处理",
                build_app_absolute_url("bug_detail", base_url, bug_id=int(bug["id"]), tab="process"),
            )
        )
        payload: dict[str, object] = {
            "msg_type": "interactive",
            "card": {
                "config": {"wide_screen_mode": True},
                "header": {
                    "template": "red",
                    "title": {"tag": "plain_text", "content": "新建 Bug"},
                },
                "elements": [
                    {
                        "tag": "div",
                        "text": {
                            "tag": "lark_md",
                            "content": f"**{clipped_card_text(bug['title'], limit=90)}**",
                        },
                    },
                    {
                        "tag": "div",
                        "text": {
                            "tag": "plain_text",
                            "content": clipped_card_text(bug["description"], limit=160),
                        },
                    },
                    {
                        "tag": "div",
                        "fields": [
                            {
                                "is_short": False,
                                "text": {
                                    "tag": "lark_md",
                                    "content": f"**{label}：** {compact_card_text(value)}",
                                },
                            }
                            for label, value in fields
                        ],
                    },
                    {"tag": "hr"},
                    {"tag": "action", "actions": actions},
                ],
            },
        }
        return build_group_robot_payload(payload, secret=secret)

    def maybe_send_new_bug_group_notification(
        bug_id: int,
        operator_name: str,
    ) -> tuple[bool, str]:
        bug = fetch_bug(bug_id)
        if bug is None:
            return False, "未找到对应 Bug。"
        project = fetch_project(int(bug["project_id"] or 0))
        if project is None:
            return False, "未找到对应项目。"
        bug_platform = str(bug["platform"] or "").strip()
        bug_notify_key = bug_notify_key_for_platform(bug_platform) or str(bug["module"] or "").strip()
        module_rule = fetch_project_bug_notify_rule(int(project["id"]), bug_notify_key) if bug_notify_key else None
        notify_target = "项目默认群"
        webhook_url = ""
        secret = ""
        is_module_rule = False
        if module_rule is not None and bool(module_rule["enabled"]):
            notify_target = f"{bug_notify_label_for_key(bug_notify_key)}群"
            webhook_url = str(module_rule["webhook_url"] or "").strip()
            secret = str(module_rule["secret"] or "").strip()
            is_module_rule = True
        elif bool(project["bug_notify_enabled"]):
            webhook_url = str(project["bug_notify_webhook"] or "").strip()
            secret = str(project["bug_notify_secret"] or "").strip()
        else:
            return False, "当前项目未开启新建 Bug 群通知。"
        if not webhook_url:
            result_text = f"新建 Bug 群通知未发送：{notify_target}未配置群机器人 Webhook。"
            if is_module_rule:
                update_project_bug_notify_rule_result(int(project["id"]), bug_notify_key, result_text)
            else:
                update_project_bug_notify_result(int(project["id"]), result_text)
            return False, result_text

        bug_url = build_bug_detail_absolute_url(int(bug["id"]), str(project["bug_notify_base_url"] or ""))
        message_text = build_new_bug_group_message(
            bug=bug,
            operator_name=operator_name,
            bug_url=bug_url,
        )
        request_body = build_new_bug_group_card_payload(
            bug=bug,
            operator_name=operator_name,
            bug_url=bug_url,
            base_url=str(project["bug_notify_base_url"] or ""),
            secret=secret,
        )
        try:
            send_group_report_message(
                webhook_url=webhook_url,
                message_text=message_text,
                request_body=request_body,
            )
        except Exception as exc:
            result_text = f"新建 Bug 群通知发送失败：{exc}"
            if is_module_rule:
                update_project_bug_notify_rule_result(int(project["id"]), bug_notify_key, result_text)
            else:
                update_project_bug_notify_result(int(project["id"]), result_text)
            return False, result_text

        result_text = f"新建 Bug 群通知已发送到{notify_target}：{format_bug_no(bug['bug_no'] or bug['id'])} {bug['title'] or ''}".strip()
        if is_module_rule:
            sent_at = update_project_bug_notify_rule_result(int(project["id"]), bug_notify_key, result_text, mark_sent=True)
        else:
            sent_at = update_project_bug_notify_result(int(project["id"]), result_text, mark_sent=True)
        return True, f"已发送到{notify_target}（{sent_at}）。"

    def build_group_report_message(
        project: sqlite3.Row,
        version: str,
        summary: dict,
        case_total: int,
        distribution: list[dict],
        risk_bugs: list[sqlite3.Row],
        generated_at: str,
        open_bug_platform_counts: list[dict[str, object]] | None = None,
        base_url: str = "",
        manual_note: str = "",
        tracking_progress: str = "",
    ) -> str:
        not_run_count = next((item["count"] for item in distribution if item["status"] == "未测"), 0)
        executed_count = max(case_total - not_run_count, 0)
        progress_percent = "0%" if case_total <= 0 else f"{(executed_count / case_total) * 100:.0f}%"
        fixed_bug_count = int(summary["verification_count"]) + int(summary["closed_count"])
        reopened_or_open_count = int(summary["active_count"])
        project_label = project["name"] if not version else f"{project['name']}{version}"
        manual_note_lines = [line.strip() for line in str(manual_note or "").splitlines() if line.strip()]
        tracking_progress_lines = [line.strip() for line in str(tracking_progress or "").splitlines() if line.strip()]
        open_bug_platform_counts = open_bug_platform_counts or []
        risk_bug_lines = [
            (
                f"  - {format_bug_no(item['bug_no'] or item['id'])} | "
                f"{item['title'] or '-'} | "
                f"{STATUS_LABELS.get(str(item['status'] or ''), str(item['status'] or '-'))} | "
                f"{item['assignee_name'] or '未分配'}"
            )
            for item in risk_bugs[:5]
        ]
        if len(risk_bugs) > 5:
            risk_bug_lines.append(f"  - 还有 {len(risk_bugs) - 5} 个严重 Bug 未展示")
        lines = [
            f"测试项目：{project_label}",
            f"• 整体测试进度：{progress_percent}",
            f"• 用例执行进度：已测 {executed_count} / 总用例 {case_total}",
            "• 缺陷情况",
            f"• 发现 Bug 数：{summary['total']}",
            f"• 修复 Bug 数：{fixed_bug_count}",
            f"• 已回归验证：{summary['closed_count']}",
            f"• 还打开 Bug 数：{reopened_or_open_count}",
        ]
        for item in open_bug_platform_counts:
            lines.append(f"  {item['label']}：{item['count']}")
        lines.append(f"• 严重 Bug 汇总：{len(risk_bugs)} 个待处理")
        lines.extend(risk_bug_lines or ["  - 暂无未关闭严重 Bug"])
        if manual_note_lines:
            lines.append("• 风险备注：")
            lines.extend(f"  - {line}" for line in manual_note_lines)
        if tracking_progress_lines:
            lines.append(f"• 埋点进度：{'；'.join(tracking_progress_lines)}")
        return "\n".join(lines)

    def report_percent(value: float) -> str:
        rounded = round(value)
        return f"{rounded}%" if abs(value - rounded) < 0.05 else f"{value:.1f}%"

    def build_group_report_card_payload(
        project: sqlite3.Row,
        version: str,
        summary: dict,
        case_total: int,
        distribution: list[dict],
        risk_bugs: list[sqlite3.Row],
        generated_at: str,
        open_bug_platform_counts: list[dict[str, object]] | None = None,
        manual_note: str = "",
        tracking_progress: str = "",
        secret: str = "",
    ) -> bytes:
        not_run_count = next((int(item["count"]) for item in distribution if item["status"] == "未测"), 0)
        executed_count = max(case_total - not_run_count, 0)
        progress_value = 0.0 if case_total <= 0 else (executed_count / case_total) * 100
        headline_progress_percent = "0%" if case_total <= 0 else f"{round(progress_value)}%"
        fixed_bug_count = int(summary["verification_count"]) + int(summary["closed_count"])
        active_count = int(summary["active_count"])
        total_bug_count = int(summary["total"])
        fix_rate_value = 0.0 if total_bug_count <= 0 else (fixed_bug_count / total_bug_count) * 100
        fix_rate = report_percent(fix_rate_value)
        closed_count = int(summary["closed_count"])
        project_label = project["name"] if not version else f"{project['name']} {version}"
        generated_date = generated_at.split(" ", 1)[0].replace("-", ".")
        open_bug_platform_counts = open_bug_platform_counts or []
        platform_text = "；".join(
            f"{item['label']} {item['count']}" for item in open_bug_platform_counts
        )
        risk_level = "低风险" if active_count == 0 and not risk_bugs else "有风险"
        clear_label = "缺陷清零" if active_count == 0 else f"{active_count} 个打开缺陷"
        if active_count == 0:
            conclusion = (
                f"测试整体完成度 {headline_progress_percent}，当前无阻塞缺陷；"
                f"剩余 {not_run_count} 条用例建议在发版前完成收尾执行与冒烟复核。"
            )
        else:
            conclusion = (
                f"测试整体完成度 {headline_progress_percent}，仍有 {active_count} 个打开缺陷待处理"
                + (f"（{platform_text}）" if platform_text else "")
                + f"；剩余 {not_run_count} 条用例待执行。"
            )
        manual_note_lines = [line.strip() for line in str(manual_note or "").splitlines() if line.strip()]
        tracking_progress_lines = [line.strip() for line in str(tracking_progress or "").splitlines() if line.strip()]
        note_lines: list[str]
        if risk_bugs:
            note_lines = [f"**严重 Bug 汇总：{len(risk_bugs)} 个待处理**"]
            note_lines.extend(
                (
                    f"- {format_bug_no(item['bug_no'] or item['id'])} "
                    f"{item['title'] or '-'}"
                    f"（{STATUS_LABELS.get(str(item['status'] or ''), str(item['status'] or '-'))} / "
                    f"{item['assignee_name'] or '未分配'}）"
                )
                for item in risk_bugs[:5]
            )
            if len(risk_bugs) > 5:
                note_lines.append(f"- 还有 {len(risk_bugs) - 5} 个严重 Bug 未展示")
        else:
            note_lines = ["**严重 Bug 汇总：0 个待处理**"]
            if not manual_note_lines:
                note_lines.append("- 暂无未关闭严重 Bug")
        if manual_note_lines:
            note_lines.append("**风险备注**")
            note_lines.extend(f"- {line}" for line in manual_note_lines)
        tracking_progress_text = "；".join(tracking_progress_lines)
        metrics = [
            ("总", "整体进度", headline_progress_percent, ""),
            ("例", "用例执行", f"{executed_count} / {case_total}", f"剩余 {not_run_count} 条待执行"),
            ("修", "缺陷修复", f"{fixed_bug_count} / {total_bug_count}", f"修复率 {fix_rate}"),
            ("开", "打开缺陷", str(active_count), "当前无阻塞缺陷" if active_count == 0 else "需要继续跟进"),
        ]
        metric_tag_colors = ["turquoise", "blue", "orange", "red"]
        chain_items = [
            ("发现", str(total_bug_count), "累计提交", "blue-600"),
            ("修复", str(fixed_bug_count), "全部完成" if active_count == 0 else "处理中", "orange-600"),
            ("回归", str(closed_count), "验证通过", "turquoise-600"),
            ("打开", str(active_count), "无遗留" if active_count == 0 else "待处理", "grey-600"),
        ]
        risk_tag_text = f"严重 {len(risk_bugs)}"
        risk_tag_markdown_color = "orange" if risk_bugs else "green"

        def div_text(
            content: str,
            text_size: str = "normal",
            text_color: str = "grey-1000",
            text_align: str = "left",
            tag: str = "plain_text",
            margin: str = "0px",
        ) -> dict[str, object]:
            return {
                "tag": "div",
                "margin": margin,
                "text": {
                    "tag": tag,
                    "content": content,
                    "text_size": text_size,
                    "text_color": text_color,
                    "text_align": text_align,
                },
            }

        def container(
            elements: list[dict[str, object]],
            background_style: str = "default",
            padding: str = "10px",
            corner_radius: str = "6px",
            has_border: bool = False,
            border_color: str = "grey-300",
            direction: str = "vertical",
            width: str = "fill",
            height: str = "auto",
            horizontal_align: str = "left",
            vertical_align: str = "top",
            horizontal_spacing: str = "6px",
            vertical_spacing: str = "6px",
            margin: str = "0px",
        ) -> dict[str, object]:
            data: dict[str, object] = {
                "tag": "interactive_container",
                "width": width,
                "height": height,
                "direction": direction,
                "horizontal_align": horizontal_align,
                "vertical_align": vertical_align,
                "horizontal_spacing": horizontal_spacing,
                "vertical_spacing": vertical_spacing,
                "background_style": background_style,
                "corner_radius": corner_radius,
                "padding": padding,
                "margin": margin,
                "behaviors": [],
                "elements": elements,
            }
            if has_border:
                data["has_border"] = True
                data["border_color"] = border_color
            return data

        def column(
            elements: list[dict[str, object]],
            weight: int = 1,
            width: str = "weighted",
            padding: str = "0px",
            background_style: str = "default",
            horizontal_align: str = "left",
            vertical_align: str = "top",
            vertical_spacing: str = "6px",
        ) -> dict[str, object]:
            return {
                "tag": "column",
                "width": width,
                "weight": weight,
                "padding": padding,
                "background_style": background_style,
                "horizontal_align": horizontal_align,
                "vertical_align": vertical_align,
                "vertical_spacing": vertical_spacing,
                "elements": elements,
            }

        def column_set(
            columns: list[dict[str, object]],
            flex_mode: str = "flow",
            horizontal_spacing: str = "10px",
            margin: str = "0px",
            horizontal_align: str = "left",
        ) -> dict[str, object]:
            return {
                "tag": "column_set",
                "flex_mode": flex_mode,
                "horizontal_spacing": horizontal_spacing,
                "horizontal_align": horizontal_align,
                "margin": margin,
                "columns": columns,
            }

        def metric_card(icon: str, title: str, value: str, sub_text: str, tag_color: str) -> dict[str, object]:
            elements = [
                div_text(
                    f"<text_tag color='{tag_color}'>{icon}</text_tag> {title}",
                    text_size="normal",
                    text_color="grey-600",
                    tag="lark_md",
                ),
                div_text(value, text_size="heading-2", text_color="grey-1000", margin="4px 0px 0px 0px"),
            ]
            if sub_text:
                elements.append(div_text(sub_text, text_size="notation", text_color="grey-600"))
            return container(
                elements,
                background_style="bg-white",
                padding="12px",
                corner_radius="6px",
                has_border=True,
                border_color="grey-300",
                vertical_spacing="4px",
            )

        def chain_node(label: str, value: str, sub_text: str, text_color: str) -> dict[str, object]:
            return column(
                [
                    div_text(value, text_size="heading", text_color=text_color, text_align="center"),
                    div_text(label, text_size="normal", text_color="grey-1000", text_align="center"),
                    div_text(sub_text, text_size="notation", text_color="grey-600", text_align="center"),
                ],
                weight=1,
                horizontal_align="center",
                vertical_spacing="2px",
            )

        metric_columns = [
            column([metric_card(icon, title, value, sub_text, tag_color)], weight=1)
            for (icon, title, value, sub_text), tag_color in zip(metrics, metric_tag_colors)
        ]
        chain_columns = [chain_node(label, value, sub_text, text_color) for label, value, sub_text, text_color in chain_items]
        note_text = "\n".join(note_lines)
        elements = [
            container(
                [
                    container(
                        [
                            div_text(
                                "QA",
                                text_size="normal",
                                text_color="grey-00",
                                text_align="center",
                            )
                        ],
                        background_style="turquoise-600",
                        padding="0px",
                        corner_radius="6px",
                        width="34px",
                        height="34px",
                        horizontal_align="center",
                        vertical_align="center",
                    ),
                    column_set(
                        [
                            column(
                                [
                                    div_text("测试日报", text_size="heading-2", text_color="grey-00"),
                                    div_text(
                                        f"{project_label} · 发布前收尾验证",
                                        text_size="notation",
                                        text_color="grey-350",
                                    ),
                                ],
                                weight=3,
                                vertical_align="center",
                                vertical_spacing="0px",
                            ),
                            column(
                                [
                                    div_text(
                                        (
                                            f"<text_tag color='green'>{risk_level}</text_tag> "
                                            f"<text_tag color='yellow'>{clear_label}</text_tag>"
                                        ),
                                        text_size="normal",
                                        text_color="grey-00",
                                        text_align="right",
                                        tag="lark_md",
                                    ),
                                    div_text(
                                        generated_date,
                                        text_size="normal",
                                        text_color="grey-350",
                                        text_align="right",
                                    ),
                                ],
                                weight=2,
                                horizontal_align="right",
                                vertical_align="center",
                            ),
                        ],
                        flex_mode="flow",
                        horizontal_spacing="12px",
                    ),
                ],
                background_style="indigo-900",
                padding="12px 18px 12px 18px",
                corner_radius="0px",
                direction="horizontal",
                horizontal_spacing="10px",
                vertical_align="center",
            ),
            container(
                [
                    column_set(
                        [
                            column(
                                [div_text("今日结论", text_size="normal", text_color="turquoise-600")],
                                width="82px",
                                vertical_align="center",
                            ),
                            column(
                                [div_text(conclusion, text_size="normal", text_color="grey-1000")],
                                weight=1,
                                vertical_align="center",
                            ),
                        ],
                        flex_mode="flow",
                        horizontal_spacing="10px",
                    )
                ],
                background_style="turquoise-50",
                padding="10px 14px 10px 14px",
                margin="10px 16px 0px 16px",
                corner_radius="6px",
                has_border=True,
                border_color="turquoise-200",
            ),
            column_set(metric_columns, flex_mode="bisect", horizontal_spacing="10px", margin="6px 16px 0px 16px"),
            container(
                [
                    div_text("缺陷闭环", text_size="heading", text_color="grey-1000"),
                    div_text(
                        "发现 · 修复 · 回归 · 打开缺陷状态一屏看清",
                        text_size="notation",
                        text_color="grey-600",
                    ),
                    column_set(chain_columns, flex_mode="bisect", horizontal_spacing="6px", margin="0px"),
                    div_text(
                        f"发现 {total_bug_count} → 修复 {fixed_bug_count} → 回归 {closed_count} → 打开 {active_count}",
                        text_size="notation",
                        text_color="grey-600",
                        text_align="center",
                    ),
                ],
                background_style="bg-white",
                padding="14px",
                corner_radius="6px",
                has_border=True,
                border_color="grey-300",
                vertical_spacing="4px",
                margin="0px 16px 0px 16px",
            ),
            container(
                [
                    div_text("严重 Bug / 风险备注", text_size="heading", text_color="grey-1000"),
                    column_set(
                        [
                            column(
                                [
                                    div_text(
                                        f"<text_tag color='{risk_tag_markdown_color}'>{risk_tag_text}</text_tag>",
                                        text_size="normal",
                                        text_color="turquoise-600",
                                        tag="lark_md",
                                    )
                                ],
                                width="74px",
                                vertical_align="top",
                            ),
                            column(
                                [
                                    div_text(
                                        note_text,
                                        text_size="normal",
                                        text_color="grey-1000",
                                        tag="lark_md",
                                    )
                                ],
                                weight=1,
                                vertical_align="center",
                            ),
                        ],
                        flex_mode="flow",
                        horizontal_spacing="8px",
                    ),
                ],
                background_style="bg-white",
                padding="14px",
                corner_radius="6px",
                has_border=True,
                border_color="grey-300",
                vertical_spacing="5px",
                margin="0px 16px 0px 16px",
            ),
        ]
        if tracking_progress_lines:
            elements.append(
                container(
                    [
                        div_text(
                            f"埋点进度：{tracking_progress_text}",
                            text_size="heading",
                            text_color="grey-1000",
                            tag="lark_md",
                        ),
                    ],
                    background_style="bg-white",
                    padding="14px",
                    corner_radius="6px",
                    has_border=True,
                    border_color="grey-300",
                    vertical_spacing="5px",
                    margin="0px 16px 0px 16px",
                )
            )
        elements.append(
            column_set(
                [
                    column(
                        [
                            div_text(
                                "建议日报结构：结论先行 / 关键指标 / 缺陷闭环 / 风险备注",
                                text_size="notation",
                                text_color="grey-600",
                            )
                        ],
                        weight=1,
                    ),
                    column(
                        [
                            div_text(
                                "QA Daily Report",
                                text_size="notation",
                                text_color="grey-600",
                                text_align="right",
                            )
                        ],
                        weight=1,
                    ),
                ],
                flex_mode="flow",
                margin="0px 16px 0px 16px",
            )
        )

        payload: dict[str, object] = {
            "msg_type": "interactive",
            "card": {
                "schema": "2.0",
                "config": {
                    "update_multi": True,
                    "summary": {"content": "QA 测试日报"},
                },
                "body": {
                    "direction": "vertical",
                    "padding": "0px 0px 6px 0px",
                    "vertical_spacing": "8px",
                    "elements": elements,
                },
            },
        }
        return build_group_robot_payload(payload, secret=secret)

    def load_report_image_font(size: int, bold: bool = False):
        from PIL import ImageFont

        font_candidates = [
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/STHeiti Light.ttc",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
            "/Library/Fonts/Arial Unicode.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        ]
        if bold:
            font_candidates = [
                "/System/Library/Fonts/PingFang.ttc",
                "/System/Library/Fonts/STHeiti Medium.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
                "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
                *font_candidates,
            ]
        for font_path in font_candidates:
            path = Path(font_path)
            if not path.exists():
                continue
            try:
                return ImageFont.truetype(str(path), size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    def report_image_text_width(draw, text: str, font) -> float:
        try:
            return float(draw.textlength(text, font=font))
        except Exception:
            bbox = draw.textbbox((0, 0), text, font=font)
            return float(bbox[2] - bbox[0])

    def wrap_report_image_text(
        draw,
        text: object,
        font,
        max_width: int,
        max_lines: int | None = None,
    ) -> list[str]:
        lines: list[str] = []
        paragraphs = str(text or "").replace("\r\n", "\n").splitlines() or [""]
        for paragraph in paragraphs:
            paragraph = paragraph.strip()
            if not paragraph:
                lines.append("")
                continue
            current = ""
            tokens = re.findall(r"[A-Za-z0-9%._:/+-]+|[ \t]+|.", paragraph)
            for token in tokens:
                if token.isspace() and not current:
                    continue
                candidate = f"{current}{token}"
                if not current or report_image_text_width(draw, candidate, font) <= max_width:
                    current = candidate
                else:
                    lines.append(current.rstrip())
                    current = token.lstrip()
                    while current and report_image_text_width(draw, current, font) > max_width:
                        clipped = ""
                        for char in current:
                            char_candidate = f"{clipped}{char}"
                            if not clipped or report_image_text_width(draw, char_candidate, font) <= max_width:
                                clipped = char_candidate
                            else:
                                break
                        lines.append(clipped)
                        current = current[len(clipped) :]
            if current:
                lines.append(current.rstrip())
        if max_lines is not None and len(lines) > max_lines:
            clipped = lines[:max_lines]
            last_line = clipped[-1].rstrip()
            while last_line and report_image_text_width(draw, f"{last_line}...", font) > max_width:
                last_line = last_line[:-1].rstrip()
            clipped[-1] = f"{last_line}..." if last_line else "..."
            return clipped
        return lines

    def draw_report_image_lines(
        draw,
        lines: list[str],
        x: int,
        y: int,
        font,
        fill: str,
        line_height: int,
    ) -> int:
        for line in lines:
            draw.text((x, y), line, font=font, fill=fill)
            y += line_height
        return y

    def build_group_report_image_bytes(
        project: sqlite3.Row,
        version: str,
        summary: dict,
        case_total: int,
        distribution: list[dict],
        risk_bugs: list[sqlite3.Row],
        generated_at: str,
        open_bug_platform_counts: list[dict[str, object]] | None = None,
        manual_note: str = "",
        tracking_progress: str = "",
    ) -> bytes:
        from PIL import Image, ImageDraw

        width = 820
        max_height = 2600
        outer_margin = 24
        content_margin = 42
        card_width = width - outer_margin * 2
        content_width = width - content_margin * 2
        image = Image.new("RGB", (width, max_height), "#f3f6fb")
        draw = ImageDraw.Draw(image)

        font_title = load_report_image_font(34, bold=True)
        font_heading = load_report_image_font(24, bold=True)
        font_body = load_report_image_font(21)
        font_body_bold = load_report_image_font(21, bold=True)
        font_small = load_report_image_font(17)
        font_metric = load_report_image_font(31, bold=True)
        font_qa = load_report_image_font(22, bold=True)

        not_run_count = next((int(item["count"]) for item in distribution if item["status"] == "未测"), 0)
        executed_count = max(case_total - not_run_count, 0)
        progress_value = 0.0 if case_total <= 0 else (executed_count / case_total) * 100
        headline_progress_percent = "0%" if case_total <= 0 else f"{round(progress_value)}%"
        fixed_bug_count = int(summary["verification_count"]) + int(summary["closed_count"])
        active_count = int(summary["active_count"])
        total_bug_count = int(summary["total"])
        closed_count = int(summary["closed_count"])
        fix_rate_value = 0.0 if total_bug_count <= 0 else (fixed_bug_count / total_bug_count) * 100
        fix_rate = report_percent(fix_rate_value)
        project_label = project["name"] if not version else f"{project['name']} {version}"
        generated_date = generated_at.split(" ", 1)[0].replace("-", ".")
        open_bug_platform_counts = open_bug_platform_counts or []
        platform_text = "；".join(f"{item['label']} {item['count']}" for item in open_bug_platform_counts)
        risk_level = "低风险" if active_count == 0 and not risk_bugs else "有风险"
        clear_label = "缺陷清零" if active_count == 0 else f"{active_count} 个打开缺陷"
        if active_count == 0:
            conclusion = (
                f"测试整体完成度 {headline_progress_percent}，当前无阻塞缺陷；"
                f"剩余 {not_run_count} 条用例建议在发版前完成收尾执行与冒烟复核。"
            )
        else:
            conclusion = (
                f"测试整体完成度 {headline_progress_percent}，仍有 {active_count} 个打开缺陷待处理"
                + (f"（{platform_text}）" if platform_text else "")
                + f"；剩余 {not_run_count} 条用例待执行。"
            )
        manual_note_lines = [line.strip() for line in str(manual_note or "").splitlines() if line.strip()]
        tracking_progress_lines = [line.strip() for line in str(tracking_progress or "").splitlines() if line.strip()]
        tracking_progress_text = "；".join(tracking_progress_lines)

        def rounded_box(x1: int, y1: int, x2: int, y2: int, fill: str, outline: str | None = None, radius: int = 14) -> None:
            draw.rounded_rectangle((x1, y1, x2, y2), radius=radius, fill=fill, outline=outline, width=1)

        def badge(x: int, y: int, text: str, fill: str, text_fill: str) -> int:
            padding_x = 13
            badge_width = int(report_image_text_width(draw, text, font_small)) + padding_x * 2
            rounded_box(x, y, x + badge_width, y + 28, fill, radius=14)
            draw.text((x + padding_x, y + 5), text, font=font_small, fill=text_fill)
            return x + badge_width + 8

        y = 24
        rounded_box(outer_margin, y, width - outer_margin, y + 148, "#17233f", radius=18)
        rounded_box(content_margin, y + 34, content_margin + 50, y + 84, "#14b8a6", radius=8)
        draw.text((content_margin + 12, y + 47), "QA", font=font_qa, fill="#ffffff")
        draw.text((content_margin + 68, y + 32), "测试日报", font=font_title, fill="#ffffff")
        draw.text((content_margin + 68, y + 80), f"{project_label} · 发布前收尾验证", font=font_body, fill="#cbd5e1")
        tag_x = badge(width - content_margin - 188, y + 36, risk_level, "#dcfce7" if risk_level == "低风险" else "#fef3c7", "#166534" if risk_level == "低风险" else "#92400e")
        badge(tag_x, y + 36, clear_label, "#e0f2fe", "#075985")
        draw.text((width - content_margin - 100, y + 80), generated_date, font=font_small, fill="#cbd5e1")
        y += 174

        conclusion_lines = wrap_report_image_text(draw, conclusion, font_body, content_width - 136, max_lines=3)
        conclusion_height = max(78, 28 + len(conclusion_lines) * 29)
        rounded_box(outer_margin, y, width - outer_margin, y + conclusion_height, "#dff7f2", "#99e6d8", radius=14)
        draw.text((content_margin, y + 24), "今日结论", font=font_body_bold, fill="#0f766e")
        draw_report_image_lines(draw, conclusion_lines, content_margin + 112, y + 22, font_body, "#1f2937", 29)
        y += conclusion_height + 16

        metrics = [
            ("整体进度", headline_progress_percent, "", "#0f766e"),
            ("用例执行", f"{executed_count} / {case_total}", f"剩余 {not_run_count} 条待执行", "#2563eb"),
            ("缺陷修复", f"{fixed_bug_count} / {total_bug_count}", f"修复率 {fix_rate}", "#ea580c"),
            ("打开缺陷", str(active_count), "当前无阻塞缺陷" if active_count == 0 else "需要继续跟进", "#dc2626"),
        ]
        metric_gap = 14
        metric_w = (card_width - content_margin + outer_margin - metric_gap) // 2
        metric_h = 112
        for index, (title, value, sub_text, color) in enumerate(metrics):
            row = index // 2
            col = index % 2
            x1 = outer_margin + col * (metric_w + metric_gap)
            y1 = y + row * (metric_h + metric_gap)
            rounded_box(x1, y1, x1 + metric_w, y1 + metric_h, "#ffffff", "#d8dee9", radius=14)
            draw.text((x1 + 18, y1 + 16), title, font=font_small, fill="#64748b")
            draw.text((x1 + 18, y1 + 42), value, font=font_metric, fill=color)
            if sub_text:
                draw.text((x1 + 18, y1 + 82), sub_text, font=font_small, fill="#64748b")
        y += metric_h * 2 + metric_gap + 18

        rounded_box(outer_margin, y, width - outer_margin, y + 132, "#ffffff", "#d8dee9", radius=14)
        draw.text((content_margin, y + 18), "缺陷闭环", font=font_heading, fill="#111827")
        chain_labels = [
            ("发现", str(total_bug_count), "累计提交", "#2563eb"),
            ("修复", str(fixed_bug_count), "全部完成" if active_count == 0 else "处理中", "#ea580c"),
            ("回归", str(closed_count), "验证通过", "#0f766e"),
            ("打开", str(active_count), "无遗留" if active_count == 0 else "待处理", "#64748b"),
        ]
        node_w = content_width // 4
        for index, (label, value, sub_text, color) in enumerate(chain_labels):
            cx = content_margin + index * node_w + node_w // 2
            draw.text((cx - report_image_text_width(draw, value, font_heading) / 2, y + 52), value, font=font_heading, fill=color)
            draw.text((cx - report_image_text_width(draw, label, font_small) / 2, y + 82), label, font=font_small, fill="#111827")
            draw.text((cx - report_image_text_width(draw, sub_text, font_small) / 2, y + 104), sub_text, font=font_small, fill="#64748b")
        y += 150

        risk_lines = [f"严重 Bug 汇总：{len(risk_bugs)} 个待处理"]
        risk_lines.extend(
            (
                f"{format_bug_no(item['bug_no'] or item['id'])} {item['title'] or '-'}"
                f"（{STATUS_LABELS.get(str(item['status'] or ''), str(item['status'] or '-'))} / "
                f"{item['assignee_name'] or '未分配'}）"
            )
            for item in risk_bugs[:5]
        )
        if len(risk_bugs) > 5:
            risk_lines.append(f"还有 {len(risk_bugs) - 5} 个严重 Bug 未展示")
        if not risk_bugs and not manual_note_lines:
            risk_lines.append("暂无未关闭严重 Bug")
        if manual_note_lines:
            risk_lines.append("风险备注")
            risk_lines.extend(manual_note_lines)

        rendered_risk_lines: list[str] = []
        for line in risk_lines:
            rendered_risk_lines.extend(wrap_report_image_text(draw, line, font_body, content_width, max_lines=2))
        risk_height = 60 + len(rendered_risk_lines) * 30
        rounded_box(outer_margin, y, width - outer_margin, y + risk_height, "#ffffff", "#d8dee9", radius=14)
        draw.text((content_margin, y + 18), "严重 Bug / 风险备注", font=font_heading, fill="#111827")
        draw_report_image_lines(draw, rendered_risk_lines, content_margin, y + 54, font_body, "#1f2937", 30)
        y += risk_height + 16

        if tracking_progress_text:
            tracking_line = f"埋点进度：{tracking_progress_text}"
            tracking_lines = wrap_report_image_text(draw, tracking_line, font_body_bold, content_width, max_lines=2)
            tracking_height = 34 + len(tracking_lines) * 32
            rounded_box(outer_margin, y, width - outer_margin, y + tracking_height, "#ffffff", "#d8dee9", radius=14)
            draw_report_image_lines(draw, tracking_lines, content_margin, y + 22, font_body_bold, "#111827", 32)
            y += tracking_height + 16

        footer = "QA Daily Report · 固定图片版式"
        draw.text((content_margin, y + 8), footer, font=font_small, fill="#64748b")
        final_height = min(max_height, y + 48)
        cropped = image.crop((0, 0, width, final_height))
        stream = io.BytesIO()
        cropped.save(stream, format="PNG", optimize=True)
        return stream.getvalue()

    def fetch_lark_tenant_access_token(app_id: str, app_secret: str) -> str:
        if not app_id or not app_secret:
            raise ValueError("图片消息需要先配置飞书应用 App ID 和 App Secret。")
        request_body = json.dumps(
            {"app_id": app_id, "app_secret": app_secret},
            ensure_ascii=False,
        ).encode("utf-8")
        http_request = urllib_request.Request(
            "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
            data=request_body,
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Content-Length": str(len(request_body)),
            },
            method="POST",
        )
        try:
            with urllib_request.urlopen(http_request, timeout=20) as response:
                raw_body = response.read().decode("utf-8", errors="replace")
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"获取飞书访问凭证失败：HTTP {exc.code} {detail}") from exc
        except urllib_error.URLError as exc:
            raise ValueError(f"获取飞书访问凭证失败：{exc.reason}") from exc
        try:
            payload = json.loads(raw_body or "{}")
        except json.JSONDecodeError as exc:
            raise ValueError(f"获取飞书访问凭证失败：响应不是 JSON：{raw_body}") from exc
        if payload.get("code") not in (None, 0):
            raise ValueError(f"获取飞书访问凭证失败：{payload.get('msg') or payload}")
        token = str(payload.get("tenant_access_token") or "").strip()
        if not token:
            raise ValueError("获取飞书访问凭证失败：响应缺少 tenant_access_token。")
        return token

    def upload_lark_report_image(image_bytes: bytes, app_id: str, app_secret: str) -> str:
        token = fetch_lark_tenant_access_token(app_id, app_secret)
        boundary = f"----BugPlatformReport{uuid.uuid4().hex}"
        body = b"".join(
            [
                f"--{boundary}\r\n".encode("utf-8"),
                b'Content-Disposition: form-data; name="image_type"\r\n\r\n',
                b"message\r\n",
                f"--{boundary}\r\n".encode("utf-8"),
                b'Content-Disposition: form-data; name="image"; filename="qa_daily_report.png"\r\n',
                b"Content-Type: image/png\r\n\r\n",
                image_bytes,
                b"\r\n",
                f"--{boundary}--\r\n".encode("utf-8"),
            ]
        )
        http_request = urllib_request.Request(
            "https://open.feishu.cn/open-apis/im/v1/images",
            data=body,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": f"multipart/form-data; boundary={boundary}",
                "Content-Length": str(len(body)),
            },
            method="POST",
        )
        try:
            with urllib_request.urlopen(http_request, timeout=30) as response:
                raw_body = response.read().decode("utf-8", errors="replace")
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"上传日报图片失败：HTTP {exc.code} {detail}") from exc
        except urllib_error.URLError as exc:
            raise ValueError(f"上传日报图片失败：{exc.reason}") from exc
        try:
            payload = json.loads(raw_body or "{}")
        except json.JSONDecodeError as exc:
            raise ValueError(f"上传日报图片失败：响应不是 JSON：{raw_body}") from exc
        if payload.get("code") not in (None, 0):
            raise ValueError(f"上传日报图片失败：{payload.get('msg') or payload}")
        image_key = str((payload.get("data") or {}).get("image_key") or "").strip()
        if not image_key:
            raise ValueError("上传日报图片失败：响应缺少 image_key。")
        return image_key

    def build_group_robot_payload(payload: dict[str, object], secret: str = "") -> bytes:
        if secret:
            timestamp = str(int(time.time()))
            string_to_sign = f"{timestamp}\n{secret}"
            sign = base64.b64encode(
                hmac.new(string_to_sign.encode("utf-8"), digestmod=hashlib.sha256).digest()
            ).decode("utf-8")
            payload["timestamp"] = timestamp
            payload["sign"] = sign
        return json.dumps(payload, ensure_ascii=False).encode("utf-8")

    def build_group_report_payload(message_text: str, secret: str = "") -> bytes:
        payload: dict[str, object] = {
            "msg_type": "text",
            "content": {
                "text": message_text,
            },
        }
        return build_group_robot_payload(payload, secret=secret)

    def build_group_report_image_payload(image_key: str, secret: str = "") -> bytes:
        payload: dict[str, object] = {
            "msg_type": "image",
            "content": {
                "image_key": image_key,
            },
        }
        return build_group_robot_payload(payload, secret=secret)

    def send_group_report_message(
        webhook_url: str,
        message_text: str,
        secret: str = "",
        request_body: bytes | None = None,
    ) -> dict[str, object]:
        request_body = request_body or build_group_report_payload(message_text, secret=secret)
        http_request = urllib_request.Request(
            webhook_url,
            data=request_body,
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Content-Length": str(len(request_body)),
            },
            method="POST",
        )
        try:
            with urllib_request.urlopen(http_request, timeout=20) as response:
                raw_body = response.read().decode("utf-8", errors="replace")
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"群机器人请求失败：HTTP {exc.code} {detail}") from exc
        except urllib_error.URLError as exc:
            raise ValueError(f"群机器人请求失败：{exc.reason}") from exc
        try:
            payload = json.loads(raw_body or "{}")
        except json.JSONDecodeError:
            return {"ok": True, "raw": raw_body}
        status_code = payload.get("code", payload.get("StatusCode"))
        status_message = payload.get("msg") or payload.get("StatusMessage") or payload
        if status_code not in (None, 0):
            raise ValueError(f"群机器人发送失败：{status_message}")
        return payload

    def send_testing_report_to_group(
        force: bool = False,
        mark_daily_sent: bool = False,
        manual_note: str = "",
    ) -> tuple[str, str, str]:
        settings = fetch_group_report_settings()
        if not settings["enabled"] and not force:
            raise ValueError("群测试报告通知未开启。")
        webhook_url = settings["webhook_url"].strip()
        if not webhook_url:
            raise ValueError("请先配置群机器人 Webhook。")
        project_id_text = settings["project_id"].strip()
        if not project_id_text:
            raise ValueError("请先选择要发送测试报告的项目。")
        try:
            project_id = int(project_id_text)
        except ValueError as exc:
            raise ValueError("群测试报告通知中的项目配置无效。") from exc
        project = fetch_project(project_id)
        if project is None:
            raise ValueError("所选项目不存在。")
        version = settings["version"].strip()
        effective_manual_note = str(manual_note or "").strip() or settings["manual_note"].strip()
        effective_tracking_progress = settings["tracking_progress"].strip()
        message_format = settings["message_format"].strip() or DEFAULT_GROUP_REPORT_SETTINGS["message_format"]
        if message_format not in GROUP_REPORT_MESSAGE_FORMATS:
            message_format = DEFAULT_GROUP_REPORT_SETTINGS["message_format"]
        summary = fetch_summary(version=version, project_id=project_id)
        case_total = count_test_cases(version=version, project_id=project_id)
        distribution = execution_distribution(project_id=project_id, version=version)
        risk_bugs = fetch_report_risk_bugs(project_id=project_id, version=version)
        open_bug_platform_counts = fetch_open_bug_counts_by_platform(project_id=project_id, version=version)
        generated_at = current_time()
        message_text = build_group_report_message(
            project=project,
            version=version,
            summary=summary,
            case_total=case_total,
            distribution=distribution,
            risk_bugs=risk_bugs,
            generated_at=generated_at,
            open_bug_platform_counts=open_bug_platform_counts,
            base_url=settings["base_url"].strip(),
            manual_note=effective_manual_note,
            tracking_progress=effective_tracking_progress,
        )
        if message_format == "image":
            image_bytes = build_group_report_image_bytes(
                project=project,
                version=version,
                summary=summary,
                case_total=case_total,
                distribution=distribution,
                risk_bugs=risk_bugs,
                generated_at=generated_at,
                open_bug_platform_counts=open_bug_platform_counts,
                manual_note=effective_manual_note,
                tracking_progress=effective_tracking_progress,
            )
            image_key = upload_lark_report_image(
                image_bytes=image_bytes,
                app_id=settings["lark_app_id"].strip(),
                app_secret=settings["lark_app_secret"].strip(),
            )
            request_body = build_group_report_image_payload(image_key, secret=settings["secret"].strip())
        else:
            request_body = build_group_report_card_payload(
                project=project,
                version=version,
                summary=summary,
                case_total=case_total,
                distribution=distribution,
                risk_bugs=risk_bugs,
                generated_at=generated_at,
                open_bug_platform_counts=open_bug_platform_counts,
                manual_note=effective_manual_note,
                tracking_progress=effective_tracking_progress,
                secret=settings["secret"].strip(),
            )
        send_group_report_message(
            webhook_url=webhook_url,
            message_text=message_text,
            request_body=request_body,
        )
        note_suffix = "（含补充信息）" if (effective_manual_note or effective_tracking_progress) else ""
        format_label = GROUP_REPORT_MESSAGE_FORMATS.get(message_format, GROUP_REPORT_MESSAGE_FORMATS["card"])
        sent_at = update_group_report_run_result(
            f"测试报告已发送到群：{project['name']} / {version or '全部版本'} / {format_label}{note_suffix}",
            mark_daily_sent=mark_daily_sent,
        )
        return project["name"], version or "全部版本", sent_at

    def can_manage_bug(bug: sqlite3.Row | None) -> bool:
        return bug is not None and g.current_user is not None and (
            is_admin() or int(bug["creator_id"] or 0) == int(g.current_user["id"])
        )

    def can_edit_bug_platform(bug: sqlite3.Row | None) -> bool:
        return bug is not None and g.current_user is not None and (
            is_admin()
            or int(bug["creator_id"] or 0) == int(g.current_user["id"])
            or int(bug["assignee_id"] or 0) == int(g.current_user["id"])
        )

    def can_manage_bug_comment(comment: sqlite3.Row | None) -> bool:
        return comment is not None and g.current_user is not None and (
            is_admin() or int(comment["user_id"] or 0) == int(g.current_user["id"])
        )

    def can_manage_requirement(requirement: sqlite3.Row | None) -> bool:
        return requirement is not None and g.current_user is not None and (
            is_admin() or int(requirement["creator_id"] or 0) == int(g.current_user["id"])
        )

    def can_manage_case_document(document: sqlite3.Row | None) -> bool:
        return document is not None and g.current_user is not None and (
            is_admin() or int(document["creator_id"] or 0) == int(g.current_user["id"])
        )

    def can_edit_case_execution(document: sqlite3.Row | None) -> bool:
        return document is not None and g.current_user is not None

    def project_usage_count(project_id: int) -> int:
        db = get_db()
        bug_count = int(db.execute("SELECT COUNT(*) AS count FROM bugs WHERE project_id = ?", (project_id,)).fetchone()["count"])
        requirement_count = int(db.execute("SELECT COUNT(*) AS count FROM requirements WHERE project_id = ?", (project_id,)).fetchone()["count"])
        case_count = int(db.execute("SELECT COUNT(*) AS count FROM test_cases WHERE project_id = ?", (project_id,)).fetchone()["count"])
        return bug_count + requirement_count + case_count

    def user_usage_count(user_id: int) -> int:
        db = get_db()
        assignee_count = int(db.execute("SELECT COUNT(*) AS count FROM bugs WHERE assignee_id = ?", (user_id,)).fetchone()["count"])
        creator_count = int(db.execute("SELECT COUNT(*) AS count FROM bugs WHERE creator_id = ?", (user_id,)).fetchone()["count"])
        previous_count = int(db.execute("SELECT COUNT(*) AS count FROM bugs WHERE previous_assignee_id = ?", (user_id,)).fetchone()["count"])
        return assignee_count + creator_count + previous_count

    def delete_project_with_related_data(project_id: int) -> dict[str, int]:
        db = get_db()
        bug_rows = db.execute("SELECT id FROM bugs WHERE project_id = ?", (project_id,)).fetchall()
        bug_ids = [int(row["id"]) for row in bug_rows]
        case_rows = db.execute("SELECT id FROM test_cases WHERE project_id = ?", (project_id,)).fetchall()
        case_ids = [int(row["id"]) for row in case_rows]
        attachment_paths: list[str] = []
        deleted_attachments = 0
        deleted_histories = 0
        deleted_bugs = 0
        if bug_ids:
            placeholders = ",".join("?" for _ in bug_ids)
            attachment_paths = [
                row["file_path"]
                for row in db.execute(
                    f"SELECT file_path FROM bug_attachments WHERE bug_id IN ({placeholders})",
                    bug_ids,
                ).fetchall()
            ]
            deleted_attachments = int(
                db.execute(f"DELETE FROM bug_attachments WHERE bug_id IN ({placeholders})", bug_ids).rowcount or 0
            )
            deleted_histories = int(
                db.execute(f"DELETE FROM bug_history WHERE bug_id IN ({placeholders})", bug_ids).rowcount or 0
            )
            db.execute(f"DELETE FROM notifications WHERE bug_id IN ({placeholders})", bug_ids)
            db.execute(f"DELETE FROM bug_comments WHERE bug_id IN ({placeholders})", bug_ids)
            deleted_bugs = int(db.execute("DELETE FROM bugs WHERE project_id = ?", (project_id,)).rowcount or 0)
        if case_ids:
            case_placeholders = ",".join("?" for _ in case_ids)
            db.execute(f"DELETE FROM case_document_cells WHERE case_id IN ({case_placeholders})", case_ids)
        db.execute("DELETE FROM case_document_columns WHERE project_id = ?", (project_id,))
        db.execute("DELETE FROM project_bug_notify_rules WHERE project_id = ?", (project_id,))
        deleted_requirements = int(db.execute("DELETE FROM requirements WHERE project_id = ?", (project_id,)).rowcount or 0)
        deleted_cases = int(db.execute("DELETE FROM test_cases WHERE project_id = ?", (project_id,)).rowcount or 0)
        deleted_projects = int(db.execute("DELETE FROM projects WHERE id = ?", (project_id,)).rowcount or 0)
        db.commit()
        for file_path in attachment_paths:
            try:
                attachment_file = Path(file_path)
                if attachment_file.exists():
                    attachment_file.unlink()
            except OSError:
                pass
        return {
            "projects": deleted_projects,
            "bugs": deleted_bugs,
            "requirements": deleted_requirements,
            "cases": deleted_cases,
            "attachments": deleted_attachments,
            "histories": deleted_histories,
        }

    def fetch_user_by_credentials(username: str, password: str) -> sqlite3.Row | None:
        return get_db().execute(
            "SELECT * FROM users WHERE username = ? AND password = ?",
            (username, password),
        ).fetchone()

    scheduler_started = {"value": False}

    def due_for_daily_mail(now: datetime, mail_settings: dict[str, str]) -> bool:
        if not mail_settings["enabled"]:
            return False
        send_time = mail_settings["send_time"] or DEFAULT_MAIL_SETTINGS["send_time"]
        if len(send_time) != 5 or ":" not in send_time:
            return False
        if now.weekday() >= 5:
            return False
        last_sent_date = mail_settings["last_sent_date"] or ""
        return now.strftime("%H:%M") == send_time and last_sent_date != now.strftime("%Y-%m-%d")

    def due_for_daily_group_report(now: datetime, settings: dict[str, str]) -> bool:
        if not settings["enabled"]:
            return False
        send_time = settings["send_time"] or DEFAULT_GROUP_REPORT_SETTINGS["send_time"]
        if len(send_time) != 5 or ":" not in send_time:
            return False
        last_sent_date = settings["last_sent_date"] or ""
        return now.strftime("%H:%M") == send_time and last_sent_date != now.strftime("%Y-%m-%d")

    def claim_daily_mail_send(now: datetime) -> bool:
        today_text = now.strftime("%Y-%m-%d")
        cursor = get_db().execute(
            """
            UPDATE mail_settings
            SET last_sent_date = ?, last_result = ?
            WHERE id = 1 AND COALESCE(last_sent_date, '') != ?
            """,
            (today_text, f"定时发送中：{current_time()}", today_text),
        )
        get_db().commit()
        return bool(cursor.rowcount)

    def claim_daily_group_report_send(now: datetime) -> bool:
        today_text = now.strftime("%Y-%m-%d")
        cursor = get_db().execute(
            """
            UPDATE mail_settings
            SET report_notify_last_sent_date = ?, report_notify_last_result = ?
            WHERE id = 1 AND COALESCE(report_notify_last_sent_date, '') != ?
            """,
            (today_text, f"群测试报告定时发送中：{current_time()}", today_text),
        )
        get_db().commit()
        return bool(cursor.rowcount)

    def start_mail_scheduler() -> None:
        if scheduler_started["value"]:
            return
        scheduler_started["value"] = True

        def scheduler_loop() -> None:
            while True:
                try:
                    with app.app_context():
                        now = app_now()
                        mail_settings = fetch_mail_settings()
                        if due_for_daily_mail(now, mail_settings) and claim_daily_mail_send(now):
                            try:
                                send_todo_summary_emails(force=False, mark_daily_sent=True, fail_when_empty=False)
                            except Exception as exc:
                                get_db().execute(
                                    "UPDATE mail_settings SET last_result = ? WHERE id = 1",
                                    (f"定时发送失败：{exc}",),
                                )
                                get_db().commit()
                        group_report_settings = fetch_group_report_settings()
                        if due_for_daily_group_report(now, group_report_settings) and claim_daily_group_report_send(now):
                            try:
                                send_testing_report_to_group(force=False, mark_daily_sent=True)
                            except Exception as exc:
                                update_group_report_run_result(f"群测试报告发送失败：{exc}", mark_daily_sent=False)
                except Exception:
                    pass
                time.sleep(30)

        thread = threading.Thread(target=scheduler_loop, name="todo-mail-scheduler", daemon=True)
        thread.start()

    def is_safe_local_path(target: str) -> bool:
        if not target.startswith("/") or target.startswith("//") or target.startswith("/\\"):
            return False
        if "\\" in target:
            return False
        return not any(ord(char) < 32 or ord(char) == 127 for char in target)

    def local_redirect_target(raw_target: str, default_url: str) -> str:
        target = str(raw_target or "").strip()
        if not target:
            return default_url
        parsed = urllib_parse.urlparse(target)
        if parsed.netloc or parsed.scheme:
            if parsed.scheme not in {"http", "https"} or parsed.netloc != request.host:
                return default_url
            target = parsed.path or default_url
            if parsed.query:
                target = f"{target}?{parsed.query}"
            if parsed.fragment:
                target = f"{target}#{parsed.fragment}"
        return target if is_safe_local_path(target) else default_url

    def admin_redirect_target() -> str:
        return local_redirect_target(request.form.get("next", ""), url_for("admin_center"))

    def local_back_url(default_url: str) -> str:
        next_url = local_redirect_target(request.values.get("next", ""), "")
        if next_url:
            return next_url
        referrer = (request.referrer or "").strip()
        if not referrer:
            return default_url
        parsed = urllib_parse.urlparse(referrer)
        if parsed.netloc or parsed.scheme:
            if parsed.scheme not in {"http", "https"} or parsed.netloc != request.host:
                return default_url
        elif not is_safe_local_path(referrer):
            return default_url
        candidate = parsed.path or default_url
        if parsed.query:
            candidate = f"{candidate}?{parsed.query}"
        if parsed.fragment:
            candidate = f"{candidate}#{parsed.fragment}"
        return local_redirect_target(candidate, default_url)

    def current_local_path() -> str:
        # 反代子路径下 request.path 不含挂载前缀，表单回跳需要补齐。
        path = request.full_path if request.query_string else request.path
        script_root = (request.script_root or "").rstrip("/")
        if script_root and path != script_root and not path.startswith(f"{script_root}/"):
            return f"{script_root}{path}"
        return path

    def require_admin_access() -> Response | None:
        if not is_admin():
            flash("仅管理员可访问。", "error")
            return redirect(url_for("bug_list"))
        return None

    def admin_dashboard_cards() -> list[dict[str, str | int]]:
        group_report_settings = fetch_group_report_settings()
        return [
            {
                "title": "项目管理",
                "desc": "创建项目、修改项目信息，并配置新建 Bug 群通知。",
                "count_label": "项目数",
                "count": len(fetch_projects()),
                "href": url_for("admin_projects_page"),
            },
            {
                "title": "账号管理",
                "desc": "创建账号、分配角色、维护登录信息与邮箱配置。",
                "count_label": "账号数",
                "count": len(fetch_users()),
                "href": url_for("admin_users_page"),
            },
            {
                "title": "群测试报告通知",
                "desc": "配置飞书群机器人，每日自动把测试报告推送到群里。",
                "count_label": "通知状态",
                "count": "已开启" if group_report_settings["enabled"] else "未开启",
                "href": url_for("admin_report_notify_page"),
            },
        ]

    def fetch_requirement_by_code(code: str) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM requirements WHERE code = ?", (code,)).fetchone()

    def fetch_case_by_no(case_no: str) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM test_cases WHERE case_no = ?", (case_no,)).fetchone()

    def fetch_case(case_id: int) -> sqlite3.Row | None:
        return get_db().execute("SELECT * FROM test_cases WHERE id = ?", (case_id,)).fetchone()

    def build_bug_form_prefill_from_case(case_item: sqlite3.Row | None) -> dict:
        if case_item is None:
            return {}
        case_title = str(case_item["title"] or "").strip()
        case_no = str(case_item["case_no"] or "").strip()
        title = case_title or case_no
        if case_no and case_title:
            title = f"{case_no} {case_title}"
        return {
            "title": title,
            "version": str(case_item["version"] or "").strip(),
            "module": "APP",
            "platform": "",
            "severity": "高",
            "priority": "高",
            "assignee_id": "",
            "requirement_id": "",
            "case_id": str(case_item["id"]),
            "environment": str(case_item["environment_info"] or "").strip(),
            "description": str(case_item["steps"] or "").strip(),
            "expected_result": str(case_item["expected_result"] or "").strip(),
            "actual_result": "",
        }

    def build_bug_form_prefill_from_request(case_item: sqlite3.Row | None = None) -> dict:
        form_values = build_bug_form_prefill_from_case(case_item)
        if not form_values.get("version"):
            selected_version = request.args.get("version", "").strip()
            if selected_version:
                form_values["version"] = selected_version
        return form_values

    def find_user_id(db: sqlite3.Connection, user_name: str) -> int:
        row = db.execute("SELECT id FROM users WHERE name = ?", (user_name,)).fetchone()
        if not row:
            raise ValueError(f"用户不存在: {user_name}")
        return row["id"]

    def find_project_id(db: sqlite3.Connection, project_name: str) -> int:
        row = db.execute("SELECT id FROM projects WHERE name = ?", (project_name,)).fetchone()
        if not row:
            raise ValueError(f"项目不存在: {project_name}")
        return row["id"]

    def insert_bug(
        db: sqlite3.Connection,
        title: str,
        project_id: int,
        version: str,
        module: str,
        platform: str,
        severity: str,
        priority: str,
        status: str,
        assignee_id: int,
        creator_id: int,
        previous_assignee_id: int | None,
        requirement_id: int | None,
        case_id: int | None,
        environment: str,
        description: str,
        expected_result: str,
        actual_result: str,
        resolution_note: str,
    ) -> int:
        now = current_time()
        creator_name = db.execute("SELECT name FROM users WHERE id = ?", (creator_id,)).fetchone()["name"]
        max_bug_no_row = db.execute(
            """
            SELECT MAX(CAST(bug_no AS INTEGER)) AS max_bug_no
            FROM bugs
            WHERE COALESCE(bug_no, '') <> '' AND bug_no GLOB '[0-9]*'
            """
        ).fetchone()
        next_bug_no = str(int(max_bug_no_row["max_bug_no"] or 0) + 1).zfill(3)
        cursor = db.execute(
            """
            INSERT INTO bugs (
                bug_no, title, project_id, version, module, platform, severity, priority, status, assignee_id,
                creator_id, previous_assignee_id, reporter, requirement_id, case_id,
                environment, description, expected_result, actual_result, resolution_note,
                created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                next_bug_no,
                title,
                project_id,
                version,
                module,
                platform,
                severity,
                priority,
                status,
                assignee_id,
                creator_id,
                previous_assignee_id,
                creator_name,
                requirement_id,
                case_id,
                environment,
                description,
                expected_result,
                actual_result,
                resolution_note,
                now,
                now,
            ),
        )
        return int(cursor.lastrowid)

    def add_history(
        db: sqlite3.Connection,
        bug_id: int,
        action: str,
        detail: str,
        operator_name: str,
        environment_snapshot: str = "",
        status_snapshot: str = "",
        assignee_snapshot: str = "",
    ) -> None:
        db.execute(
            """
            INSERT INTO bug_history (
                bug_id, action, detail, operator_name, environment_snapshot,
                status_snapshot, assignee_snapshot, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                bug_id,
                action,
                detail,
                operator_name,
                environment_snapshot,
                status_snapshot,
                assignee_snapshot,
                current_time(),
            ),
        )

    def normalize_attachment_source(source: object) -> str:
        source_text = str(source or "").strip()
        if source_text in BUG_ATTACHMENT_SOURCE_FIELDS:
            return source_text
        return "attachments"

    def save_bug_attachments(db: sqlite3.Connection, bug_id: int, files: list, source_fields: list[str] | None = None) -> list[str]:
        upload_dir = Path(app.config["UPLOAD_FOLDER"])
        saved_names: list[str] = []
        source_fields = source_fields or []
        for index, file in enumerate(files):
            if file is None or not file.filename:
                continue
            source_field = normalize_attachment_source(source_fields[index] if index < len(source_fields) else "")
            original_name = secure_filename(file.filename) or f"attachment-{uuid.uuid4().hex}"
            stored_name = f"{app_now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex}{Path(original_name).suffix}"
            destination = upload_dir / stored_name
            file.save(destination)
            db.execute(
                """
                INSERT INTO bug_attachments (bug_id, filename, stored_name, content_type, source_field, file_path, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    bug_id,
                    file.filename,
                    stored_name,
                    file.mimetype or "application/octet-stream",
                    source_field,
                    str(destination),
                    current_time(),
                ),
            )
            saved_names.append(file.filename)
        return saved_names

    def normalize_bug_form(form, files) -> dict:
        severity = normalize_bug_severity_value(form.get("severity", ""), form.get("priority", ""))
        platform = form.get("platform", "").strip()
        return {
            "title": form.get("title", "").strip(),
            "version": form.get("version", "").strip(),
            "module": bug_notify_key_for_platform(platform) if platform else form.get("module", "").strip(),
            "platform": platform,
            "severity": severity,
            "priority": severity,
            "assignee_id": form.get("assignee_id", "").strip(),
            "requirement_id": form.get("requirement_id", "").strip(),
            "case_id": form.get("case_id", "").strip(),
            "environment": form.get("environment", "").strip(),
            "description": form.get("description", "").strip(),
            "expected_result": form.get("expected_result", "").strip(),
            "actual_result": form.get("actual_result", "").strip(),
            "attachments": files.getlist("attachments") if files else [],
            "inline_images": files.getlist("inline_images") if files else [],
            "inline_image_sources": form.getlist("inline_image_sources") if form else [],
        }

    def sync_users() -> None:
        db = get_db()
        existing = db.execute("SELECT id, name, role, role_code, account_type, username, password, email FROM users").fetchall()
        if not existing:
            now = current_time()
            for profile in SAMPLE_USER_PROFILES:
                db.execute(
                    """
                    INSERT INTO users (name, role, role_code, account_type, username, password, email, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        profile["name"],
                        profile["role"],
                        profile["role_code"],
                        profile["account_type"],
                        profile["username"],
                        profile["password"],
                        profile["email"],
                        now,
                    ),
                )
            db.commit()
            return

        existing_by_username = {row["username"]: row for row in existing if row["username"]}
        now = current_time()
        for profile in SAMPLE_USER_PROFILES:
            user = existing_by_username.get(profile["username"])
            if user:
                db.execute(
                    """
                    UPDATE users
                    SET
                        name = COALESCE(NULLIF(name, ''), ?),
                        role = ?,
                        role_code = ?,
                        account_type = COALESCE(NULLIF(account_type, ''), ?),
                        password = COALESCE(NULLIF(password, ''), ?),
                        email = COALESCE(NULLIF(email, ''), ?)
                    WHERE id = ?
                    """,
                    (
                        profile["name"],
                        profile["role"],
                        profile["role_code"],
                        profile["account_type"],
                        profile["password"],
                        profile["email"],
                        user["id"],
                    ),
                )
                continue
            if profile["username"] != "admin":
                continue
            db.execute(
                """
                INSERT INTO users (name, role, role_code, account_type, username, password, email, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    profile["name"],
                    profile["role"],
                    profile["role_code"],
                    profile["account_type"],
                    profile["username"],
                    profile["password"],
                    profile["email"],
                    now,
                ),
            )
        db.commit()

    def seed_projects() -> None:
        db = get_db()
        if db.execute("SELECT COUNT(*) FROM projects").fetchone()[0] == 0:
            now = current_time()
            db.executemany(
                "INSERT INTO projects (name, description, created_at) VALUES (?, ?, ?)",
                [(name, desc, now) for name, desc in SAMPLE_PROJECTS],
            )
            db.commit()

    def seed_requirements() -> None:
        db = get_db()
        if db.execute("SELECT COUNT(*) FROM requirements").fetchone()[0] == 0:
            now = current_time()
            for project_name, code, title in SAMPLE_REQUIREMENTS:
                project = fetch_project_by_name(project_name)
                if project is None:
                    continue
                db.execute(
                    """
                    INSERT INTO requirements (
                        project_id, code, title, version, status, priority, description,
                        acceptance_criteria, creator_id, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project["id"],
                        code,
                        title,
                        "",
                        "pending",
                        "中",
                        f"{title}相关需求内容待补充。",
                        f"{title}上线前需补充验收标准。",
                        6,
                        now,
                        now,
                    ),
                )
            db.commit()

    def seed_cases() -> None:
        db = get_db()
        if db.execute("SELECT COUNT(*) FROM test_cases").fetchone()[0] == 0:
            now = current_time()
            for item in SAMPLE_CASES:
                project = fetch_project_by_name(item["project_name"])
                if project is None:
                    continue
                db.execute(
                    """
                    INSERT INTO test_cases (
                        project_id, version, folder_name, doc_name, case_no, title, priority_level, module_name,
                        steps, expected_result, ios_result, android_result, h5_result, remark,
                        source_type, doc_link, execute_status, creator_id, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project["id"],
                        item["doc_name"].split("-")[0] if "-" in item["doc_name"] else "",
                        item["folder_name"],
                        item["doc_name"],
                        item["case_no"],
                        item["title"],
                        item["priority_level"],
                        item["module_name"],
                        item["steps"],
                        item["expected_result"],
                        item["ios_result"],
                        item["android_result"],
                        item["h5_result"],
                        item["remark"],
                        item["source_type"],
                        "在线文档",
                        item["execute_status"],
                        7,
                        now,
                        now,
                    ),
                )
            db.commit()

    def seed_bugs() -> None:
        db = get_db()
        if db.execute("SELECT COUNT(*) FROM bugs").fetchone()[0] == 0:
            for bug in SAMPLE_BUGS:
                project = fetch_project_by_name(bug["project_name"])
                if project is None:
                    continue
                requirement = fetch_requirement_by_code(bug["requirement_code"])
                case = fetch_case_by_no(bug["case_code"])
                bug_id = insert_bug(
                    db=db,
                    title=bug["title"],
                    project_id=project["id"],
                    version=bug["version"],
                    module=bug["module"],
                    platform=bug.get("platform", ""),
                    severity=bug["severity"],
                    priority=bug["priority"],
                    status=bug["status"],
                    assignee_id=find_user_id(db, bug["assignee_name"]),
                    creator_id=find_user_id(db, bug["creator_name"]),
                    previous_assignee_id=find_user_id(db, bug["previous_assignee_name"]),
                    requirement_id=requirement["id"] if requirement else None,
                    case_id=case["id"] if case else None,
                    environment=bug["environment"],
                    description=bug["description"],
                    expected_result=bug["expected_result"],
                    actual_result=bug["actual_result"],
                    resolution_note=bug["resolution_note"],
                )
                add_history(
                    db,
                    bug_id,
                    "初始化",
                    f"创建缺陷，当前状态为 {STATUS_LABELS[bug['status']]}",
                    bug["creator_name"],
                    environment_snapshot=bug["environment"],
                    status_snapshot=bug["status"],
                    assignee_snapshot=bug["assignee_name"],
                )
            db.commit()

    def seed_data() -> None:
        sync_users()
        seed_projects()
        seed_requirements()
        seed_cases()
        seed_bugs()

    def current_project_id() -> int | None:
        project_id = session.get("project_id")
        if project_id:
            return int(project_id)
        first_project = get_db().execute("SELECT id FROM projects ORDER BY id LIMIT 1").fetchone()
        return int(first_project["id"]) if first_project else None

    def default_project_id_for_user(user_id: int) -> int | None:
        # 登录后优先进入有待办的项目，避免用户误以为待办数据丢失。
        todo_project = get_db().execute(
            f"""
            SELECT project_id
            FROM bugs
            WHERE assignee_id = ?
              AND status IN ({','.join('?' for _ in TODO_STATUS_CODES)})
            GROUP BY project_id
            ORDER BY MAX(datetime(updated_at)) DESC, project_id ASC
            LIMIT 1
            """,
            (user_id, *TODO_STATUS_CODES),
        ).fetchone()
        if todo_project is not None:
            return int(todo_project["project_id"])
        first_project = get_db().execute("SELECT id FROM projects ORDER BY id LIMIT 1").fetchone()
        return int(first_project["id"]) if first_project else None

    def set_current_project(project_id: int) -> None:
        session["project_id"] = project_id

    def fetch_current_project() -> sqlite3.Row | None:
        project_id = current_project_id()
        if project_id is None:
            return None
        return fetch_project(project_id)

    def clean_filter_values(raw_values: object) -> list[str]:
        if raw_values is None:
            return []
        if isinstance(raw_values, str):
            source_values = [raw_values]
        elif isinstance(raw_values, (list, tuple, set)):
            source_values = list(raw_values)
        else:
            source_values = [raw_values]
        values: list[str] = []
        for raw_value in source_values:
            for part in str(raw_value or "").split(","):
                value = part.strip()
                if value and value not in values:
                    values.append(value)
        return values

    def filter_values(filters: dict, key: str) -> list[str]:
        return clean_filter_values(filters.get(f"{key}_values") or filters.get(key, ""))

    def add_multi_filter_clause(clauses: list[str], params: list[str], expression: str, values: list[str]) -> None:
        if not values:
            return
        placeholders = ", ".join("?" for _ in values)
        clauses.append(f"{expression} IN ({placeholders})")
        params.extend(values)

    def build_bug_where(filters: dict) -> tuple[str, list[str]]:
        clauses = ["bugs.project_id = ?"]
        params: list[str] = [str(current_project_id() or 0)]
        add_multi_filter_clause(clauses, params, "COALESCE(bugs.version, '')", filter_values(filters, "version"))
        add_multi_filter_clause(clauses, params, "COALESCE(bugs.platform, '')", filter_values(filters, "platform"))
        add_multi_filter_clause(clauses, params, "bugs.creator_id", filter_values(filters, "creator_id"))
        add_multi_filter_clause(clauses, params, "bugs.assignee_id", filter_values(filters, "assignee_id"))
        add_multi_filter_clause(clauses, params, "bugs.status", filter_values(filters, "status"))
        if filters.get("created_from"):
            clauses.append("date(bugs.created_at) >= date(?)")
            params.append(filters["created_from"])
        if filters.get("created_to"):
            clauses.append("date(bugs.created_at) <= date(?)")
            params.append(filters["created_to"])
        if filters.get("keyword"):
            keyword = f"%{filters['keyword']}%"
            normalized_keyword = filters["keyword"].strip()
            padded_keyword = normalized_keyword.zfill(3) if normalized_keyword.isdigit() else normalized_keyword
            clauses.append("(bugs.bug_no LIKE ? OR bugs.title LIKE ? OR creator.name LIKE ? OR assignee.name LIKE ?)")
            params.extend([f"%{padded_keyword}%", keyword, keyword, keyword])
        return " AND ".join(clauses), params

    def fetch_filters() -> dict:
        filters = {
            "version": request.args.get("version", "").strip(),
            "platform": request.args.get("platform", "").strip(),
            "creator_id": request.args.get("creator_id", "").strip(),
            "assignee_id": request.args.get("assignee_id", "").strip(),
            "status": request.args.get("status", "").strip(),
            "created_from": request.args.get("created_from", "").strip(),
            "created_to": request.args.get("created_to", "").strip(),
            "keyword": request.args.get("keyword", "").strip(),
        }
        for key in BUG_MULTI_FILTER_KEYS:
            values = clean_filter_values(request.args.getlist(key))
            filters[f"{key}_values"] = values
            filters[key] = values[0] if len(values) == 1 else ""
        return filters

    def fetch_bug_versions(project_id: int | None = None) -> list[str]:
        target_project_id = project_id or current_project_id()
        if target_project_id is None:
            return []
        rows = get_db().execute(
            """
            SELECT DISTINCT version
            FROM bugs
            WHERE project_id = ? AND COALESCE(version, '') <> ''
            ORDER BY version DESC
            """,
            (target_project_id,),
        ).fetchall()
        return [row["version"] for row in rows]

    def build_pagination_items(page: int, pages: int, side_count: int = 1) -> list[dict[str, object]]:
        if pages <= 1:
            return [{"type": "page", "page": 1, "current": True}]
        page_numbers = {1, pages}
        page_numbers.update(range(max(1, page - side_count), min(pages, page + side_count) + 1))
        if page <= 3:
            page_numbers.update(range(1, min(pages, 4) + 1))
        if page >= pages - 2:
            page_numbers.update(range(max(1, pages - 3), pages + 1))

        items: list[dict[str, object]] = []
        previous = 0
        for page_number in sorted(page_numbers):
            if previous and page_number - previous == 2:
                items.append({"type": "page", "page": previous + 1, "current": False})
            elif previous and page_number - previous > 2:
                items.append({"type": "ellipsis"})
            items.append({"type": "page", "page": page_number, "current": page_number == page})
            previous = page_number
        return items

    def request_page(default: int = 1) -> int:
        try:
            return max(1, int(request.args.get("page", str(default)) or default))
        except (TypeError, ValueError):
            return default

    def fetch_bug_page(filters: dict, page: int) -> dict:
        db = get_db()
        where_sql, params = build_bug_where(filters)
        total = db.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM bugs
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["count"]
        page_size = app.config["BUG_PAGE_SIZE"]
        pages = max(1, math.ceil(total / page_size)) if total else 1
        page = max(1, min(page, pages))
        offset = (page - 1) * page_size
        items = db.execute(
            f"""
            SELECT
                bugs.*,
                projects.name AS project_name,
                creator.name AS creator_name,
                assignee.name AS assignee_name,
                requirements.code AS requirement_code,
                test_cases.case_no AS case_no
            FROM bugs
            JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            LEFT JOIN requirements ON bugs.requirement_id = requirements.id
            LEFT JOIN test_cases ON bugs.case_id = test_cases.id
            WHERE {where_sql}
            ORDER BY bugs.created_at DESC, bugs.id DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()
        return {
            "items": items,
            "total": total,
            "page": page,
            "pages": pages,
            "start_index": offset + 1 if total else 0,
            "end_index": min(offset + page_size, total),
            "has_prev": page > 1,
            "has_next": page < pages,
            "page_items": build_pagination_items(page, pages),
        }

    def fetch_bug_summary(filters: dict) -> dict:
        db = get_db()
        where_sql, params = build_bug_where(filters)
        summary = db.execute(
            f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN bugs.status IN ('open', 'in_progress') THEN 1 ELSE 0 END) AS active_count,
                SUM(CASE WHEN bugs.status = 'pending_verification' THEN 1 ELSE 0 END) AS verification_count,
                SUM(CASE WHEN bugs.status IN ('closed', 'duplicate', 'on_hold') THEN 1 ELSE 0 END) AS closed_count
            FROM bugs
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            WHERE {where_sql}
            """,
            params,
        ).fetchone()
        return {
            "total": int(summary["total"] or 0),
            "active_count": int(summary["active_count"] or 0),
            "verification_count": int(summary["verification_count"] or 0),
            "closed_count": int(summary["closed_count"] or 0),
        }

    def fetch_case_page(page: int) -> dict:
        db = get_db()
        project_id = current_project_id()
        total = db.execute("SELECT COUNT(*) AS count FROM test_cases WHERE project_id = ?", (project_id,)).fetchone()["count"]
        page_size = app.config["CASE_PAGE_SIZE"]
        pages = max(1, math.ceil(total / page_size)) if total else 1
        page = max(1, min(page, pages))
        offset = (page - 1) * page_size
        items = db.execute(
            """
            SELECT * FROM test_cases
            WHERE project_id = ?
            ORDER BY id DESC
            LIMIT ? OFFSET ?
            """,
            (project_id, page_size, offset),
        ).fetchall()
        return {
            "items": items,
            "page": page,
            "pages": pages,
            "total": total,
            "start_index": offset + 1 if total else 0,
            "end_index": min(offset + page_size, total),
        }

    def normalize_case_status(ios_result: str, android_result: str, h5_result: str) -> str:
        values = [value for value in [ios_result, android_result, h5_result] if value]
        if not values:
            return "未测"
        if "failed" in values:
            return "失败"
        if "block" in values:
            return "受阻"
        if all(value == "skip" for value in values):
            return "跳过"
        if any(value == "pass" for value in values):
            return "通过"
        return "未测"

    def sync_case_execute_statuses() -> int:
        rows = get_db().execute(
            """
            SELECT id, ios_result, android_result, h5_result, execute_status
            FROM test_cases
            WHERE COALESCE(ios_result, '') <> ''
                OR COALESCE(android_result, '') <> ''
                OR COALESCE(h5_result, '') <> ''
            """
        ).fetchall()
        updates: list[tuple[str, int]] = []
        for row in rows:
            next_status = normalize_case_status(
                row["ios_result"] or "",
                row["android_result"] or "",
                row["h5_result"] or "",
            )
            if next_status != (row["execute_status"] or ""):
                updates.append((next_status, int(row["id"])))
        if updates:
            get_db().executemany(
                "UPDATE test_cases SET execute_status = ? WHERE id = ?",
                updates,
            )
            get_db().commit()
        return len(updates)

    def normalize_excel_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text.lower() == "none":
            return ""
        return text

    def excel_image_anchor_cell(image: object) -> tuple[int, int] | None:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is not None:
            row_index = int(getattr(marker, "row", -1)) + 1
            column_index = int(getattr(marker, "col", -1))
            if row_index > 0 and column_index >= 0:
                return row_index, column_index
        if isinstance(anchor, str):
            try:
                row_index, column_index = openpyxl.utils.cell.coordinate_to_tuple(anchor)
            except ValueError:
                return None
            return row_index, column_index - 1
        return None

    def collect_sheet_image_cells(sheet) -> dict[tuple[int, int], int]:
        image_cells: dict[tuple[int, int], int] = {}
        for image in getattr(sheet, "_images", []) or []:
            cell = excel_image_anchor_cell(image)
            if cell is None:
                continue
            image_cells[cell] = image_cells.get(cell, 0) + 1
        return image_cells

    def count_excel_images_in_row(image_cells: dict[tuple[int, int], int], row_index: int) -> int:
        return sum(count for (image_row, _column_index), count in image_cells.items() if image_row == row_index)

    def count_excel_images_in_cell(image_cells: dict[tuple[int, int], int], row_index: int, column_index: int | None) -> int:
        if column_index is None:
            return 0
        return image_cells.get((row_index, column_index), 0)

    def append_excel_image_marker(text: str, image_count: int) -> str:
        if image_count <= 0:
            return text
        marker = "[原Excel含图片]" if image_count == 1 else f"[原Excel含{image_count}张图片]"
        if marker in text:
            return text
        return f"{text}\n{marker}" if text else marker

    def normalize_header_key(value: object) -> str:
        text = normalize_excel_text(value)
        return "".join(char.lower() for char in text if char not in " \t\r\n_-/:：()[]【】")

    def looks_like_priority_level(value: object) -> bool:
        text = normalize_excel_text(value).upper().replace(" ", "")
        if not text:
            return False
        if text in {"高", "中", "低", "最高", "最低", "建议"}:
            return True
        return text.startswith("P") and text[1:].isdigit()

    def looks_like_step_text(value: object) -> bool:
        text = normalize_excel_text(value)
        if not text:
            return False
        compact = text.replace(" ", "")
        return any(
            marker in compact
            for marker in [
                "前置条件",
                "前提条件",
                "【前置】",
                "点击",
                "进入",
                "打开",
                "查看",
                "长按",
                "刷新页面",
                "1、",
                "1.",
            ]
        ) or ("\n" in text and any(char.isdigit() for char in text[:4]))

    def find_excel_header_index(sheet) -> tuple[int | None, dict[str, int]]:
        header_aliases = {
            "case_no": {"用例编号", "测试编号", "编号", "caseid", "case_no", "caseno", "testcaseid", "测试用例编号"},
            "title": {"用例标题", "标题", "测试标题", "用例名称", "名称", "title", "casetitle", "testcasetitle", "测试点", "功能点", "验证点", "测试项", "测试目标", "场景"},
            "priority_level": {"优先级", "优先级别", "priority", "prioritylevel", "优先级p"},
            "module_name": {"所属模块", "模块", "功能模块", "module", "modulename", "一级模块", "二级模块", "业务模块", "功能模块名称"},
            "steps": {"测试步骤", "步骤", "操作步骤", "step", "steps", "前置条件", "前提条件", "操作内容"},
            "expected_result": {"预期结果", "预期", "expected", "expectedresult"},
            "actual_result": {"实际结果", "实际", "actual", "actualresult"},
            "execute_status": {"执行结果", "执行状态", "结果", "状态", "executestatus", "status", "result"},
            "ios_result": {"ios", "ios结果", "ios执行结果", "苹果结果"},
            "android_result": {"android", "android结果", "android执行结果"},
            "h5_result": {"h5", "h5结果", "h5执行结果", "web结果"},
            "remark": {"备注", "说明", "remark", "note", "comment"},
            "version": {"版本", "version"},
            "executor": {"执行人", "测试人", "负责人", "执行者", "executor", "tester"},
        }
        normalized_aliases = {
            key: {normalize_header_key(item) for item in values}
            for key, values in header_aliases.items()
        }
        preview_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 24), values_only=True))
        best_row_index = None
        best_mapping: dict[str, int] = {}
        best_score = -1

        for row_index, row in enumerate(preview_rows, start=1):
            next_row = preview_rows[row_index] if row_index < len(preview_rows) else ()
            mapping: dict[str, int] = {}
            max_cols = max(len(row), len(next_row))
            for col_index in range(max_cols):
                cell_value = row[col_index] if col_index < len(row) else ""
                next_value = next_row[col_index] if col_index < len(next_row) else ""
                next_header_key = normalize_header_key(next_value)
                if next_header_key in normalized_aliases["ios_result"] and "ios_result" not in mapping:
                    mapping["ios_result"] = col_index
                    continue
                if next_header_key in normalized_aliases["android_result"] and "android_result" not in mapping:
                    mapping["android_result"] = col_index
                    continue
                if next_header_key in normalized_aliases["h5_result"] and "h5_result" not in mapping:
                    mapping["h5_result"] = col_index
                    continue
                header_candidates: list[str] = []
                for candidate in (
                    cell_value,
                    next_value,
                    f"{normalize_excel_text(cell_value)}{normalize_excel_text(next_value)}" if normalize_excel_text(cell_value) or normalize_excel_text(next_value) else "",
                    f"{normalize_excel_text(next_value)}{normalize_excel_text(cell_value)}" if normalize_excel_text(cell_value) or normalize_excel_text(next_value) else "",
                ):
                    header_key = normalize_header_key(candidate)
                    if header_key and header_key not in header_candidates:
                        header_candidates.append(header_key)
                if not header_candidates:
                    continue
                for field_name, aliases in normalized_aliases.items():
                    if field_name in mapping:
                        continue
                    if any(header_key in aliases for header_key in header_candidates):
                        mapping[field_name] = col_index
                        break
            if "case_no" not in mapping:
                continue
            major_keys = {
                "title",
                "priority_level",
                "module_name",
                "steps",
                "expected_result",
                "actual_result",
                "ios_result",
                "android_result",
                "h5_result",
                "remark",
                "version",
                "executor",
            }
            major_count = sum(1 for key in major_keys if key in mapping)
            if major_count < 2 and not {"steps", "expected_result"}.intersection(mapping.keys()):
                continue
            score = major_count
            if "steps" in mapping:
                score += 3
            if "expected_result" in mapping:
                score += 3
            if "module_name" in mapping:
                score += 2
            if "priority_level" in mapping:
                score += 1
            if score > best_score:
                best_score = score
                best_row_index = row_index
                best_mapping = mapping

        if best_row_index is None or "case_no" not in best_mapping:
            return None, {}
        return best_row_index, best_mapping

    def normalize_platform_result(raw_status: str) -> str:
        text = normalize_excel_text(raw_status).replace(" ", "").lower()
        result_map = {
            "pass": "pass",
            "passed": "pass",
            "通过": "pass",
            "成功": "pass",
            "ok": "pass",
            "√": "pass",
            "true": "pass",
            "yes": "pass",
            "fail": "failed",
            "failed": "failed",
            "失败": "failed",
            "x": "failed",
            "×": "failed",
            "false": "failed",
            "no": "failed",
            "block": "block",
            "blocked": "block",
            "受阻": "block",
            "skip": "skip",
            "skipped": "skip",
            "跳过": "skip",
        }
        return result_map.get(text, "")

    def extract_sheet_meta_info(sheet) -> dict[str, str]:
        preview_rows = min(sheet.max_row, 5)
        preview_cols = min(sheet.max_column, 6)
        lines: list[str] = []
        for row in sheet.iter_rows(min_row=1, max_row=preview_rows, max_col=preview_cols, values_only=True):
            line = " ".join(filter(None, (normalize_excel_text(value) for value in row)))
            if line:
                lines.append(line)
        return parse_case_meta_info("\n".join(lines))

    def infer_case_version(case_no: str, version: str = "") -> str:
        version_text = normalize_excel_text(version)
        if version_text:
            return version_text
        case_no_text = normalize_excel_text(case_no)
        if "-" in case_no_text:
            prefix = case_no_text.split("-", 1)[0].strip()
            if "." in prefix:
                return prefix
        return ""

    def normalize_case_execute_status(raw_status: str) -> tuple[str, str, str, str]:
        text = normalize_excel_text(raw_status).replace(" ", "").lower()
        status_map = {
            "": ("未测", "", "", ""),
            "未测": ("未测", "", "", ""),
            "notrun": ("未测", "", "", ""),
            "norun": ("未测", "", "", ""),
            "pass": ("通过", "pass", "", ""),
            "passed": ("通过", "pass", "", ""),
            "通过": ("通过", "pass", "", ""),
            "成功": ("通过", "pass", "", ""),
            "ok": ("通过", "pass", "", ""),
            "fail": ("失败", "failed", "", ""),
            "failed": ("失败", "failed", "", ""),
            "失败": ("失败", "failed", "", ""),
            "block": ("受阻", "block", "", ""),
            "blocked": ("受阻", "block", "", ""),
            "受阻": ("受阻", "block", "", ""),
            "skip": ("跳过", "skip", "", ""),
            "skipped": ("跳过", "skip", "", ""),
            "跳过": ("跳过", "skip", "", ""),
        }
        if text in status_map:
            return status_map[text]
        return normalize_excel_text(raw_status) or "未测", "", "", ""

    def has_meaningful_case_content(
        *,
        title: str,
        module_name: str,
        steps: str,
        expected_result: str,
        actual_result: str,
        remark: str,
        executor: str,
        execute_status: str,
        ios_result: str,
        android_result: str,
        h5_result: str,
    ) -> bool:
        return any(
            [
                title,
                module_name,
                steps,
                expected_result,
                actual_result,
                remark,
                executor,
                execute_status,
                ios_result,
                android_result,
                h5_result,
            ]
        )

    def is_sparse_imported_case_row(
        *,
        case_no: str,
        title: str,
        module_name: str,
        steps: str,
        expected_result: str,
        actual_result: str,
        remark: str,
        executor: str,
        ios_result: str,
        android_result: str,
        h5_result: str,
    ) -> bool:
        if not case_no:
            return False
        if any([steps, expected_result, actual_result, remark, executor, ios_result, android_result, h5_result]):
            return False
        title_text = normalize_excel_text(title)
        module_text = normalize_excel_text(module_name)
        if not title_text and not module_text:
            return True
        if title_text and title_text == case_no:
            return True
        return bool(title_text and module_text and title_text == module_text)

    def repair_misaligned_excel_cases(db: sqlite3.Connection, doc_names: list[str] | None = None) -> int:
        query = """
            SELECT
                id, doc_name, case_no, title, priority_level, module_name, steps, expected_result, actual_result,
                remark, executor, ios_result, android_result, h5_result
            FROM test_cases
            WHERE source_type = 'Excel上传'
        """
        params: list[str] = []
        if doc_names:
            valid_doc_names = [name for name in doc_names if name]
            if valid_doc_names:
                placeholders = ",".join("?" for _ in valid_doc_names)
                query += f" AND doc_name IN ({placeholders})"
                params.extend(valid_doc_names)
        query += " ORDER BY doc_name ASC, id ASC"
        rows = db.execute(query, params).fetchall()

        repaired = 0
        current_doc_name = ""
        current_module_name = ""

        for row in rows:
            doc_name = normalize_excel_text(row["doc_name"])
            if doc_name != current_doc_name:
                current_doc_name = doc_name
                current_module_name = ""

            title = normalize_excel_text(row["title"])
            priority_level = normalize_excel_text(row["priority_level"])
            module_name = normalize_excel_text(row["module_name"])
            steps = normalize_excel_text(row["steps"])
            expected_result = normalize_excel_text(row["expected_result"])
            actual_result = normalize_excel_text(row["actual_result"])
            remark = normalize_excel_text(row["remark"])
            executor = normalize_excel_text(row["executor"])
            ios_result = normalize_excel_text(row["ios_result"])
            android_result = normalize_excel_text(row["android_result"])
            h5_result = normalize_excel_text(row["h5_result"])

            if is_sparse_imported_case_row(
                case_no=normalize_excel_text(row["case_no"]),
                title=title,
                module_name=module_name,
                steps=steps,
                expected_result=expected_result,
                actual_result=actual_result,
                remark=remark,
                executor=executor,
                ios_result=ios_result,
                android_result=android_result,
                h5_result=h5_result,
            ):
                db.execute("DELETE FROM test_cases WHERE id = ?", (row["id"],))
                repaired += 1
                continue

            if title and module_name and title == module_name and steps and normalize_platform_result(expected_result):
                if priority_level and not looks_like_priority_level(priority_level):
                    current_module_name = priority_level
                fixed_module_name = current_module_name or (priority_level if not looks_like_priority_level(priority_level) else "") or module_name
                fixed_priority_level = priority_level if looks_like_priority_level(priority_level) else "P1"
                fixed_actual_result = "" if normalize_platform_result(actual_result) else actual_result
                db.execute(
                    """
                    UPDATE test_cases
                    SET title = ?, priority_level = ?, module_name = ?, steps = ?, expected_result = ?, actual_result = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        fixed_module_name or row["case_no"],
                        fixed_priority_level,
                        fixed_module_name,
                        title,
                        steps,
                        fixed_actual_result,
                        current_time(),
                        row["id"],
                    ),
                )
                repaired += 1
                continue

            if module_name and not looks_like_step_text(module_name) and not looks_like_priority_level(module_name):
                current_module_name = module_name
            elif priority_level and not looks_like_priority_level(priority_level) and not looks_like_step_text(priority_level):
                current_module_name = priority_level

        return repaired

    def parse_case_meta_info(raw_text: str) -> dict[str, str]:
        meta = {"environment_info": "", "device_info": "", "network_info": ""}
        for line in raw_text.splitlines():
            text = normalize_excel_text(line)
            if not text:
                continue
            if "测试环境" in text and "：" in text:
                meta["environment_info"] = text.split("：", 1)[1].strip()
            elif "测试设备" in text and "：" in text:
                meta["device_info"] = text.split("：", 1)[1].strip()
            elif "网络环境" in text and "：" in text:
                meta["network_info"] = text.split("：", 1)[1].strip()
        return meta

    def fetch_case_versions(project_id: int | None = None) -> list[str]:
        target_project_id = project_id or current_project_id()
        if target_project_id is None:
            return []
        rows = get_db().execute(
            """
            SELECT DISTINCT version
            FROM test_cases
            WHERE project_id = ? AND COALESCE(version, '') <> ''
            ORDER BY version DESC
            """,
            (target_project_id,),
        ).fetchall()
        return [row["version"] for row in rows]

    def fetch_report_versions(project_id: int | None = None) -> list[str]:
        values = set(fetch_bug_versions(project_id=project_id))
        values.update(fetch_case_versions(project_id=project_id))
        return sorted((value for value in values if value), reverse=True)

    def fetch_document_dynamic_columns(project_id: int, version: str, folder_name: str, doc_name: str) -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT *
            FROM case_document_columns
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            ORDER BY sort_order ASC, id ASC
            """,
            (project_id, version, folder_name, doc_name),
        ).fetchall()

    def fetch_document_dynamic_cell_map(case_ids: list[int], column_ids: list[int]) -> dict[tuple[int, int], str]:
        valid_case_ids = [case_id for case_id in case_ids if case_id > 0]
        valid_column_ids = [column_id for column_id in column_ids if column_id > 0]
        if not valid_case_ids or not valid_column_ids:
            return {}
        case_placeholders = ",".join("?" for _ in valid_case_ids)
        column_placeholders = ",".join("?" for _ in valid_column_ids)
        rows = get_db().execute(
            f"""
            SELECT column_id, case_id, cell_value
            FROM case_document_cells
            WHERE case_id IN ({case_placeholders})
                AND column_id IN ({column_placeholders})
            """,
            [*valid_case_ids, *valid_column_ids],
        ).fetchall()
        return {
            (int(row["case_id"]), int(row["column_id"])): str(row["cell_value"] or "")
            for row in rows
        }

    def suggest_next_document_case_no(document: sqlite3.Row | dict[str, Any]) -> str:
        last_rows = get_db().execute(
            """
            SELECT case_no
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            ORDER BY id DESC
            """,
            (
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
            ),
        ).fetchall()
        for row in last_rows:
            case_no = str(row["case_no"] or "").strip()
            match = re.match(r"^(.*?)(\d+)$", case_no)
            if match:
                prefix, number = match.groups()
                return f"{prefix}{str(int(number) + 1).zfill(len(number))}"
        version = str(document["version"] or "").strip()
        if version:
            return f"{version}-TC-001"
        next_index = count_document_cases(
            int(document["project_id"]),
            str(document["version"] or ""),
            str(document["folder_name"] or ""),
            str(document["doc_name"] or ""),
        ) + 1
        return f"TC-{str(next_index).zfill(3)}"

    def create_case_document_row(document: sqlite3.Row | dict[str, Any], db: sqlite3.Connection | None = None) -> int:
        target_db = db or get_db()
        now = current_time()
        cursor = target_db.execute(
            """
            INSERT INTO test_cases (
                project_id, version, folder_name, doc_name, case_no, title, priority_level, module_name,
                steps, expected_result, ios_result, android_result, h5_result, remark, executor,
                environment_info, device_info, network_info, source_type, doc_link, execute_status,
                creator_id, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
                suggest_next_document_case_no(document),
                "",
                "P1",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                document["environment_info"] or "",
                document["device_info"] or "",
                document["network_info"] or "",
                document["source_type"] or "在线文档",
                document["doc_link"] or "",
                "未测",
                int(g.current_user["id"]) if g.current_user is not None else None,
                now,
                now,
            ),
        )
        return int(cursor.lastrowid)

    def make_unique_document_column_name(
        document: sqlite3.Row | dict[str, Any],
        column_name: str,
        db: sqlite3.Connection | None = None,
    ) -> str:
        target_db = db or get_db()
        desired = column_name.strip()
        existing_names = {
            str(row["column_name"] or "").strip().lower()
            for row in target_db.execute(
                """
                SELECT column_name
                FROM case_document_columns
                WHERE project_id = ?
                    AND COALESCE(version, '') = COALESCE(?, '')
                    AND COALESCE(folder_name, '') = COALESCE(?, '')
                    AND COALESCE(doc_name, '') = COALESCE(?, '')
                """,
                (
                    document["project_id"],
                    document["version"],
                    document["folder_name"],
                    document["doc_name"],
                ),
            ).fetchall()
        }
        if desired.lower() not in existing_names:
            return desired
        suffix = 2
        while f"{desired}{suffix}".lower() in existing_names:
            suffix += 1
        return f"{desired}{suffix}"

    def create_case_document_column(
        document: sqlite3.Row | dict[str, Any],
        column_name: str,
        db: sqlite3.Connection | None = None,
    ) -> int:
        target_db = db or get_db()
        now = current_time()
        next_sort_order = int(
            target_db.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0) AS max_sort_order
                FROM case_document_columns
                WHERE project_id = ?
                    AND COALESCE(version, '') = COALESCE(?, '')
                    AND COALESCE(folder_name, '') = COALESCE(?, '')
                    AND COALESCE(doc_name, '') = COALESCE(?, '')
                """,
                (
                    document["project_id"],
                    document["version"],
                    document["folder_name"],
                    document["doc_name"],
                ),
            ).fetchone()["max_sort_order"]
            or 0
        )
        cursor = target_db.execute(
            """
            INSERT INTO case_document_columns (
                project_id, version, folder_name, doc_name, column_name, sort_order,
                creator_id, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
                make_unique_document_column_name(document, column_name, db=target_db),
                next_sort_order + 1,
                int(g.current_user["id"]) if g.current_user is not None else None,
                now,
                now,
            ),
        )
        return int(cursor.lastrowid)

    def fetch_case_documents(version: str = "") -> list[sqlite3.Row]:
        params: list[object] = [current_project_id()]
        where_parts = [
            "project_id = ?",
            "COALESCE(doc_name, '') <> ''",
        ]
        if version:
            where_parts.append("COALESCE(version, '') = ?")
            params.append(version)
        return get_db().execute(
            f"""
            SELECT
                MIN(id) AS id,
                version,
                folder_name,
                doc_name,
                MIN(creator_id) AS creator_id,
                COUNT(*) AS case_count,
                MAX(updated_at) AS updated_at
            FROM test_cases
            WHERE {' AND '.join(where_parts)}
            GROUP BY version, folder_name, doc_name
            ORDER BY updated_at DESC, doc_name ASC
            """,
            params,
        ).fetchall()

    def fetch_case_folders() -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT *
            FROM case_folders
            WHERE project_id = ?
            ORDER BY name ASC
            """,
            (current_project_id(),),
        ).fetchall()

    def create_case_folder(folder_name: str) -> bool:
        name = folder_name.strip()
        if not name:
            return False
        db = get_db()
        existing = db.execute(
            """
            SELECT 1
            FROM case_folders
            WHERE project_id = ? AND name = ?
            LIMIT 1
            """,
            (current_project_id(), name),
        ).fetchone()
        if existing:
            return False
        now = current_time()
        db.execute(
            """
            INSERT INTO case_folders (project_id, name, creator_id, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                current_project_id(),
                name,
                int(g.current_user["id"]) if g.current_user is not None else None,
                now,
                now,
            ),
        )
        db.commit()
        return True

    def ensure_case_folder(folder_name: str, db: sqlite3.Connection | None = None) -> None:
        name = folder_name.strip() or "测试用例"
        project_id = current_project_id()
        if project_id is None:
            return
        target_db = db or get_db()
        now = current_time()
        target_db.execute(
            """
            INSERT OR IGNORE INTO case_folders (project_id, name, creator_id, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                project_id,
                name,
                int(g.current_user["id"]) if g.current_user is not None else None,
                now,
                now,
            ),
        )

    def create_case_document(folder_name: str, doc_name: str) -> None:
        db = get_db()
        now = current_time()
        unique_case_no = f"DOC-{uuid.uuid4().hex[:8].upper()}"
        ensure_case_folder(folder_name, db)
        db.execute(
            """
            INSERT INTO test_cases (
                project_id, version, folder_name, doc_name, case_no, title, priority_level, module_name,
                steps, expected_result, ios_result, android_result, h5_result, remark,
                source_type, doc_link, execute_status, creator_id, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                current_project_id(),
                doc_name.split("-")[0] if "-" in doc_name else "",
                folder_name,
                doc_name,
                unique_case_no,
                "",
                "P1",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "在线文档",
                "",
                "未测",
                int(g.current_user["id"]) if g.current_user is not None else None,
                now,
                now,
            ),
        )
        db.commit()

    def rename_case_document(document_id: int, version: str, folder_name: str, doc_name: str) -> int:
        document = fetch_case_document(document_id)
        if document is None:
            return 0
        db = get_db()
        ensure_case_folder(folder_name, db)
        cursor = db.execute(
            """
            UPDATE test_cases
            SET version = ?, folder_name = ?, doc_name = ?, updated_at = ?
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (
                version,
                folder_name,
                doc_name,
                current_time(),
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
            ),
        )
        db.execute(
            """
            UPDATE case_document_columns
            SET version = ?, folder_name = ?, doc_name = ?, updated_at = ?
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (
                version,
                folder_name,
                doc_name,
                current_time(),
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
            ),
        )
        db.commit()
        return int(cursor.rowcount or 0)

    def move_case_document(document_id: int, folder_name: str) -> int:
        document = fetch_case_document(document_id)
        if document is None:
            return 0
        target_folder = folder_name.strip() or "测试用例"
        current_folder = document["folder_name"] or "测试用例"
        if target_folder == current_folder:
            return 0
        db = get_db()
        duplicate = db.execute(
            """
            SELECT 1
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            LIMIT 1
            """,
            (document["project_id"], document["version"], target_folder, document["doc_name"]),
        ).fetchone()
        if duplicate:
            return -1
        now = current_time()
        ensure_case_folder(target_folder, db)
        cursor = db.execute(
            """
            UPDATE test_cases
            SET folder_name = ?, updated_at = ?
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (
                target_folder,
                now,
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
            ),
        )
        db.execute(
            """
            UPDATE case_document_columns
            SET folder_name = ?, updated_at = ?
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (
                target_folder,
                now,
                document["project_id"],
                document["version"],
                document["folder_name"],
                document["doc_name"],
            ),
        )
        db.commit()
        return int(cursor.rowcount or 0)

    def delete_case_folder(folder_name: str) -> None:
        db = get_db()
        db.execute(
            """
            DELETE FROM case_folders
            WHERE project_id = ? AND name = ?
            """,
            (current_project_id(), folder_name),
        )
        db.execute(
            """
            DELETE FROM test_cases
            WHERE project_id = ? AND folder_name = ?
            """,
            (current_project_id(), folder_name),
        )
        db.commit()

    def delete_case_document(document_id: int) -> int:
        document = fetch_case_document(document_id)
        if document is None:
            return 0
        db = get_db()
        case_rows = db.execute(
            """
            SELECT id
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (document["project_id"], document["version"], document["folder_name"], document["doc_name"]),
        ).fetchall()
        case_ids = [int(row["id"]) for row in case_rows]
        if case_ids:
            placeholders = ",".join("?" for _ in case_ids)
            db.execute(f"DELETE FROM case_document_cells WHERE case_id IN ({placeholders})", case_ids)
        db.execute(
            """
            DELETE FROM case_document_columns
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (document["project_id"], document["version"], document["folder_name"], document["doc_name"]),
        )
        cursor = db.execute(
            """
            DELETE FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            """,
            (document["project_id"], document["version"], document["folder_name"], document["doc_name"]),
        )
        db.commit()
        return int(cursor.rowcount or 0)

    def delete_case_item(case_id: int) -> tuple[int, int | None, str]:
        case_item = fetch_case(case_id)
        if case_item is None:
            return 0, None, ""
        db = get_db()
        now = current_time()
        cursor = db.execute("DELETE FROM test_cases WHERE id = ?", (case_id,))
        db.execute("DELETE FROM case_document_cells WHERE case_id = ?", (case_id,))
        db.execute("UPDATE bugs SET case_id = NULL, updated_at = ? WHERE case_id = ?", (now, case_id))
        next_row = db.execute(
            """
            SELECT id
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            ORDER BY id ASC
            LIMIT 1
            """,
            (
                case_item["project_id"],
                case_item["version"],
                case_item["folder_name"],
                case_item["doc_name"],
            ),
        ).fetchone()
        db.commit()
        return int(cursor.rowcount or 0), int(next_row["id"]) if next_row is not None else None, str(case_item["version"] or "").strip()

    def save_case_document_dynamic_cells(
        *,
        document_columns: list[sqlite3.Row],
        document_cases: list[dict[str, object]],
        form,
        db: sqlite3.Connection,
        now: str,
    ) -> None:
        if not document_columns or not document_cases:
            return
        case_ids = [int(item["id"]) for item in document_cases]
        column_ids = [int(column["id"]) for column in document_columns]
        existing_cell_map = fetch_document_dynamic_cell_map(case_ids, column_ids)
        for item in document_cases:
            case_id = int(item["id"])
            for column in document_columns:
                column_id = int(column["id"])
                field_name = f"dynamic_{column_id}_{case_id}"
                if field_name not in form:
                    continue
                cell_value = form.get(field_name, "").strip()
                existing_value = existing_cell_map.get((case_id, column_id), "")
                if cell_value == existing_value:
                    continue
                if cell_value:
                    if (case_id, column_id) in existing_cell_map:
                        db.execute(
                            """
                            UPDATE case_document_cells
                            SET cell_value = ?, updated_at = ?
                            WHERE column_id = ? AND case_id = ?
                            """,
                            (cell_value, now, column_id, case_id),
                        )
                    else:
                        db.execute(
                            """
                            INSERT INTO case_document_cells (column_id, case_id, cell_value, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?)
                            """,
                            (column_id, case_id, cell_value, now, now),
                        )
                elif (case_id, column_id) in existing_cell_map:
                    db.execute(
                        "DELETE FROM case_document_cells WHERE column_id = ? AND case_id = ?",
                        (column_id, case_id),
                    )

    def case_document_form_value(form, field_name: str, case_id: int, fallback: object) -> str:
        input_name = f"{field_name}_{case_id}"
        if input_name not in form:
            return str(fallback or "")
        return str(form.get(input_name, "") or "").strip()

    def update_case_document_cell(
        *,
        bundle: dict,
        case_id: int,
        field_name: str,
        raw_value: object,
        db: sqlite3.Connection,
        now: str,
    ) -> dict[str, object]:
        case_map = {int(item["id"]): item for item in bundle["cases"]}
        case_item = case_map.get(case_id)
        if case_item is None:
            raise ValueError("未找到对应用例。")

        field = str(field_name or "").strip()
        value = str(raw_value or "").strip()
        base_fields = {
            "case_no",
            "priority_level",
            "module_name",
            "steps",
            "expected_result",
            "ios_result",
            "android_result",
            "h5_result",
            "remark",
            "executor",
        }
        result_fields = {"ios_result", "android_result", "h5_result"}

        if field in base_fields:
            if field == "case_no" and not value:
                raise ValueError("测试编号不能为空。")
            if field in result_fields and value not in PLATFORM_RESULT_OPTIONS:
                raise ValueError("请选择有效的执行结果。")

            set_parts = [f"{field} = ?"]
            params: list[object] = [value]
            execute_status = str(case_item["execute_status"] or "")
            if field in result_fields:
                ios_result = value if field == "ios_result" else str(case_item["ios_result"] or "")
                android_result = value if field == "android_result" else str(case_item["android_result"] or "")
                h5_result = value if field == "h5_result" else str(case_item["h5_result"] or "")
                execute_status = normalize_case_status(ios_result, android_result, h5_result)
                set_parts.append("execute_status = ?")
                params.append(execute_status)

            set_parts.append("updated_at = ?")
            params.extend([now, case_id])
            db.execute(
                f"UPDATE test_cases SET {', '.join(set_parts)} WHERE id = ?",
                params,
            )
            return {
                "field": field,
                "case_id": case_id,
                "value": value,
                "execute_status": execute_status,
                "saved_at": now,
            }

        dynamic_match = re.fullmatch(r"dynamic_(\d+)", field)
        if dynamic_match:
            column_id = int(dynamic_match.group(1))
            valid_column_ids = {int(column["id"]) for column in bundle["columns"]}
            if column_id not in valid_column_ids:
                raise ValueError("未找到对应自定义列。")
            existing = db.execute(
                """
                SELECT id
                FROM case_document_cells
                WHERE column_id = ? AND case_id = ?
                """,
                (column_id, case_id),
            ).fetchone()
            if value:
                if existing is None:
                    db.execute(
                        """
                        INSERT INTO case_document_cells (column_id, case_id, cell_value, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (column_id, case_id, value, now, now),
                    )
                else:
                    db.execute(
                        """
                        UPDATE case_document_cells
                        SET cell_value = ?, updated_at = ?
                        WHERE column_id = ? AND case_id = ?
                        """,
                        (value, now, column_id, case_id),
                    )
            elif existing is not None:
                db.execute(
                    "DELETE FROM case_document_cells WHERE column_id = ? AND case_id = ?",
                    (column_id, case_id),
                )
            db.execute("UPDATE test_cases SET updated_at = ? WHERE id = ?", (now, case_id))
            return {
                "field": field,
                "case_id": case_id,
                "value": value,
                "execute_status": str(case_item["execute_status"] or ""),
                "saved_at": now,
            }

        raise ValueError("该字段不支持实时保存。")

    def build_case_tree(documents: list[sqlite3.Row]) -> list[dict]:
        grouped: dict[str, list[sqlite3.Row]] = {}
        folder_names = {row["name"] for row in fetch_case_folders()}
        for item in documents:
            folder_name = item["folder_name"] or "测试用例"
            folder_names.add(folder_name)
            grouped.setdefault(folder_name, []).append(item)
        tree = []
        for folder_name in folder_names:
            tree.append({"name": folder_name, "documents": grouped.get(folder_name, [])})
        tree.sort(key=lambda item: item["name"])
        return tree

    def fetch_case_document(document_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT
                test_cases.*,
                users.name AS creator_name
            FROM test_cases
            LEFT JOIN users ON test_cases.creator_id = users.id
            WHERE test_cases.id = ?
            """,
            (document_id,),
        ).fetchone()

    def fetch_case_document_meta(project_id: int, version: str, folder_name: str, doc_name: str) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT *
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            ORDER BY id ASC
            LIMIT 1
            """,
            (project_id, version, folder_name, doc_name),
        ).fetchone()

    def fetch_document_cases(project_id: int, version: str, folder_name: str, doc_name: str) -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT *
            FROM test_cases
            WHERE project_id = ?
                AND COALESCE(version, '') = COALESCE(?, '')
                AND COALESCE(folder_name, '') = COALESCE(?, '')
                AND COALESCE(doc_name, '') = COALESCE(?, '')
            ORDER BY id ASC
            """,
            (project_id, version, folder_name, doc_name),
        ).fetchall()

    def fetch_bug_links_for_cases(project_id: int, case_ids: list[int]) -> dict[int, list[dict[str, object]]]:
        valid_case_ids = [case_id for case_id in case_ids if case_id > 0]
        if not valid_case_ids:
            return {}
        placeholders = ",".join("?" for _ in valid_case_ids)
        rows = get_db().execute(
            f"""
            SELECT id, case_id, bug_no, title
            FROM bugs
            WHERE project_id = ?
                AND case_id IN ({placeholders})
            ORDER BY CASE
                WHEN COALESCE(bug_no, '') <> '' AND bug_no GLOB '[0-9]*' THEN CAST(bug_no AS INTEGER)
                ELSE id
            END ASC, id ASC
            """,
            [project_id, *valid_case_ids],
        ).fetchall()
        case_bug_map: dict[int, list[dict[str, object]]] = {}
        for row in rows:
            case_bug_map.setdefault(int(row["case_id"]), []).append(
                {
                    "id": int(row["id"]),
                    "bug_no": str(row["bug_no"] or "").strip(),
                    "title": str(row["title"] or "").strip(),
                }
            )
        return case_bug_map

    def fetch_case_document_bundle(document_id: int) -> dict | None:
        document = fetch_case_document(document_id)
        if document is None:
            return None
        cases = fetch_document_cases(document["project_id"], document["version"], document["folder_name"], document["doc_name"])
        dynamic_columns = fetch_document_dynamic_columns(
            document["project_id"],
            document["version"],
            document["folder_name"],
            document["doc_name"],
        )
        dynamic_cell_map = fetch_document_dynamic_cell_map(
            [int(item["id"]) for item in cases],
            [int(column["id"]) for column in dynamic_columns],
        )
        case_bug_map = fetch_bug_links_for_cases(document["project_id"], [int(item["id"]) for item in cases])
        case_rows: list[dict[str, object]] = []
        for item in cases:
            row = dict(item)
            row["linked_bugs"] = case_bug_map.get(int(item["id"]), [])
            row["dynamic_values"] = {
                int(column["id"]): dynamic_cell_map.get((int(item["id"]), int(column["id"])), "")
                for column in dynamic_columns
            }
            case_rows.append(row)
        meta = fetch_case_document_meta(document["project_id"], document["version"], document["folder_name"], document["doc_name"])
        owner = g.current_user["name"] if g.current_user is not None else ""
        collaborators = fetch_users()[:4]
        return {
            "document": document,
            "cases": case_rows,
            "columns": dynamic_columns,
            "meta": meta,
            "owner": owner,
            "collaborators": collaborators,
        }

    def count_document_cases(project_id: int, version: str, folder_name: str, doc_name: str) -> int:
        return int(
            get_db()
            .execute(
                """
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = ?
                    AND COALESCE(version, '') = COALESCE(?, '')
                    AND COALESCE(folder_name, '') = COALESCE(?, '')
                    AND COALESCE(doc_name, '') = COALESCE(?, '')
                """,
                (project_id, version, folder_name, doc_name),
            )
            .fetchone()["count"]
        )

    def build_requirement_query(filters: dict | None = None) -> tuple[str, list[str | int]]:
        filters = filters or {}
        where_parts = ["requirements.project_id = ?"]
        params: list[str | int] = [current_project_id()]
        keyword = filters.get("keyword", "").strip()
        version = filters.get("version", "").strip()
        if keyword:
            where_parts.append(
                """
                (
                    requirements.code LIKE ?
                    OR requirements.title LIKE ?
                    OR COALESCE(requirements.requirement_doc_link, '') LIKE ?
                    OR COALESCE(requirements.design_doc_link, '') LIKE ?
                )
                """
            )
            keyword_like = f"%{keyword}%"
            params.extend([keyword_like, keyword_like, keyword_like, keyword_like])
        if version:
            where_parts.append("COALESCE(requirements.version, '') = ?")
            params.append(version)
        return " AND ".join(where_parts), params

    def build_requirement_display_name(requirement: sqlite3.Row | dict) -> str:
        version = str(requirement["version"] or "").strip()
        title = str(requirement["title"] or "").strip()
        if version and title:
            return f"{version} / {title}"
        return title or str(requirement["code"] or "")

    def generate_requirement_code(project_id: int) -> str:
        db = get_db()
        rows = db.execute("SELECT code FROM requirements WHERE project_id = ?", (project_id,)).fetchall()
        current_max = 0
        for row in rows:
            code = str(row["code"] or "").strip().upper()
            if code.startswith("REQ-"):
                suffix = code[4:]
                if suffix.isdigit():
                    current_max = max(current_max, int(suffix))
        return f"REQ-{str(current_max + 1).zfill(3)}"

    def fetch_requirements(filters: dict | None = None) -> list[sqlite3.Row]:
        where_sql, params = build_requirement_query(filters)
        return get_db().execute(
            f"""
            SELECT
                requirements.*,
                projects.name AS project_name,
                users.name AS creator_name,
                COUNT(bugs.id) AS linked_bug_count
            FROM requirements
            JOIN projects ON requirements.project_id = projects.id
            LEFT JOIN users ON requirements.creator_id = users.id
            LEFT JOIN bugs ON bugs.requirement_id = requirements.id
            WHERE {where_sql}
            GROUP BY requirements.id, projects.name, users.name
            ORDER BY datetime(requirements.created_at) DESC, requirements.id DESC
            """,
            params,
        ).fetchall()

    def fetch_requirement_versions() -> list[str]:
        rows = get_db().execute(
            """
            SELECT DISTINCT version
            FROM requirements
            WHERE project_id = ? AND COALESCE(version, '') <> ''
            ORDER BY version DESC
            """,
            (current_project_id(),),
        ).fetchall()
        return [row["version"] for row in rows]

    def fetch_requirement_summary() -> dict[str, int]:
        db = get_db()
        project_id = current_project_id()
        total = int(db.execute("SELECT COUNT(*) AS count FROM requirements WHERE project_id = ?", (project_id,)).fetchone()["count"])
        requirement_doc_count = int(
            db.execute(
                "SELECT COUNT(*) AS count FROM requirements WHERE project_id = ? AND COALESCE(requirement_doc_link, '') <> ''",
                (project_id,),
            ).fetchone()["count"]
        )
        design_doc_count = int(
            db.execute(
                "SELECT COUNT(*) AS count FROM requirements WHERE project_id = ? AND COALESCE(design_doc_link, '') <> ''",
                (project_id,),
            ).fetchone()["count"]
        )
        linked_bug_total = int(
            db.execute(
                """
                SELECT COUNT(*) AS count
                FROM bugs
                JOIN requirements ON bugs.requirement_id = requirements.id
                WHERE requirements.project_id = ?
                """,
                (project_id,),
            ).fetchone()["count"]
        )
        return {
            "total": total,
            "requirement_doc_count": requirement_doc_count,
            "design_doc_count": design_doc_count,
            "linked_bug_total": linked_bug_total,
        }

    def fetch_requirement_page(filters: dict | None = None, page: int = 1) -> dict:
        db = get_db()
        where_sql, params = build_requirement_query(filters)
        total = int(
            db.execute(
                f"""
                SELECT COUNT(*) AS count
                FROM requirements
                WHERE {where_sql}
                """,
                params,
            ).fetchone()["count"]
        )
        page_size = app.config["PAGE_SIZE"]
        pages = max(1, math.ceil(total / page_size)) if total else 1
        page = max(1, min(page, pages))
        offset = (page - 1) * page_size
        items = db.execute(
            f"""
            SELECT
                requirements.*,
                projects.name AS project_name,
                users.name AS creator_name,
                COUNT(bugs.id) AS linked_bug_count
            FROM requirements
            JOIN projects ON requirements.project_id = projects.id
            LEFT JOIN users ON requirements.creator_id = users.id
            LEFT JOIN bugs ON bugs.requirement_id = requirements.id
            WHERE {where_sql}
            GROUP BY requirements.id, projects.name, users.name
            ORDER BY datetime(requirements.created_at) DESC, requirements.id DESC
            LIMIT ? OFFSET ?
            """,
            [*params, page_size, offset],
        ).fetchall()
        return {
            "items": items,
            "total": total,
            "page": page,
            "pages": pages,
            "start_index": offset + 1 if total else 0,
            "end_index": min(offset + page_size, total),
            "has_prev": page > 1,
            "has_next": page < pages,
        }

    def fetch_requirement(requirement_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT
                requirements.*,
                projects.name AS project_name,
                users.name AS creator_name,
                COUNT(bugs.id) AS linked_bug_count
            FROM requirements
            JOIN projects ON requirements.project_id = projects.id
            LEFT JOIN users ON requirements.creator_id = users.id
            LEFT JOIN bugs ON bugs.requirement_id = requirements.id
            WHERE requirements.id = ? AND requirements.project_id = ?
            GROUP BY requirements.id, projects.name, users.name
            """,
            (requirement_id, current_project_id()),
        ).fetchone()

    def fetch_requirement_bugs(requirement_id: int) -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT
                bugs.*,
                creator.name AS creator_name,
                assignee.name AS assignee_name
            FROM bugs
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            WHERE bugs.requirement_id = ?
            ORDER BY datetime(bugs.created_at) DESC, bugs.id DESC
            """,
            (requirement_id,),
        ).fetchall()

    def fetch_cases_for_project() -> list[sqlite3.Row]:
        return get_db().execute(
            "SELECT * FROM test_cases WHERE project_id = ? ORDER BY id DESC",
            (current_project_id(),),
        ).fetchall()

    def count_test_cases(version: str = "", project_id: int | None = None) -> int:
        target_project_id = project_id or current_project_id()
        if target_project_id is None:
            return 0
        params: list[object] = [target_project_id]
        version_sql = ""
        if version:
            version_sql = " AND COALESCE(version, '') = ?"
            params.append(version)
        return int(
            get_db()
            .execute(
                f"""
                SELECT COUNT(*) AS count
                FROM test_cases
                WHERE project_id = ?{version_sql}
                """,
                params,
            )
            .fetchone()["count"]
        )

    def fetch_summary(version: str | list[str] = "", project_id: int | None = None, user_id: int | None = None) -> dict:
        db = get_db()
        target_project_id = project_id or current_project_id()
        current_user = g.get("current_user")
        if target_project_id is None:
            return {
                "total": 0,
                "active_count": 0,
                "verification_count": 0,
                "closed_count": 0,
                "my_todo_count": 0,
                "notification_unread_count": 0,
            }
        version_values = clean_filter_values(version)
        version_sql = ""
        version_params: list[str] = []
        if version_values:
            placeholders = ", ".join("?" for _ in version_values)
            version_sql = f" AND COALESCE(version, '') IN ({placeholders})"
            version_params = version_values
        total = db.execute(
            f"SELECT COUNT(*) AS count FROM bugs WHERE project_id = ?{version_sql}",
            [target_project_id, *version_params],
        ).fetchone()["count"]
        active_count = db.execute(
            f"SELECT COUNT(*) AS count FROM bugs WHERE project_id = ?{version_sql} AND status IN ('open', 'in_progress')",
            [target_project_id, *version_params],
        ).fetchone()["count"]
        verification_count = db.execute(
            f"SELECT COUNT(*) AS count FROM bugs WHERE project_id = ?{version_sql} AND status = 'pending_verification'",
            [target_project_id, *version_params],
        ).fetchone()["count"]
        closed_count = db.execute(
            f"SELECT COUNT(*) AS count FROM bugs WHERE project_id = ?{version_sql} AND status IN ('closed', 'duplicate', 'on_hold')",
            [target_project_id, *version_params],
        ).fetchone()["count"]
        my_todo_count = 0
        target_user_id = user_id or (int(current_user["id"]) if current_user is not None else None)
        if target_user_id is not None:
            my_todo_count = db.execute(
                f"""
                SELECT COUNT(*) AS count
                FROM bugs
                WHERE project_id = ? AND assignee_id = ?
                    {version_sql}
                    AND status IN ('open', 'in_progress', 'pending_verification')
                """,
                [target_project_id, target_user_id, *version_params],
            ).fetchone()["count"]
        notification_unread_count = count_user_notifications(target_user_id, unread_only=True) if target_user_id is not None else 0
        return {
            "total": total,
            "active_count": active_count,
            "verification_count": verification_count,
            "closed_count": closed_count,
            "my_todo_count": my_todo_count,
            "notification_unread_count": notification_unread_count,
        }

    def fetch_recent_report_bugs(project_id: int, version: str = "", limit: int = 5) -> list[sqlite3.Row]:
        params: list[object] = [project_id]
        version_sql = ""
        if version:
            version_sql = " AND COALESCE(bugs.version, '') = ?"
            params.append(version)
        params.append(limit)
        return get_db().execute(
            f"""
            SELECT
                bugs.id,
                bugs.bug_no,
                bugs.title,
                bugs.severity,
                bugs.status,
                bugs.version,
                assignee.name AS assignee_name
            FROM bugs
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            WHERE bugs.project_id = ?
                {version_sql}
                AND bugs.status IN ('open', 'in_progress', 'pending_verification')
            ORDER BY
                CASE bugs.severity
                    WHEN '最高' THEN 0
                    WHEN '高' THEN 1
                    WHEN '中' THEN 2
                    WHEN '低' THEN 3
                    WHEN '最低' THEN 4
                    WHEN '建议' THEN 5
                    WHEN '严重' THEN 0
                    WHEN '一般' THEN 2
                    ELSE 9
                END,
                datetime(bugs.updated_at) DESC,
                bugs.id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()

    def fetch_report_risk_bugs(project_id: int, version: str = "") -> list[sqlite3.Row]:
        params: list[object] = [project_id]
        version_sql = ""
        if version:
            version_sql = " AND COALESCE(bugs.version, '') = ?"
            params.append(version)
        return get_db().execute(
            f"""
            SELECT
                bugs.id,
                bugs.bug_no,
                bugs.title,
                bugs.severity,
                bugs.status,
                assignee.name AS assignee_name
            FROM bugs
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            WHERE bugs.project_id = ?
                {version_sql}
                AND COALESCE(bugs.severity, '') = '{MAIL_NOTIFY_SEVERITY}'
                AND bugs.status IN ('open', 'in_progress', 'pending_verification')
            ORDER BY
                CASE bugs.status
                    WHEN 'open' THEN 0
                    WHEN 'in_progress' THEN 1
                    WHEN 'pending_verification' THEN 2
                    ELSE 9
                END,
                datetime(bugs.updated_at) DESC,
                bugs.id DESC
            """,
            params,
        ).fetchall()

    def fetch_open_bug_counts_by_platform(project_id: int, version: str = "") -> list[dict[str, object]]:
        params: list[object] = [project_id]
        version_sql = ""
        if version:
            version_sql = " AND COALESCE(version, '') = ?"
            params.append(version)
        rows = get_db().execute(
            f"""
            SELECT COALESCE(NULLIF(platform, ''), '未填写') AS platform, COUNT(*) AS count
            FROM bugs
            WHERE project_id = ?
                {version_sql}
                AND status IN ('open', 'in_progress')
            GROUP BY COALESCE(NULLIF(platform, ''), '未填写')
            """,
            params,
        ).fetchall()
        counts_by_platform = {str(row["platform"]): int(row["count"] or 0) for row in rows}
        ordered_platforms = [*BUG_PLATFORM_OPTIONS, "未填写"]
        extras = sorted(platform for platform in counts_by_platform if platform not in ordered_platforms)
        result: list[dict[str, object]] = []
        for platform in [*ordered_platforms, *extras]:
            count = counts_by_platform.get(platform, 0)
            if count <= 0:
                continue
            result.append(
                {
                    "platform": platform,
                    "label": REPORT_PLATFORM_LABELS.get(platform, platform),
                    "count": count,
                }
            )
        return result

    def fetch_my_todos() -> list[sqlite3.Row]:
        if g.current_user is None:
            return []
        return get_db().execute(
            """
            SELECT
                bugs.*,
                projects.name AS project_name,
                assignee.name AS assignee_name,
                creator.name AS creator_name
            FROM bugs
            JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            WHERE bugs.assignee_id = ?
                AND bugs.status IN ('open', 'in_progress', 'pending_verification')
            ORDER BY bugs.updated_at DESC, bugs.id DESC
            """,
            (g.current_user["id"],),
        ).fetchall()

    def normalize_todo_status_filters(raw_values: list[str]) -> tuple[list[str], list[str]]:
        status_aliases = {label: code for code, label in STATUS_OPTIONS}
        status_aliases.update({code: code for code, _label in STATUS_OPTIONS})
        values = []
        for raw_value in raw_values:
            values.extend(part.strip() for part in str(raw_value or "").split(","))
        if not any(values):
            return list(TODO_STATUS_CODES), []
        statuses: list[str] = []
        invalid_values: list[str] = []
        for value in values:
            if not value:
                continue
            status = status_aliases.get(value) or status_aliases.get(value.lower())
            if status not in TODO_STATUS_CODES:
                invalid_values.append(value)
                continue
            if status not in statuses:
                statuses.append(status)
        return statuses, invalid_values

    def normalize_bug_status_value(raw_value: object) -> str | None:
        # 写接口同时兼容中文状态名和内部状态码。
        status_aliases = {label: code for code, label in STATUS_OPTIONS}
        status_aliases.update({code: code for code, _label in STATUS_OPTIONS})
        value = str(raw_value or "").strip()
        return status_aliases.get(value) or status_aliases.get(value.lower())

    def fetch_user_todos_with_detail(user_id: int, statuses: list[str]) -> list[sqlite3.Row]:
        placeholders = ",".join("?" for _ in statuses)
        return get_db().execute(
            f"""
            SELECT
                bugs.*,
                projects.name AS project_name,
                assignee.name AS assignee_name,
                assignee.username AS assignee_username,
                creator.name AS creator_name,
                creator.username AS creator_username,
                requirements.code AS requirement_code,
                requirements.title AS requirement_title,
                test_cases.case_no AS case_no,
                test_cases.title AS case_title
            FROM bugs
            JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN requirements ON bugs.requirement_id = requirements.id
            LEFT JOIN test_cases ON bugs.case_id = test_cases.id
            WHERE bugs.assignee_id = ?
              AND bugs.status IN ({placeholders})
            ORDER BY bugs.updated_at DESC, bugs.id DESC
            """,
            (user_id, *statuses),
        ).fetchall()

    def fetch_todo_detail_bug(bug_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT
                bugs.*,
                projects.name AS project_name,
                assignee.name AS assignee_name,
                assignee.username AS assignee_username,
                creator.name AS creator_name,
                creator.username AS creator_username,
                requirements.code AS requirement_code,
                requirements.title AS requirement_title,
                test_cases.case_no AS case_no,
                test_cases.title AS case_title
            FROM bugs
            JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN requirements ON bugs.requirement_id = requirements.id
            LEFT JOIN test_cases ON bugs.case_id = test_cases.id
            WHERE bugs.id = ?
            """,
            (bug_id,),
        ).fetchone()

    def serialize_todo_attachment(attachment: sqlite3.Row) -> dict[str, object]:
        attachment_id = int(attachment["id"])
        return {
            "id": attachment_id,
            "filename": attachment["filename"] or "",
            "content_type": attachment["content_type"] or "",
            "source_field": attachment["source_field"] or "",
            "url": url_for("view_attachment", attachment_id=attachment_id),
            "created_at": attachment["created_at"] or "",
        }

    def serialize_todo_comment(comment: sqlite3.Row) -> dict[str, object]:
        return {
            "id": int(comment["id"]),
            "parent_id": int(comment["parent_id"]) if comment["parent_id"] else None,
            "user_id": int(comment["user_id"]),
            "author_name": comment["commenter_name"] or comment["author_name"] or "",
            "author_role": comment["commenter_role"] or "",
            "content": comment["content"] or "",
            "created_at": comment["created_at"] or "",
            "updated_at": comment["updated_at"] or "",
        }

    def serialize_todo_history(history: sqlite3.Row) -> dict[str, object]:
        return {
            "id": int(history["id"]),
            "action": history["action"] or "",
            "detail": history["detail"] or "",
            "operator_name": history["operator_name"] or "",
            "environment_snapshot": history["environment_snapshot"] or "",
            "status_snapshot": history["status_snapshot"] or "",
            "status_snapshot_label": STATUS_LABELS.get(str(history["status_snapshot"] or ""), str(history["status_snapshot"] or "")),
            "assignee_snapshot": history["assignee_snapshot"] or "",
            "created_at": history["created_at"] or "",
        }

    def serialize_todo_detail(
        row: sqlite3.Row,
        comments: list[sqlite3.Row] | None = None,
        history: list[sqlite3.Row] | None = None,
        attachments: list[sqlite3.Row] | None = None,
    ) -> dict[str, object]:
        bug_id = int(row["id"])
        # 详情独立序列化，读取接口和状态变更接口返回同一份明细结构。
        return {
            "id": bug_id,
            "bug_no": format_bug_no(row["bug_no"] or bug_id),
            "title": row["title"] or "",
            "project": {
                "id": int(row["project_id"]),
                "name": row["project_name"] or "",
            },
            "version": row["version"] or "",
            "module": row["module"] or "",
            "platform": row["platform"] or "",
            "severity": row["severity"] or "",
            "priority": row["priority"] or "",
            "status": row["status"] or "",
            "status_label": STATUS_LABELS.get(str(row["status"] or ""), str(row["status"] or "")),
            "assignee": {
                "id": int(row["assignee_id"] or 0),
                "name": row["assignee_name"] or "",
                "username": row["assignee_username"] or "",
            },
            "creator": {
                "id": int(row["creator_id"] or 0),
                "name": row["creator_name"] or "",
                "username": row["creator_username"] or "",
            },
            "reporter": row["reporter"] or "",
            "environment": row["environment"] or "",
            "created_at": row["created_at"] or "",
            "updated_at": row["updated_at"] or "",
            "description": row["description"] or "",
            "expected_result": row["expected_result"] or "",
            "actual_result": row["actual_result"] or "",
            "resolution_note": row["resolution_note"] or "",
            "requirement": {
                "id": int(row["requirement_id"]) if row["requirement_id"] else None,
                "code": row["requirement_code"] or "",
                "title": row["requirement_title"] or "",
            },
            "case": {
                "id": int(row["case_id"]) if row["case_id"] else None,
                "case_no": row["case_no"] or "",
                "title": row["case_title"] or "",
            },
            "comments": [serialize_todo_comment(item) for item in comments or []],
            "history": [serialize_todo_history(item) for item in history or []],
            "attachments": [serialize_todo_attachment(item) for item in attachments or []],
            "url": url_for("bug_detail", bug_id=bug_id),
        }

    def serialize_todo_bug(row: sqlite3.Row, detail: dict[str, object] | None = None) -> dict[str, object]:
        detail = detail or serialize_todo_detail(row)
        # 保留 todos[].detail 兼容旧调用方，评论等明细也同步放入。
        return {
            "id": detail["id"],
            "bug_no": detail["bug_no"],
            "title": detail["title"],
            "project": detail["project"],
            "version": detail["version"],
            "module": detail["module"],
            "platform": detail["platform"],
            "severity": detail["severity"],
            "priority": detail["priority"],
            "status": detail["status"],
            "status_label": detail["status_label"],
            "assignee": detail["assignee"],
            "creator": detail["creator"],
            "reporter": detail["reporter"],
            "environment": detail["environment"],
            "created_at": detail["created_at"],
            "updated_at": detail["updated_at"],
            "detail": detail,
        }

    def serialize_todo_detail_with_relations(row: sqlite3.Row) -> dict[str, object]:
        bug_id = int(row["id"])
        return serialize_todo_detail(
            row,
            comments=fetch_bug_comments(bug_id),
            history=fetch_bug_history(bug_id),
            attachments=fetch_bug_attachments(bug_id),
        )

    def fetch_bug(bug_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT
                bugs.*,
                projects.name AS project_name,
                assignee.name AS assignee_name,
                creator.name AS creator_name,
                previous_user.name AS previous_assignee_name,
                requirements.code AS requirement_code,
                requirements.title AS requirement_title,
                test_cases.case_no AS case_no,
                test_cases.title AS case_title
            FROM bugs
            JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users assignee ON bugs.assignee_id = assignee.id
            LEFT JOIN users creator ON bugs.creator_id = creator.id
            LEFT JOIN users previous_user ON bugs.previous_assignee_id = previous_user.id
            LEFT JOIN requirements ON bugs.requirement_id = requirements.id
            LEFT JOIN test_cases ON bugs.case_id = test_cases.id
            WHERE bugs.id = ?
            """,
            (bug_id,),
        ).fetchone()

    def create_notification(
        user_id: int | None,
        category: str,
        title: str,
        body: str,
        link_path: str = "",
        bug_id: int | None = None,
        actor_id: int | None = None,
        comment_id: int | None = None,
    ) -> int:
        if not user_id:
            return 0
        cursor = get_db().execute(
            """
            INSERT INTO notifications (
                user_id, actor_id, bug_id, comment_id, category, title, body, link_path,
                is_read, created_at, read_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, NULL)
            """,
            (
                int(user_id),
                actor_id,
                bug_id,
                comment_id,
                category,
                title,
                body,
                link_path,
                current_time(),
            ),
        )
        get_db().commit()
        return int(cursor.lastrowid)

    def fetch_notification(notification_id: int, user_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT *
            FROM notifications
            WHERE id = ? AND user_id = ?
            """,
            (notification_id, user_id),
        ).fetchone()

    def fetch_user_notifications(user_id: int, state: str = "", limit: int = 80) -> list[sqlite3.Row]:
        where_parts = ["notifications.user_id = ?"]
        params: list[object] = [user_id]
        if state == "unread":
            where_parts.append("notifications.is_read = 0")
        params.append(limit)
        return get_db().execute(
            f"""
            SELECT
                notifications.*,
                bugs.bug_no,
                bugs.title AS bug_title,
                bugs.status AS bug_status,
                bugs.severity AS bug_severity,
                projects.name AS project_name,
                actor.name AS actor_name
            FROM notifications
            LEFT JOIN bugs ON notifications.bug_id = bugs.id
            LEFT JOIN projects ON bugs.project_id = projects.id
            LEFT JOIN users actor ON notifications.actor_id = actor.id
            WHERE {' AND '.join(where_parts)}
            ORDER BY notifications.is_read ASC, datetime(notifications.created_at) DESC, notifications.id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()

    def count_user_notifications(user_id: int, unread_only: bool = False) -> int:
        where_sql = "WHERE user_id = ?"
        params: list[object] = [user_id]
        if unread_only:
            where_sql += " AND is_read = 0"
        return int(
            get_db()
            .execute(f"SELECT COUNT(*) AS count FROM notifications {where_sql}", params)
            .fetchone()["count"]
        )

    def mark_notification_read(notification_id: int, user_id: int) -> bool:
        row = fetch_notification(notification_id, user_id)
        if row is None:
            return False
        if not int(row["is_read"] or 0):
            get_db().execute(
                "UPDATE notifications SET is_read = 1, read_at = ? WHERE id = ? AND user_id = ?",
                (current_time(), notification_id, user_id),
            )
            get_db().commit()
        return True

    def mark_all_notifications_read(user_id: int) -> int:
        cursor = get_db().execute(
            """
            UPDATE notifications
            SET is_read = 1, read_at = ?
            WHERE user_id = ? AND is_read = 0
            """,
            (current_time(), user_id),
        )
        get_db().commit()
        return int(cursor.rowcount or 0)

    def count_unread_bug_comment_notifications(user_id: int, bug_id: int) -> int:
        placeholders = ",".join("?" for _ in COMMENT_NOTIFICATION_CATEGORIES)
        return int(
            get_db()
            .execute(
                f"""
                SELECT COUNT(*) AS count
                FROM notifications
                WHERE user_id = ?
                    AND bug_id = ?
                    AND is_read = 0
                    AND category IN ({placeholders})
                """,
                (user_id, bug_id, *COMMENT_NOTIFICATION_CATEGORIES),
            )
            .fetchone()["count"]
        )

    def mark_bug_comment_notifications_read(user_id: int, bug_id: int) -> int:
        placeholders = ",".join("?" for _ in COMMENT_NOTIFICATION_CATEGORIES)
        cursor = get_db().execute(
            f"""
            UPDATE notifications
            SET is_read = 1, read_at = ?
            WHERE user_id = ?
                AND bug_id = ?
                AND is_read = 0
                AND category IN ({placeholders})
            """,
            (current_time(), user_id, bug_id, *COMMENT_NOTIFICATION_CATEGORIES),
        )
        get_db().commit()
        return int(cursor.rowcount or 0)

    def notification_category_label(category: object) -> str:
        category_text = str(category or "").strip()
        return NOTIFICATION_CATEGORY_LABELS.get(category_text, category_text or "消息")

    def notification_snippet(content: str, max_length: int = 120) -> str:
        text = re.sub(r"\s+", " ", str(content or "")).strip()
        if len(text) <= max_length:
            return text
        return text[:max_length].rstrip() + "..."

    def mention_tokens_for_users(users: list[sqlite3.Row]) -> list[dict[str, object]]:
        tokens: list[dict[str, object]] = []
        seen_pairs: set[tuple[int, str]] = set()
        for user in users:
            user_id = int(user["id"])
            display_name = str(user["name"] or user["username"] or "").strip()
            for token in {str(user["name"] or "").strip(), str(user["username"] or "").strip()}:
                if not token or (user_id, token) in seen_pairs:
                    continue
                tokens.append(
                    {
                        "user_id": user_id,
                        "token": token,
                        "name": display_name or token,
                    }
                )
                seen_pairs.add((user_id, token))
        tokens.sort(key=lambda item: len(str(item["token"])), reverse=True)
        return tokens

    def find_comment_mentions(content: str, users: list[sqlite3.Row]) -> list[dict[str, object]]:
        content_text = str(content or "")
        tokens = mention_tokens_for_users(users)
        mentions: list[dict[str, object]] = []
        index = 0

        def has_mention_boundary(end_index: int) -> bool:
            if end_index >= len(content_text):
                return True
            next_char = content_text[end_index]
            return next_char.isspace() or not (next_char.isalnum() or next_char == "_")

        while index < len(content_text):
            if content_text[index] != "@":
                index += 1
                continue
            matched = None
            for item in tokens:
                token = str(item["token"])
                mention_text = f"@{token}"
                end_index = index + len(mention_text)
                if content_text.startswith(mention_text, index) and has_mention_boundary(end_index):
                    matched = {
                        "start": index,
                        "end": end_index,
                        "user_id": int(item["user_id"]),
                        "name": str(item["name"]),
                        "text": mention_text,
                    }
                    break
            if matched is None:
                index += 1
                continue
            mentions.append(matched)
            index = int(matched["end"])
        return mentions

    def extract_mentioned_user_ids(content: str, users: list[sqlite3.Row]) -> set[int]:
        return {int(item["user_id"]) for item in find_comment_mentions(content, users)}

    def fetch_comment_mention_states(bug_id: int) -> dict[int, dict[int, bool]]:
        rows = get_db().execute(
            """
            SELECT comment_id, user_id, is_read
            FROM notifications
            WHERE bug_id = ?
                AND category = 'comment_mention'
                AND comment_id IS NOT NULL
            ORDER BY id ASC
            """,
            (bug_id,),
        ).fetchall()
        states: dict[int, dict[int, bool]] = {}
        for row in rows:
            comment_id = int(row["comment_id"] or 0)
            user_id = int(row["user_id"] or 0)
            if not comment_id or not user_id:
                continue
            user_states = states.setdefault(comment_id, {})
            user_states[user_id] = bool(row["is_read"])
        return states

    def build_comment_content_parts(
        content: str,
        users: list[sqlite3.Row],
        mention_states: dict[int, bool],
    ) -> list[dict[str, object]]:
        content_text = str(content or "")
        parts: list[dict[str, object]] = []
        cursor = 0
        for mention in find_comment_mentions(content_text, users):
            start = int(mention["start"])
            end = int(mention["end"])
            if start > cursor:
                parts.append({"type": "text", "text": content_text[cursor:start]})
            user_id = int(mention["user_id"])
            is_read = mention_states.get(user_id, True)
            parts.append(
                {
                    "type": "mention",
                    "text": str(mention["text"]),
                    "user_id": user_id,
                    "name": str(mention["name"]),
                    "read_state": "read" if is_read else "unread",
                }
            )
            cursor = end
        if cursor < len(content_text):
            parts.append({"type": "text", "text": content_text[cursor:]})
        return parts or [{"type": "text", "text": ""}]

    def notify_bug_comment_recipients(
        bug: sqlite3.Row,
        comment_id: int,
        content: str,
    ) -> int:
        if g.get("current_user") is None:
            return 0
        actor_id = int(g.current_user["id"])
        actor_name = str(g.current_user["name"] or "")
        users = fetch_users()
        mentioned_user_ids = extract_mentioned_user_ids(content, users)
        recipient_categories: dict[int, str] = {}
        for user_id in mentioned_user_ids:
            if user_id != actor_id:
                recipient_categories[user_id] = "comment_mention"

        creator_id = int(bug["creator_id"] or 0)
        if creator_id and creator_id != actor_id and creator_id not in recipient_categories:
            recipient_categories[creator_id] = "bug_comment"

        if not recipient_categories:
            return 0

        bug_no = format_bug_no(bug["bug_no"] or bug["id"])
        link_path = url_for("bug_detail", bug_id=int(bug["id"]), tab="detail") + f"#comment-{comment_id}"
        body = f"{bug_no}「{bug['title'] or '-'}」：{notification_snippet(content)}"
        created_count = 0
        for user_id, category in recipient_categories.items():
            if category == "comment_mention":
                title = f"{actor_name} 在评论中提到了你"
            else:
                title = f"{actor_name} 评论了你创建的 Bug"
            create_notification(
                user_id=user_id,
                actor_id=actor_id,
                bug_id=int(bug["id"]),
                comment_id=comment_id,
                category=category,
                title=title,
                body=body,
                link_path=link_path,
            )
            created_count += 1
        return created_count

    def build_bug_mention_users(bug: sqlite3.Row, comments: list[sqlite3.Row]) -> list[sqlite3.Row]:
        users = fetch_users()
        users_by_id = {int(user["id"]): user for user in users}
        ordered_ids: list[int] = []

        def add_user_id(raw_user_id: object) -> None:
            try:
                user_id = int(raw_user_id or 0)
            except (TypeError, ValueError):
                return
            if user_id and user_id in users_by_id and user_id not in ordered_ids:
                ordered_ids.append(user_id)

        add_user_id(bug["creator_id"])
        add_user_id(bug["assignee_id"])
        add_user_id(bug["previous_assignee_id"])
        for comment in comments:
            add_user_id(comment["user_id"])
        return [users_by_id[user_id] for user_id in ordered_ids]

    def create_severe_bug_assignment_message(
        bug: sqlite3.Row | None,
        assignee_user_id: int | None,
        trigger_reason: str,
        operator_name: str,
    ) -> tuple[bool, str]:
        if bug is None:
            return False, "未找到对应 Bug。"
        if str(bug["severity"] or "") != MAIL_NOTIFY_SEVERITY:
            return False, "当前 Bug 非严重级别。"
        if str(bug["status"] or "") not in TODO_STATUS_CODES:
            return False, "当前状态无需发送严重 Bug 站内消息。"
        if not assignee_user_id:
            return False, "当前处理人为空，无法发送站内消息。"

        assignee_user = fetch_user(int(assignee_user_id))
        if assignee_user is None:
            return False, "当前处理人不存在。"

        bug_no = format_bug_no(bug["bug_no"] or bug["id"])
        status_label = STATUS_LABELS.get(str(bug["status"] or ""), str(bug["status"] or "-"))
        title = f"严重 Bug 已进入你的待办：{bug_no}"
        body = (
            f"{operator_name or '系统'} {trigger_reason}，"
            f"{bug_no}「{bug['title'] or '-'}」已进入你的待办。"
            f"项目：{bug['project_name'] or '-'}；状态：{status_label}。"
        )
        actor_id = int(g.current_user["id"]) if g.get("current_user") is not None else None
        create_notification(
            user_id=int(assignee_user_id),
            actor_id=actor_id,
            bug_id=int(bug["id"]),
            category="severe_bug",
            title=title,
            body=body,
            link_path=url_for("bug_detail", bug_id=int(bug["id"])),
        )
        return True, f"已发送站内消息给 {assignee_user['name']}。"

    def fetch_bug_history(bug_id: int) -> list[sqlite3.Row]:
        return get_db().execute(
            "SELECT * FROM bug_history WHERE bug_id = ? ORDER BY created_at DESC, id DESC",
            (bug_id,),
        ).fetchall()

    def fetch_bug_comments(bug_id: int) -> list[sqlite3.Row]:
        return get_db().execute(
            """
            SELECT
                bug_comments.*,
                COALESCE(users.name, bug_comments.author_name) AS commenter_name,
                COALESCE(users.role, '') AS commenter_role
            FROM bug_comments
            LEFT JOIN users ON bug_comments.user_id = users.id
            WHERE bug_comments.bug_id = ?
            ORDER BY bug_comments.created_at DESC, bug_comments.id DESC
            """,
            (bug_id,),
        ).fetchall()

    def fetch_bug_comment(comment_id: int, bug_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            """
            SELECT
                bug_comments.*,
                COALESCE(users.name, bug_comments.author_name) AS commenter_name
            FROM bug_comments
            LEFT JOIN users ON bug_comments.user_id = users.id
            WHERE bug_comments.id = ? AND bug_comments.bug_id = ?
            """,
            (comment_id, bug_id),
        ).fetchone()

    def collect_comment_branch_ids(bug_id: int, root_comment_id: int) -> list[int]:
        rows = get_db().execute(
            "SELECT id, parent_id FROM bug_comments WHERE bug_id = ? ORDER BY id ASC",
            (bug_id,),
        ).fetchall()
        children_map: dict[int | None, list[int]] = {}
        for row in rows:
            parent_id = int(row["parent_id"]) if row["parent_id"] is not None else None
            children_map.setdefault(parent_id, []).append(int(row["id"]))

        result: list[int] = []
        stack = [root_comment_id]
        while stack:
            current_id = stack.pop()
            result.append(current_id)
            for child_id in reversed(children_map.get(current_id, [])):
                stack.append(child_id)
        return result

    def build_bug_comment_threads(
        comments: list[sqlite3.Row],
        mention_users: list[sqlite3.Row],
        comment_mention_states: dict[int, dict[int, bool]],
    ) -> list[dict[str, object]]:
        nodes: list[dict[str, object]] = []
        by_id: dict[int, dict[str, object]] = {}
        roots: list[dict[str, object]] = []

        for item in comments:
            comment_id = int(item["id"])
            node = {
                "id": comment_id,
                "parent_id": int(item["parent_id"]) if item["parent_id"] is not None else None,
                "user_id": int(item["user_id"]),
                "actor_name": str(item["commenter_name"] or item["author_name"] or ""),
                "actor_role": str(item["commenter_role"] or ""),
                "created_at": str(item["created_at"] or ""),
                "content": str(item["content"] or ""),
                "content_parts": build_comment_content_parts(
                    str(item["content"] or ""),
                    mention_users,
                    comment_mention_states.get(comment_id, {}),
                ),
                "reply_to_name": "",
                "replies": [],
            }
            nodes.append(node)
            by_id[comment_id] = node

        for node in nodes:
            parent_id = node["parent_id"]
            if parent_id and parent_id in by_id:
                parent_node = by_id[parent_id]
                node["reply_to_name"] = str(parent_node["actor_name"] or "")
                parent_node["replies"].append(node)
            else:
                roots.append(node)

        for node in nodes:
            node["replies"].sort(
                key=lambda reply: (
                    str(reply["created_at"]),
                    int(reply["id"]),
                ),
                reverse=True,
            )
        roots.sort(
            key=lambda root: (
                str(root["created_at"]),
                int(root["id"]),
            ),
            reverse=True,
        )
        return roots

    def build_bug_activity_items(history: list[sqlite3.Row], comments: list[sqlite3.Row]) -> list[dict[str, object]]:
        activity_items: list[dict[str, object]] = []
        for item in comments:
            activity_items.append(
                {
                    "type": "comment",
                    "id": int(item["id"]),
                    "created_at": str(item["created_at"] or ""),
                    "actor_name": str(item["commenter_name"] or item["author_name"] or ""),
                    "actor_role": str(item["commenter_role"] or ""),
                    "content": str(item["content"] or ""),
                }
            )
        for item in history:
            activity_items.append(
                {
                    "type": "history",
                    "id": int(item["id"]),
                    "created_at": str(item["created_at"] or ""),
                    "actor_name": str(item["operator_name"] or ""),
                    "actor_role": "",
                    "action": str(item["action"] or ""),
                    "content": str(item["detail"] or ""),
                    "status_snapshot": str(item["status_snapshot"] or ""),
                    "assignee_snapshot": str(item["assignee_snapshot"] or ""),
                    "environment_snapshot": str(item["environment_snapshot"] or ""),
                }
            )
        activity_items.sort(
            key=lambda item: (
                str(item["created_at"]),
                1 if item["type"] == "comment" else 0,
                int(item["id"]),
            ),
            reverse=True,
        )
        return activity_items

    def fetch_bug_attachments(bug_id: int) -> list[sqlite3.Row]:
        return get_db().execute(
            "SELECT * FROM bug_attachments WHERE bug_id = ? ORDER BY created_at ASC, id ASC",
            (bug_id,),
        ).fetchall()

    def is_image_attachment(attachment: sqlite3.Row) -> bool:
        return str(attachment["content_type"] or "").startswith("image/")

    def group_bug_attachments(attachments: list[sqlite3.Row]) -> tuple[dict[str, list[sqlite3.Row]], list[sqlite3.Row]]:
        attachments_by_field: dict[str, list[sqlite3.Row]] = {field: [] for field in BUG_INLINE_ATTACHMENT_FIELDS}
        general_attachments: list[sqlite3.Row] = []
        for attachment in attachments:
            source_field = normalize_attachment_source(attachment["source_field"])
            if source_field in attachments_by_field and is_image_attachment(attachment):
                attachments_by_field[source_field].append(attachment)
            else:
                general_attachments.append(attachment)
        return attachments_by_field, general_attachments

    def allowed_status_transitions(status: str) -> list[str]:
        return list(STATUS_LABELS.keys())

    def derive_previous_assignee_id_for_bug(
        bug: sqlite3.Row,
        next_assignee_id: int,
    ) -> int:
        current_previous_assignee_id = int(bug["previous_assignee_id"] or 0)
        if str(bug["status"] or "") == "pending_verification":
            return current_previous_assignee_id or next_assignee_id
        return next_assignee_id

    def apply_bug_action(
        db: sqlite3.Connection,
        bug: sqlite3.Row,
        action: str,
        operator_name: str,
        note: str = "",
        assignee_id: int | None = None,
        selected_status: str | None = None,
    ) -> tuple[str, int, int | None, str, str]:
        new_status = bug["status"]
        new_assignee_id = int(bug["assignee_id"])
        new_previous_assignee_id = bug["previous_assignee_id"]
        detail = "补充了处理记录"
        action_label = "更新缺陷"

        if action == "start_progress":
            new_status = "in_progress"
            action_label = "开始处理"
            detail = f"{operator_name} 开始处理该缺陷"
        elif action == "resolve":
            new_status = "pending_verification"
            new_previous_assignee_id = bug["assignee_id"]
            new_assignee_id = int(bug["creator_id"] or bug["assignee_id"])
            action_label = "提交待验证"
            detail = f"{operator_name} 提交缺陷进入待验证，系统自动回到创建人 {bug['creator_name']} 的待办"
        elif action == "reject":
            new_status = "in_progress"
            new_assignee_id = int(bug["previous_assignee_id"] or bug["assignee_id"])
            action_label = "退回处理"
            reject_user = fetch_user(new_assignee_id)
            reject_name = reject_user["name"] if reject_user else "原处理人"
            detail = f"{operator_name} 将缺陷退回处理中，系统自动回到 {reject_name} 的待办"
        elif action == "close":
            new_status = "closed"
            new_assignee_id = int(bug["creator_id"] or bug["assignee_id"])
            action_label = "关闭缺陷"
            detail = f"{operator_name} 验证通过并关闭缺陷"
        elif action == "reassign":
            if not assignee_id:
                raise ValueError("请选择转交处理人。")
            target_user = fetch_bug_assignee_user(assignee_id)
            if target_user is None:
                raise ValueError("缺陷处理人不能选择 admin 账号。")
            new_assignee_id = int(target_user["id"])
            new_previous_assignee_id = derive_previous_assignee_id_for_bug(bug, new_assignee_id)
            action_label = "转交处理"
            target_name = target_user["name"]
            detail = f"{operator_name} 转交给 {target_name}"
        elif action == "change_status":
            # 页面表单从 request.form 读取，JSON 接口直接传入目标状态。
            target_status = (selected_status if selected_status is not None else request.form.get("status", "")).strip()
            if target_status not in STATUS_LABELS:
                raise ValueError("请选择有效状态。")
            if target_status not in allowed_status_transitions(bug["status"]):
                raise ValueError("当前状态不支持直接切换到该选项。")
            action_label = "更新状态"
            previous_status = STATUS_LABELS.get(bug["status"], bug["status"])
            new_status = target_status
            if target_status in {"open", "in_progress"} and str(bug["status"] or "") == "pending_verification":
                new_assignee_id = int(bug["previous_assignee_id"] or bug["assignee_id"])
                reject_user = fetch_user(new_assignee_id)
                reject_name = reject_user["name"] if reject_user else "原处理人"
                detail = f"{operator_name} 将状态更新为 {STATUS_LABELS[target_status]}，系统自动回到 {reject_name} 的待办"
            elif target_status == "pending_verification":
                new_previous_assignee_id = bug["assignee_id"]
                new_assignee_id = int(bug["creator_id"] or bug["assignee_id"])
                detail = f"{operator_name} 将状态更新为待验证，系统自动回到创建人 {bug['creator_name']} 的待办"
            elif target_status == "closed":
                new_assignee_id = int(bug["creator_id"] or bug["assignee_id"])
                detail = f"{operator_name} 将状态更新为已关闭"
            else:
                detail = f"{operator_name} 将状态从 {previous_status} 更新为 {STATUS_LABELS[target_status]}"

        if note:
            detail += f"；说明：{note}"

        return new_status, new_assignee_id, new_previous_assignee_id, action_label, detail

    def fetch_attachment(attachment_id: int) -> sqlite3.Row | None:
        return get_db().execute(
            "SELECT * FROM bug_attachments WHERE id = ?",
            (attachment_id,),
        ).fetchone()

    def execution_distribution(
        project_id: int | None = None,
        version: str | None = None,
        folder_name: str | None = None,
        doc_name: str | None = None,
    ) -> list[dict]:
        target_project_id = project_id or current_project_id()
        params: list[object] = [target_project_id]
        where_clauses = ["project_id = ?"]
        if version is not None and str(version).strip():
            where_clauses.append("COALESCE(version, '') = COALESCE(?, '')")
            params.append(str(version).strip())
        if doc_name is not None:
            where_clauses.append("COALESCE(folder_name, '') = COALESCE(?, '')")
            where_clauses.append("COALESCE(doc_name, '') = COALESCE(?, '')")
            params.extend([folder_name, doc_name])
        rows = get_db().execute(
            f"""
            SELECT execute_status, COUNT(*) AS count
            FROM test_cases
            WHERE {' AND '.join(where_clauses)}
            GROUP BY execute_status
            """,
            params,
        ).fetchall()
        total = sum(row["count"] for row in rows) or 1
        mapped = {row["execute_status"]: row["count"] for row in rows}
        result = []
        for status in CASE_STATUS_OPTIONS:
            count = mapped.get(status, 0)
            percent = f"{(count / total) * 100:.2f}%"
            result.append({"status": status, "count": count, "percent": percent, "color": CASE_STATUS_COLORS[status]})
        return result

    def build_case_chart_bytes(version: str = "") -> bytes:
        distribution = execution_distribution(version=version)
        max_count = max((item["count"] for item in distribution), default=0) or 1
        chart_height = 170
        base_y = 240
        left = 58
        bar_width = 74
        gap = 42
        bars = []
        labels = []
        for index, item in enumerate(distribution):
            x = left + index * (bar_width + gap)
            bar_height = 0 if item["count"] == 0 else round((item["count"] / max_count) * chart_height)
            y = base_y - bar_height
            bars.append(
                f'<rect x="{x}" y="{y}" width="{bar_width}" height="{bar_height}" rx="10" fill="{item["color"]}" />'
            )
            labels.append(
                f'<text x="{x + bar_width / 2}" y="{base_y + 28}" text-anchor="middle" font-size="13" fill="#606266">{escape(CASE_STATUS_CHART_LABELS[item["status"]])}</text>'
            )
            labels.append(
                f'<text x="{x + bar_width / 2}" y="{y - 10}" text-anchor="middle" font-size="12" fill="#303133">{item["count"]}</text>'
            )
            labels.append(
                f'<text x="{x + bar_width / 2}" y="{y - 28}" text-anchor="middle" font-size="11" fill="#909399">{escape(item["percent"])}</text>'
            )
        svg = f"""
        <svg xmlns="http://www.w3.org/2000/svg" width="{SVG_CHART_WIDTH}" height="{SVG_CHART_HEIGHT}" viewBox="0 0 {SVG_CHART_WIDTH} {SVG_CHART_HEIGHT}">
            <rect width="100%" height="100%" fill="#ffffff"/>
            <text x="34" y="38" font-size="20" font-weight="600" fill="#303133">执行结果分布</text>
            <line x1="40" y1="{base_y}" x2="{SVG_CHART_WIDTH - 34}" y2="{base_y}" stroke="#dcdfe6" stroke-width="1.2"/>
            <line x1="40" y1="70" x2="40" y2="{base_y}" stroke="#dcdfe6" stroke-width="1.2"/>
            {''.join(bars)}
            {''.join(labels)}
        </svg>
        """
        return svg.encode("utf-8")

    def fetch_report_page_url(page_number: int, version: str = "") -> str:
        query: dict[str, str | int] = {}
        if version:
            query["version"] = version
        query["page"] = page_number
        return url_for("testing_report", **query)

    def build_report_jump_url(version: str = "") -> str:
        query: dict[str, str] = {}
        if version:
            query["version"] = version
        return url_for("testing_report", **query)

    def fetch_report_data(version: str = "", page: int = 1) -> dict:
        project = fetch_current_project()
        report_filters = {"version": version} if version else {}
        report_bug_page = fetch_bug_page(report_filters, page)
        return {
            "project": project,
            "case_total": count_test_cases(version=version),
            "distribution": execution_distribution(version=version),
            "summary": fetch_summary(version),
            "open_bug_platform_counts": fetch_open_bug_counts_by_platform(project_id=int(project["id"]) if project else None, version=version) if project else [],
            "bugs": report_bug_page["items"],
            "bug_page": report_bug_page,
            "selected_version": version,
            "versions": fetch_report_versions(),
        }

    def is_admin() -> bool:
        return g.current_user is not None and (
            str(g.current_user["account_type"] or "") == "admin"
            or str(g.current_user["role_code"] or "") == ADMIN_ROLE_CODE
        )

    def build_page_url(page_number: int, filters: dict) -> str:
        query_items: list[tuple[str, object]] = []
        for key, value in filters.items():
            if key.endswith("_values"):
                continue
            if key in BUG_MULTI_FILTER_KEYS:
                for selected_value in filter_values(filters, key):
                    query_items.append((key, selected_value))
            elif value:
                query_items.append((key, value))
        query_items.append(("page", page_number))
        return f"{url_for('bug_list')}?{urllib_parse.urlencode(query_items)}"

    def build_page_jump_url(filters: dict) -> str:
        query_items: list[tuple[str, object]] = []
        for key, value in filters.items():
            if key.endswith("_values"):
                continue
            if key in BUG_MULTI_FILTER_KEYS:
                for selected_value in filter_values(filters, key):
                    query_items.append((key, selected_value))
            elif value:
                query_items.append((key, value))
        query = urllib_parse.urlencode(query_items)
        return f"{url_for('bug_list')}?{query}" if query else url_for("bug_list")

    def build_requirement_page_url(page_number: int, filters: dict) -> str:
        query = {key: value for key, value in filters.items() if value}
        query["page"] = page_number
        return url_for("requirement_library", **query)

    def build_requirement_jump_url(filters: dict) -> str:
        query = {key: value for key, value in filters.items() if value}
        return url_for("requirement_library", **query)

    def build_case_page_url(page_number: int) -> str:
        return url_for("case_library", page=page_number)

    def wants_json_response() -> bool:
        accept = (request.headers.get("Accept") or "").lower()
        requested_with = (request.headers.get("X-Requested-With") or "").lower()
        return request.args.get("format") == "json" or "application/json" in accept or requested_with == "xmlhttprequest"

    @app.before_request
    def load_common_data() -> Response | None:
        g.status_options = STATUS_OPTIONS
        g.status_labels = STATUS_LABELS
        g.requirement_status_options = REQUIREMENT_STATUS_OPTIONS
        g.requirement_status_labels = REQUIREMENT_STATUS_LABELS
        g.current_user = None
        user_id = session.get("user_id")
        if user_id:
            g.current_user = fetch_user(int(user_id))
        endpoint = request.endpoint or ""
        # 郑敬佩待办 JSON 接口按需求开放匿名访问。
        public_endpoints = {
            "login",
            "static",
            "api_health",
            "api_login",
            "zhengjingpei_todos_api",
            "zhengjingpei_todo_status_api",
        }
        if endpoint not in public_endpoints and g.current_user is None:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "请先登录。"}), 401
            next_url = current_local_path()
            return redirect(f"{url_for('login')}?{urllib_parse.urlencode({'next': next_url})}")
        g.projects = fetch_projects()
        g.current_project = fetch_current_project()
        return None

    @app.teardown_appcontext
    def close_db(_error: Exception | None) -> None:
        db = g.pop("db", None)
        if db is not None:
            db.close()

    @app.context_processor
    def inject_helpers() -> dict:
        summary = fetch_summary() if g.get("current_user") is not None else None
        return {
            "current_user": g.get("current_user"),
            "current_project": g.get("current_project"),
            "projects": g.get("projects", []),
            "summary": summary,
            "format_bug_no": format_bug_no,
            "build_page_url": build_page_url,
            "build_page_jump_url": build_page_jump_url,
            "build_requirement_page_url": build_requirement_page_url,
            "build_requirement_jump_url": build_requirement_jump_url,
            "fetch_report_page_url": fetch_report_page_url,
            "build_report_jump_url": build_report_jump_url,
            "build_case_page_url": build_case_page_url,
            "case_status_colors": CASE_STATUS_COLORS,
            "status_labels": STATUS_LABELS,
            "requirement_status_labels": REQUIREMENT_STATUS_LABELS,
            "severity_options": BUG_SEVERITY_OPTIONS,
            "bug_priority_options": BUG_PRIORITY_OPTIONS,
            "bug_priority_icon_map": BUG_PRIORITY_ICON_MAP,
            "bug_platform_options": BUG_PLATFORM_OPTIONS,
            "allowed_status_transitions": allowed_status_transitions,
            "can_edit_bug_platform": can_edit_bug_platform,
            "notification_category_label": notification_category_label,
            "current_local_path": current_local_path,
            "is_admin": is_admin(),
        }

    @app.route("/login", methods=["GET", "POST"])
    def login() -> str | Response:
        next_url = local_redirect_target(request.values.get("next", ""), url_for("bug_list"))
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            user = fetch_user_by_credentials(username, password)
            if user is None:
                flash("账号或密码错误。", "error")
            else:
                session["user_id"] = user["id"]
                default_project_id = default_project_id_for_user(int(user["id"]))
                if default_project_id is not None:
                    session["project_id"] = default_project_id
                flash(f"已登录为 {user['name']}。", "success")
                return redirect(next_url)
        return render_template("login.html", login_next=next_url)

    @app.route("/logout")
    def logout() -> Response:
        session.clear()
        flash("已退出登录。", "success")
        return redirect(url_for("login"))

    @app.route("/switch-project", methods=["POST"])
    def switch_project() -> Response:
        project_id = int(request.form.get("project_id", "0") or 0)
        if fetch_project(project_id):
            set_current_project(project_id)
        return redirect(request.referrer or url_for("bug_list"))

    @app.route("/")
    def home() -> Response:
        return redirect(url_for("bug_list"))

    @app.route("/bugs")
    def bug_list() -> str:
        filters = fetch_filters()
        page = request_page()
        bug_page = fetch_bug_page(filters, page)
        bug_select_users = fetch_bug_assignee_users()
        return render_template(
            "bug_list.html",
            bugs=bug_page["items"],
            bug_page=bug_page,
            users=bug_select_users,
            bug_assignee_users=bug_select_users,
            bug_versions=fetch_bug_versions(),
            filters=filters,
            bug_summary=fetch_bug_summary(filters),
            requirements=fetch_requirements(),
            cases=fetch_cases_for_project(),
            bug_form_values={"version": filters.get("version", "").strip()},
        )

    @app.route("/todos")
    def my_todo_page() -> str:
        my_todos = fetch_my_todos()
        summary = fetch_summary()
        summary["my_todo_count"] = len(my_todos)
        return render_template("my_todos.html", my_todos=my_todos, summary=summary)

    @app.route("/api/todos/zhengjingpei")
    def zhengjingpei_todos_api() -> Response | tuple[Response, int]:
        target_user = fetch_user_by_identity("zhengjingpei") or fetch_user_by_identity("郑敬佩")
        if target_user is None:
            return jsonify({"ok": False, "message": "未找到郑敬佩账号。"}), 404
        statuses, invalid_values = normalize_todo_status_filters(request.args.getlist("status"))
        if invalid_values:
            return (
                jsonify(
                    {
                        "ok": False,
                        "message": "待办状态参数无效。",
                        "invalid_statuses": invalid_values,
                        "valid_statuses": [{"code": code, "label": STATUS_LABELS[code]} for code in TODO_STATUS_CODES],
                    }
                ),
                400,
            )
        rows = fetch_user_todos_with_detail(int(target_user["id"]), statuses)
        todo_details = [serialize_todo_detail_with_relations(row) for row in rows]
        return jsonify(
            {
                "ok": True,
                "assignee": {
                    "id": int(target_user["id"]),
                    "name": target_user["name"] or "",
                    "username": target_user["username"] or "",
                    "role": target_user["role"] or "",
                },
                "filters": {
                    "statuses": [{"code": code, "label": STATUS_LABELS[code]} for code in statuses],
                },
                "count": len(rows),
                "todos": [serialize_todo_bug(row, detail) for row, detail in zip(rows, todo_details)],
                "todo_details": todo_details,
            }
        )

    @app.route("/api/todos/zhengjingpei/<int:bug_id>/status", methods=["POST"])
    def zhengjingpei_todo_status_api(bug_id: int) -> Response | tuple[Response, int]:
        target_user = fetch_user_by_identity("zhengjingpei") or fetch_user_by_identity("郑敬佩")
        if target_user is None:
            return jsonify({"ok": False, "message": "未找到郑敬佩账号。"}), 404

        bug = fetch_bug(bug_id)
        if bug is None:
            return jsonify({"ok": False, "message": "未找到对应的待办。"}), 404
        if int(bug["assignee_id"] or 0) != int(target_user["id"]):
            return jsonify({"ok": False, "message": "该数据当前不归属郑敬佩，不能通过此接口变更。"}), 403

        payload = request.get_json(silent=True) if request.is_json else None
        data = payload if isinstance(payload, dict) else {}
        raw_status = data.get("status", request.form.get("status", ""))
        selected_status = normalize_bug_status_value(raw_status)
        if selected_status is None or selected_status not in STATUS_LABELS:
            return (
                jsonify(
                    {
                        "ok": False,
                        "message": "目标状态无效。",
                        "valid_statuses": [{"code": code, "label": label} for code, label in STATUS_OPTIONS],
                    }
                ),
                400,
            )

        note = str(data.get("resolution_note", data.get("note", request.form.get("resolution_note", ""))) or "").strip()
        operator_name = str(data.get("operator_name", data.get("operator", "郑敬佩待办接口")) or "").strip() or "郑敬佩待办接口"
        db = get_db()
        try:
            new_status, new_assignee_id, new_previous_assignee_id, action_label, detail = apply_bug_action(
                db=db,
                bug=bug,
                action="change_status",
                operator_name=operator_name,
                note=note,
                selected_status=selected_status,
            )
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400

        now_text = current_time()
        db.execute(
            """
            UPDATE bugs
            SET status = ?, assignee_id = ?, previous_assignee_id = ?, resolution_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (new_status, new_assignee_id, new_previous_assignee_id, note or bug["resolution_note"], now_text, bug_id),
        )
        target_assignee = fetch_user(new_assignee_id)
        add_history(
            db,
            bug_id,
            action_label,
            detail,
            operator_name,
            environment_snapshot=bug["environment"] or "",
            status_snapshot=new_status,
            assignee_snapshot=target_assignee["name"] if target_assignee else bug["assignee_name"],
        )
        db.commit()
        bump_bug_sync_token()

        detail_row = fetch_todo_detail_bug(bug_id)
        todo_detail = serialize_todo_detail_with_relations(detail_row) if detail_row is not None else None
        return jsonify(
            {
                "ok": True,
                "message": "待办状态已更新。",
                "previous_status": {
                    "code": bug["status"] or "",
                    "label": STATUS_LABELS.get(str(bug["status"] or ""), str(bug["status"] or "")),
                },
                "current_status": {
                    "code": new_status,
                    "label": STATUS_LABELS.get(new_status, new_status),
                },
                "todo": serialize_todo_bug(detail_row, todo_detail) if detail_row is not None and todo_detail is not None else None,
                "todo_detail": todo_detail,
            }
        )

    @app.route("/notifications")
    def notification_center() -> str:
        state = request.args.get("state", "").strip()
        if state not in {"", "unread"}:
            state = ""
        user_id = int(g.current_user["id"])
        return render_template(
            "notifications.html",
            notifications=fetch_user_notifications(user_id, state=state),
            selected_state=state,
            unread_count=count_user_notifications(user_id, unread_only=True),
            total_count=count_user_notifications(user_id),
        )

    @app.route("/notifications/<int:notification_id>/open")
    def open_notification(notification_id: int) -> Response:
        user_id = int(g.current_user["id"])
        notification = fetch_notification(notification_id, user_id)
        if notification is None:
            flash("消息不存在或已删除。", "error")
            return redirect(url_for("notification_center"))
        mark_notification_read(notification_id, user_id)
        return redirect(local_redirect_target(notification["link_path"] or "", url_for("notification_center")))

    @app.route("/notifications/read-all", methods=["POST"])
    def mark_all_notifications_read_route() -> Response:
        marked_count = mark_all_notifications_read(int(g.current_user["id"]))
        flash(f"已将 {marked_count} 条消息标为已读。", "success")
        return redirect(url_for("notification_center"))

    @app.route("/cases")
    def case_library() -> str:
        selected_version = request.args.get("version", "").strip()
        selected_folder = request.args.get("folder", "").strip()
        documents = fetch_case_documents(selected_version)
        case_tree = build_case_tree(documents)
        folder_names = {item["name"] for item in case_tree}
        document_ids = {int(item["id"]) for item in documents}
        selected_id = int(request.args.get("document_id", "0") or 0)
        if selected_id not in document_ids:
            selected_id = 0
        requested_document = fetch_case_document(selected_id) if selected_id else None
        if not selected_folder and requested_document is not None:
            selected_folder = requested_document["folder_name"] or "测试用例"
        if selected_folder not in folder_names:
            selected_folder = case_tree[0]["name"] if case_tree else ""
        visible_documents = [
            item
            for item in documents
            if not selected_folder or (item["folder_name"] or "测试用例") == selected_folder
        ]
        visible_document_ids = {int(item["id"]) for item in visible_documents}
        if selected_id not in visible_document_ids:
            selected_id = int(visible_documents[0]["id"]) if visible_documents else 0
        selected_document = fetch_case_document(selected_id) if selected_id else None
        selected_document_count = 0
        if selected_document is not None:
            selected_document_count = count_document_cases(
                selected_document["project_id"],
                selected_document["version"],
                selected_document["folder_name"],
                selected_document["doc_name"],
            )
        selected_distribution = execution_distribution(
            project_id=selected_document["project_id"],
            version=selected_document["version"],
            folder_name=selected_document["folder_name"],
            doc_name=selected_document["doc_name"],
        ) if selected_document is not None else execution_distribution(version=selected_version)
        return render_template(
            "case_library.html",
            case_documents=documents,
            visible_case_documents=visible_documents,
            case_tree=case_tree,
            selected_document=selected_document,
            selected_document_count=selected_document_count,
            distribution=selected_distribution,
            selected_version=selected_version,
            selected_folder=selected_folder,
            case_versions=fetch_case_versions(),
        )

    @app.route("/cases/<int:document_id>")
    def case_document_detail(document_id: int) -> str | Response:
        bundle = fetch_case_document_bundle(document_id)
        if bundle is None:
            flash("未找到对应的在线文档。", "error")
            return redirect(url_for("case_library"))
        editable = can_edit_case_execution(bundle["document"])
        manageable = can_manage_case_document(bundle["document"])
        return render_template(
            "case_document_v2.html",
            case_document=bundle["document"],
            document_cases=bundle["cases"],
            document_columns=bundle["columns"],
            case_meta=bundle["meta"],
            owner_name=bundle["owner"],
            collaborators=bundle["collaborators"],
            platform_result_options=PLATFORM_RESULT_OPTIONS,
            can_edit_execution=editable,
            can_manage_document=manageable,
        )

    @app.route("/cases/<int:document_id>/update", methods=["POST"])
    def update_case_document(document_id: int) -> Response:
        bundle = fetch_case_document_bundle(document_id)
        if bundle is None:
            flash("未找到对应的在线文档。", "error")
            return redirect(url_for("case_library"))
        if not can_edit_case_execution(bundle["document"]):
            flash("仅登录用户可编辑在线文档。", "error")
            return redirect(url_for("case_document_detail", document_id=document_id))
        db = get_db()
        now = current_time()
        document_action = request.form.get("document_action", "save").strip() or "save"
        for item in bundle["cases"]:
            case_id = item["id"]
            case_no = case_document_form_value(request.form, "case_no", case_id, item["case_no"]) or str(item["case_no"] or "")
            priority_level = case_document_form_value(request.form, "priority_level", case_id, item["priority_level"])
            module_name = case_document_form_value(request.form, "module_name", case_id, item["module_name"])
            steps = case_document_form_value(request.form, "steps", case_id, item["steps"])
            expected_result = case_document_form_value(request.form, "expected_result", case_id, item["expected_result"])
            ios_result = case_document_form_value(request.form, "ios_result", case_id, item["ios_result"])
            android_result = case_document_form_value(request.form, "android_result", case_id, item["android_result"])
            h5_result = case_document_form_value(request.form, "h5_result", case_id, item["h5_result"])
            remark = case_document_form_value(request.form, "remark", case_id, item["remark"])
            executor = case_document_form_value(request.form, "executor", case_id, item["executor"])
            execute_status = normalize_case_status(ios_result, android_result, h5_result)
            db.execute(
                """
                UPDATE test_cases
                SET case_no = ?, priority_level = ?, module_name = ?, steps = ?, expected_result = ?,
                    ios_result = ?, android_result = ?, h5_result = ?, remark = ?, executor = ?, execute_status = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    case_no,
                    priority_level,
                    module_name,
                    steps,
                    expected_result,
                    ios_result,
                    android_result,
                    h5_result,
                    remark,
                    executor,
                    execute_status,
                    now,
                    case_id,
                ),
            )
        save_case_document_dynamic_cells(
            document_columns=bundle["columns"],
            document_cases=bundle["cases"],
            form=request.form,
            db=db,
            now=now,
        )

        if document_action == "add_row":
            create_case_document_row(bundle["document"], db=db)
            message = "已新增一行。"
            category = "success"
        elif document_action == "add_column":
            new_column_name = request.form.get("new_column_name", "").strip()
            if new_column_name:
                create_case_document_column(bundle["document"], new_column_name, db=db)
                message = "已新增一列。"
                category = "success"
            else:
                message = "文档已保存，请先填写列名。"
                category = "error"
        else:
            message = "在线文档已保存。"
            category = "success"

        db.commit()
        flash(message, category)
        return redirect(url_for("case_document_detail", document_id=document_id))

    @app.route("/cases/<int:document_id>/autosave", methods=["POST"])
    def autosave_case_document(document_id: int) -> Response:
        bundle = fetch_case_document_bundle(document_id)
        if bundle is None:
            return jsonify({"ok": False, "message": "未找到对应的在线文档。"}), 404
        if not can_edit_case_execution(bundle["document"]):
            return jsonify({"ok": False, "message": "仅登录用户可编辑在线文档。"}), 403
        try:
            case_id = int(request.form.get("case_id", "0") or 0)
        except ValueError:
            case_id = 0
        field_name = request.form.get("field", "").strip()
        raw_value = request.form.get("value", "")
        if not case_id or not field_name:
            return jsonify({"ok": False, "message": "保存参数不完整。"}), 400

        db = get_db()
        try:
            payload = update_case_document_cell(
                bundle=bundle,
                case_id=case_id,
                field_name=field_name,
                raw_value=raw_value,
                db=db,
                now=current_time(),
            )
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400
        db.commit()
        payload["ok"] = True
        payload["message"] = "已实时保存。"
        return jsonify(payload)

    @app.route("/cases/<int:document_id>/items/<int:case_id>/delete", methods=["POST"])
    def delete_case_item_route(document_id: int, case_id: int) -> Response:
        bundle = fetch_case_document_bundle(document_id)
        if bundle is None:
            flash("未找到对应的在线文档。", "error")
            return redirect(url_for("case_library"))
        if not can_manage_case_document(bundle["document"]):
            flash("仅管理员或文档创建人可删除用例。", "error")
            return redirect(url_for("case_document_detail", document_id=document_id))
        case_item = fetch_case(case_id)
        if case_item is None:
            flash("未找到对应的用例。", "error")
            return redirect(url_for("case_document_detail", document_id=document_id))
        same_document = (
            int(case_item["project_id"] or 0) == int(bundle["document"]["project_id"] or 0)
            and str(case_item["version"] or "") == str(bundle["document"]["version"] or "")
            and str(case_item["folder_name"] or "") == str(bundle["document"]["folder_name"] or "")
            and str(case_item["doc_name"] or "") == str(bundle["document"]["doc_name"] or "")
        )
        if not same_document:
            flash("当前用例不属于这个在线文档。", "error")
            return redirect(url_for("case_document_detail", document_id=document_id))
        deleted, next_document_id, version = delete_case_item(case_id)
        if deleted <= 0:
            flash("用例删除失败。", "error")
            return redirect(url_for("case_document_detail", document_id=document_id))
        flash("用例已删除。", "success")
        if case_id != document_id:
            return redirect(url_for("case_document_detail", document_id=document_id))
        if next_document_id is not None:
            return redirect(url_for("case_document_detail", document_id=next_document_id))
        if version:
            return redirect(url_for("case_library", version=version))
        return redirect(url_for("case_library"))

    @app.route("/cases/<int:document_id>/rename", methods=["POST"])
    def rename_case_document_route(document_id: int) -> Response:
        version_filter = request.form.get("version_filter", "").strip()
        document = fetch_case_document(document_id)
        if document is None:
            flash("未找到对应的在线文档。", "error")
            return redirect(url_for("case_library", version=version_filter) if version_filter else url_for("case_library"))
        if not can_manage_case_document(document):
            flash("仅管理员或文档创建人可修改在线文档。", "error")
            return redirect(url_for("case_library", document_id=document_id, version=version_filter) if version_filter else url_for("case_library", document_id=document_id))
        version = request.form.get("version", "").strip()
        folder_name = request.form.get("folder_name", "").strip() or "测试用例"
        doc_name = request.form.get("doc_name", "").strip()
        if not version or not doc_name:
            flash("在线文档版本和名称不能为空。", "error")
        else:
            renamed = rename_case_document(document_id, version, folder_name, doc_name)
            flash("在线文档已更新。" if renamed > 0 else "在线文档更新失败。", "success" if renamed > 0 else "error")
        target_version = version or version_filter
        return redirect(url_for("case_library", document_id=document_id, version=target_version) if target_version else url_for("case_library", document_id=document_id))

    @app.route("/cases/upload", methods=["POST"])
    def upload_cases() -> Response:
        version_filter = request.form.get("version_filter", "").strip()
        folder_name = request.form.get("folder_name", "").strip()

        def redirect_to_case_library() -> Response:
            redirect_params: dict[str, object] = {}
            if version_filter:
                redirect_params["version"] = version_filter
            if folder_name:
                redirect_params["folder"] = folder_name
            return redirect(url_for("case_library", **redirect_params))

        file = request.files.get("excel_file")
        if not folder_name:
            flash("请先选择上传文件夹。", "error")
            return redirect_to_case_library()
        if file is None or not file.filename:
            flash("请选择 Excel 文件。", "error")
            return redirect_to_case_library()
        workbook = openpyxl.load_workbook(file, data_only=True)
        db = get_db()
        project_id = current_project_id()
        now = current_time()
        imported = 0
        default_doc_name = Path(file.filename).stem
        default_version = default_doc_name.split("-")[0] if "-" in default_doc_name else ""
        creator_id = int(g.current_user["id"]) if g.current_user is not None else None
        workbook_sheets = [sheet for sheet in workbook.worksheets if sheet.max_row > 0 and sheet.max_column > 0]
        multi_sheet_mode = len(workbook_sheets) > 1
        scanned_sheet_count = 0
        imported_sheet_names: list[str] = []
        imported_doc_names: list[str] = []
        ensure_case_folder(folder_name, db)

        db.execute(
            """
            DELETE FROM test_cases
            WHERE project_id = ?
                AND source_type = 'Excel上传'
                AND folder_name = ?
                AND (
                    COALESCE(doc_name, '') = COALESCE(?, '')
                    OR COALESCE(doc_name, '') LIKE ?
                )
            """,
            (project_id, folder_name, default_doc_name, f"{default_doc_name} / %"),
        )

        for sheet in workbook_sheets:
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue
            scanned_sheet_count += 1
            meta_info = extract_sheet_meta_info(sheet)
            header_row_index, header_mapping = find_excel_header_index(sheet)
            current_module_name = ""
            seen_case_nos: set[str] = set()
            sheet_imported_count = 0
            sheet_doc_name = default_doc_name if not multi_sheet_mode else f"{default_doc_name} / {sheet.title}"
            sheet_image_cells = collect_sheet_image_cells(sheet)

            def insert_case_row(
                *,
                case_no: str,
                version: str,
                title: str,
                priority_level: str,
                module_name: str,
                steps: str,
                expected_result: str,
                actual_result: str,
                ios_result: str,
                android_result: str,
                h5_result: str,
                remark: str,
                executor: str,
                execute_status: str,
            ) -> None:
                nonlocal imported, sheet_imported_count
                case_key = case_no.strip()
                if not case_key or case_key in seen_case_nos:
                    return
                seen_case_nos.add(case_key)
                db.execute(
                    """
                    INSERT INTO test_cases (
                        project_id, version, folder_name, doc_name, case_no, title, priority_level, module_name,
                        steps, expected_result, actual_result, ios_result, android_result, h5_result, remark,
                        executor, environment_info, device_info, network_info,
                        source_type, doc_link, execute_status, creator_id, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project_id,
                        version or default_version,
                        folder_name,
                        sheet_doc_name,
                        case_no,
                        title,
                        priority_level or "P1",
                        module_name,
                        steps,
                        expected_result,
                        actual_result,
                        ios_result,
                        android_result,
                        h5_result,
                        remark,
                        executor,
                        meta_info["environment_info"],
                        meta_info["device_info"],
                        meta_info["network_info"],
                        "Excel上传",
                        "Excel上传",
                        execute_status,
                        creator_id,
                        now,
                        now,
                    ),
                )
                imported += 1
                sheet_imported_count += 1

            if header_row_index is not None:
                for row_index, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
                    if not row:
                        continue

                    def mapped_text(field_name: str) -> str:
                        column_index = header_mapping.get(field_name)
                        if column_index is None or column_index >= len(row):
                            return ""
                        return normalize_excel_text(row[column_index])

                    def mapped_text_with_images(field_name: str) -> str:
                        column_index = header_mapping.get(field_name)
                        return append_excel_image_marker(
                            mapped_text(field_name),
                            count_excel_images_in_cell(sheet_image_cells, row_index, column_index),
                        )

                    case_no = mapped_text("case_no")
                    if not case_no:
                        module_name_raw = mapped_text("module_name")
                        if module_name_raw:
                            current_module_name = module_name_raw
                        continue

                    title = mapped_text("title")
                    priority_level = mapped_text("priority_level") or "P1"
                    module_name_raw = mapped_text("module_name")
                    steps = mapped_text_with_images("steps")
                    expected_result = mapped_text_with_images("expected_result")
                    actual_result = mapped_text_with_images("actual_result")
                    remark = mapped_text_with_images("remark")
                    executor = mapped_text("executor")
                    version = infer_case_version(case_no, mapped_text("version")) or default_version
                    mapped_image_count = sum(
                        count_excel_images_in_cell(sheet_image_cells, row_index, header_mapping.get(field_name))
                        for field_name in ("steps", "expected_result", "actual_result", "remark")
                    )
                    unmapped_image_count = max(0, count_excel_images_in_row(sheet_image_cells, row_index) - mapped_image_count)
                    if unmapped_image_count:
                        remark = append_excel_image_marker(remark, unmapped_image_count)

                    execute_status_raw = mapped_text("execute_status")
                    ios_result = normalize_platform_result(mapped_text("ios_result"))
                    android_result = normalize_platform_result(mapped_text("android_result"))
                    h5_result = normalize_platform_result(mapped_text("h5_result"))
                    if not has_meaningful_case_content(
                        title=title,
                        module_name=module_name_raw,
                        steps=steps,
                        expected_result=expected_result,
                        actual_result=actual_result,
                        remark=remark,
                        executor=executor,
                        execute_status=execute_status_raw,
                        ios_result=ios_result,
                        android_result=android_result,
                        h5_result=h5_result,
                    ):
                        continue

                    if module_name_raw:
                        current_module_name = module_name_raw
                    module_name = current_module_name or (title if title and not looks_like_step_text(title) else "")
                    title = title or module_name or case_no

                    if any([ios_result, android_result, h5_result]):
                        execute_status = normalize_case_status(ios_result, android_result, h5_result)
                        if execute_status == "未测" and execute_status_raw:
                            execute_status = normalize_case_execute_status(execute_status_raw)[0]
                    else:
                        execute_status = normalize_case_execute_status(execute_status_raw)[0]

                    if not any([title, steps, expected_result, actual_result, remark, module_name]):
                        continue

                    insert_case_row(
                        case_no=case_no,
                        version=version,
                        title=title,
                        priority_level=priority_level,
                        module_name=module_name,
                        steps=steps,
                        expected_result=expected_result,
                        actual_result=actual_result,
                        ios_result=ios_result,
                        android_result=android_result,
                        h5_result=h5_result,
                        remark=remark,
                        executor=executor,
                        execute_status=execute_status,
                    )
                if sheet_imported_count > 0:
                    imported_sheet_names.append(sheet.title)
                    imported_doc_names.append(sheet_doc_name)
                continue

            start_row = 4 if sheet.max_row >= 4 else 1
            for row_index, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not row:
                    continue
                case_no = normalize_excel_text(row[0]) if len(row) > 0 else ""
                if not case_no:
                    continue
                priority_level = normalize_excel_text(row[1]) if len(row) > 1 else "P1"
                module_name_raw = normalize_excel_text(row[2]) if len(row) > 2 else ""
                title = normalize_excel_text(row[2]) if len(row) > 2 else ""
                steps = append_excel_image_marker(normalize_excel_text(row[3]) if len(row) > 3 else "", count_excel_images_in_cell(sheet_image_cells, row_index, 3))
                expected_result = append_excel_image_marker(normalize_excel_text(row[4]) if len(row) > 4 else "", count_excel_images_in_cell(sheet_image_cells, row_index, 4))
                actual_result = append_excel_image_marker(normalize_excel_text(row[5]) if len(row) > 5 else "", count_excel_images_in_cell(sheet_image_cells, row_index, 5))
                ios_result = normalize_platform_result(row[5] if len(row) > 5 else "")
                android_result = normalize_platform_result(row[6] if len(row) > 6 else "")
                h5_result = normalize_platform_result(row[7] if len(row) > 7 else "")
                remark = append_excel_image_marker(normalize_excel_text(row[8]) if len(row) > 8 else "", count_excel_images_in_cell(sheet_image_cells, row_index, 8))
                executor = normalize_excel_text(row[9]) if len(row) > 9 else ""
                mapped_image_count = sum(
                    count_excel_images_in_cell(sheet_image_cells, row_index, column_index)
                    for column_index in (3, 4, 5, 8)
                )
                unmapped_image_count = max(0, count_excel_images_in_row(sheet_image_cells, row_index) - mapped_image_count)
                if unmapped_image_count:
                    remark = append_excel_image_marker(remark, unmapped_image_count)
                if not has_meaningful_case_content(
                    title=title,
                    module_name=module_name_raw,
                    steps=steps,
                    expected_result=expected_result,
                    actual_result=actual_result,
                    remark=remark,
                    executor=executor,
                    execute_status="",
                    ios_result=ios_result,
                    android_result=android_result,
                    h5_result=h5_result,
                ):
                    continue
                if module_name_raw:
                    current_module_name = module_name_raw
                module_name = current_module_name
                execute_status = normalize_case_status(ios_result, android_result, h5_result)

                insert_case_row(
                    case_no=case_no,
                    version=infer_case_version(case_no, default_version),
                    title=title or case_no,
                    priority_level=priority_level,
                    module_name=module_name,
                    steps=steps,
                    expected_result=expected_result,
                    actual_result=actual_result,
                    ios_result=ios_result,
                    android_result=android_result,
                    h5_result=h5_result,
                    remark=remark,
                    executor=executor,
                    execute_status=execute_status,
                )
            if sheet_imported_count > 0:
                imported_sheet_names.append(sheet.title)
                imported_doc_names.append(sheet_doc_name)

        if scanned_sheet_count == 0:
            flash("Excel 中没有可读取的工作表，无法导入。", "error")
            db.commit()
            return redirect_to_case_library()
        if imported == 0:
            flash("未识别到可导入的用例数据，请检查各工作表表头中是否包含“用例编号”等核心字段。", "error")
            db.commit()
            return redirect_to_case_library()
        repaired_count = repair_misaligned_excel_cases(db, imported_doc_names)
        db.commit()
        flash(
            f"已同步 {imported} 条用例，来自 {len(imported_sheet_names)} 个工作表。"
            + (f" 已自动修正 {repaired_count} 条错位数据。" if repaired_count > 0 else ""),
            "success",
        )
        return redirect_to_case_library()

    @app.route("/cases/manage", methods=["POST"])
    def manage_case_library() -> Response:
        action = request.form.get("action", "").strip()
        folder_name = request.form.get("folder_name", "").strip()
        doc_name = request.form.get("doc_name", "").strip()
        document_id = int(request.form.get("document_id", "0") or 0)
        version_filter = request.form.get("version_filter", "").strip()
        open_folder = request.form.get("open_folder", "").strip()
        redirect_folder = open_folder
        if action == "create_folder":
            if not folder_name:
                flash("请输入文件夹名称。", "error")
            else:
                created = create_case_folder(folder_name)
                redirect_folder = folder_name
                flash("文件夹已创建。" if created else "文件夹已存在。", "success" if created else "info")
        elif action == "create_document":
            if not doc_name:
                flash("请输入在线文档名称。", "error")
            else:
                create_case_document(folder_name or "测试用例", doc_name)
                redirect_folder = folder_name or "测试用例"
                flash("在线文档已创建。", "success")
        elif action == "delete_folder":
            if not folder_name:
                flash("未找到要删除的文件夹。", "error")
            elif not is_admin():
                flash("仅管理员可删除整个文件夹。", "error")
            else:
                delete_case_folder(folder_name)
                redirect_folder = ""
                flash("文件夹已删除。", "success")
        elif action == "delete_document":
            if document_id <= 0:
                flash("未找到要删除的在线文档。", "error")
            else:
                document = fetch_case_document(document_id)
                if not can_manage_case_document(document):
                    flash("仅管理员或文档创建人可删除在线文档。", "error")
                else:
                    deleted = delete_case_document(document_id)
                    if deleted > 0:
                        flash("在线文档已删除。", "success")
                    else:
                        flash("在线文档删除失败。", "error")
        elif action == "move_document":
            target_folder = folder_name or "测试用例"
            if document_id <= 0:
                flash("未找到要移动的在线文档。", "error")
            else:
                document = fetch_case_document(document_id)
                if document is not None and int(document["project_id"] or 0) != int(current_project_id() or 0):
                    flash("只能移动当前项目下的在线文档。", "error")
                elif not can_manage_case_document(document):
                    flash("仅管理员或文档创建人可移动在线文档。", "error")
                else:
                    moved = move_case_document(document_id, target_folder)
                    if moved > 0:
                        redirect_folder = target_folder
                        flash("在线文档已移动。", "success")
                    elif moved < 0:
                        flash("目标文件夹里已有同名在线文档，未移动。", "error")
                    else:
                        redirect_folder = target_folder
                        flash("在线文档已在该文件夹中。", "success")
        redirect_params: dict[str, object] = {}
        if version_filter:
            redirect_params["version"] = version_filter
        if redirect_folder:
            redirect_params["folder"] = redirect_folder
        if action == "move_document" and document_id > 0:
            redirect_params["document_id"] = document_id
        return redirect(url_for("case_library", **redirect_params))

    @app.route("/requirements")
    def requirement_library() -> str:
        filters = {
            "keyword": request.args.get("keyword", "").strip(),
            "version": request.args.get("version", "").strip(),
        }
        try:
            page = int(request.args.get("page", "1") or "1")
        except ValueError:
            page = 1
        requirement_page = fetch_requirement_page(filters, page)
        return render_template(
            "requirements.html",
            requirements=requirement_page["items"],
            requirement_page=requirement_page,
            filters=filters,
            requirement_versions=fetch_requirement_versions(),
            requirement_summary=fetch_requirement_summary(),
        )

    @app.route("/requirements/<int:requirement_id>")
    def requirement_detail(requirement_id: int) -> str | Response:
        requirement = fetch_requirement(requirement_id)
        if requirement is None:
            flash("未找到对应需求。", "error")
            return redirect(url_for("requirement_library"))
        return render_template(
            "requirement_detail.html",
            requirement=requirement,
            linked_bugs=fetch_requirement_bugs(requirement_id),
        )

    @app.route("/requirements/create", methods=["POST"])
    def create_requirement() -> Response:
        code = request.form.get("code", "").strip()
        title = request.form.get("title", "").strip()
        version = request.form.get("version", "").strip()
        status = request.form.get("status", "").strip() or "pending"
        priority = request.form.get("priority", "").strip() or "中"
        description = request.form.get("description", "").strip()
        acceptance_criteria = request.form.get("acceptance_criteria", "").strip()
        requirement_doc_link = request.form.get("requirement_doc_link", "").strip()
        design_doc_link = request.form.get("design_doc_link", "").strip()
        project_id = current_project_id()
        code = code or generate_requirement_code(project_id)
        if not title or not version:
            flash("请至少填写需求标题和版本。", "error")
        elif get_db().execute("SELECT 1 FROM requirements WHERE project_id = ? AND code = ?", (project_id, code)).fetchone():
            flash("当前项目下需求编号已存在。", "error")
        else:
            now = current_time()
            cursor = get_db().execute(
                """
                INSERT INTO requirements (
                    project_id, code, title, version, status, priority, description,
                    acceptance_criteria, requirement_doc_link, design_doc_link,
                    creator_id, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    code,
                    title,
                    version,
                    status,
                    priority,
                    description,
                    acceptance_criteria,
                    requirement_doc_link,
                    design_doc_link,
                    int(g.current_user["id"]),
                    now,
                    now,
                ),
            )
            get_db().commit()
            flash("需求已创建。", "success")
            return redirect(url_for("requirement_detail", requirement_id=cursor.lastrowid))
        return redirect(url_for("requirement_library"))

    @app.route("/requirements/<int:requirement_id>/update", methods=["POST"])
    def update_requirement(requirement_id: int) -> Response:
        requirement = fetch_requirement(requirement_id)
        if requirement is None:
            flash("未找到对应需求。", "error")
            return redirect(url_for("requirement_library"))
        if not can_manage_requirement(requirement):
            flash("仅管理员或需求创建人可编辑需求。", "error")
            return redirect(url_for("requirement_library"))
        code = request.form.get("code", "").strip()
        title = request.form.get("title", "").strip()
        version = request.form.get("version", "").strip()
        status = request.form.get("status", "").strip() or str(requirement["status"] or "pending")
        priority = request.form.get("priority", "").strip() or str(requirement["priority"] or "中")
        description = request.form.get("description", "").strip() if "description" in request.form else str(requirement["description"] or "")
        acceptance_criteria = request.form.get("acceptance_criteria", "").strip() if "acceptance_criteria" in request.form else str(requirement["acceptance_criteria"] or "")
        requirement_doc_link = request.form.get("requirement_doc_link", "").strip()
        design_doc_link = request.form.get("design_doc_link", "").strip()
        next_url = request.form.get("next", "").strip() or url_for("requirement_detail", requirement_id=requirement_id)
        edit_url = url_for("requirement_detail", requirement_id=requirement_id, edit="1")
        code = code or str(requirement["code"] or "").strip() or generate_requirement_code(current_project_id())
        if not title or not version:
            flash("请至少填写需求标题和版本。", "error")
            return redirect(edit_url)
        elif get_db().execute(
            "SELECT 1 FROM requirements WHERE project_id = ? AND code = ? AND id != ?",
            (current_project_id(), code, requirement_id),
        ).fetchone():
            flash("当前项目下需求编号已存在。", "error")
            return redirect(edit_url)
        else:
            get_db().execute(
                """
                UPDATE requirements
                SET code = ?, title = ?, version = ?, status = ?, priority = ?, description = ?,
                    acceptance_criteria = ?, requirement_doc_link = ?, design_doc_link = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    code,
                    title,
                    version,
                    status,
                    priority,
                    description,
                    acceptance_criteria,
                    requirement_doc_link,
                    design_doc_link,
                    current_time(),
                    requirement_id,
                ),
            )
            get_db().commit()
            flash("需求已更新。", "success")
        return redirect(next_url)

    @app.route("/requirements/<int:requirement_id>/delete", methods=["POST"])
    def delete_requirement(requirement_id: int) -> Response:
        requirement = fetch_requirement(requirement_id)
        next_url = request.form.get("next", "").strip() or url_for("requirement_library")
        if requirement is None:
            flash("未找到对应需求。", "error")
            return redirect(next_url)
        if not can_manage_requirement(requirement):
            flash("仅管理员或需求创建人可删除需求。", "error")
            return redirect(next_url)
        bug_ref_count = int(get_db().execute("SELECT COUNT(*) AS count FROM bugs WHERE requirement_id = ?", (requirement_id,)).fetchone()["count"])
        if bug_ref_count > 0:
            flash("该需求已被 Bug 关联，暂不可删除。", "error")
            return redirect(next_url)
        get_db().execute("DELETE FROM requirements WHERE id = ?", (requirement_id,))
        get_db().commit()
        flash("需求已删除。", "success")
        return redirect(next_url)

    @app.route("/bugs/new", methods=["GET", "POST"])
    def create_bug() -> str | Response:
        db = get_db()
        users = fetch_bug_assignee_users()
        requirements = fetch_requirements()
        cases = fetch_cases_for_project()
        back_url = local_back_url(url_for("bug_list"))
        prefill_case = None
        if request.method == "GET":
            prefill_case_id_text = request.args.get("case_id", "").strip()
            if prefill_case_id_text.isdigit():
                prefill_case = fetch_case(int(prefill_case_id_text))
                if prefill_case is not None and int(prefill_case["project_id"] or 0) != int(current_project_id() or 0):
                    prefill_case = None
        if request.method == "POST":
            bug_form = normalize_bug_form(request.form, request.files)
            title = bug_form["title"]
            version = bug_form["version"]
            module = bug_form["module"]
            platform = bug_form["platform"]
            severity = bug_form["severity"]
            priority = bug_form["priority"]
            assignee_id = bug_form["assignee_id"]
            requirement_id = bug_form["requirement_id"]
            case_id = bug_form["case_id"]
            environment = bug_form["environment"]
            description = bug_form["description"]
            expected_result = bug_form["expected_result"]
            actual_result = bug_form["actual_result"]
            attachments = bug_form["attachments"]
            inline_images = bug_form["inline_images"]
            inline_image_sources = bug_form["inline_image_sources"]
            assignee_user = fetch_bug_assignee_user(assignee_id)
            if not all([title, version, module, platform, severity, assignee_id, description]):
                if wants_json_response():
                    return jsonify({"ok": False, "message": "请完整填写必填项。"}), 400
                flash("请完整填写必填项。", "error")
            elif assignee_user is None:
                if wants_json_response():
                    return jsonify({"ok": False, "message": "缺陷处理人不能选择 admin 账号。"}), 400
                flash("缺陷处理人不能选择 admin 账号。", "error")
            else:
                bug_id = insert_bug(
                    db=db,
                    title=title,
                    project_id=current_project_id(),
                    version=version,
                    module=module,
                    platform=platform,
                    severity=severity,
                    priority=priority,
                    status="open",
                    assignee_id=int(assignee_user["id"]),
                    creator_id=int(g.current_user["id"]),
                    previous_assignee_id=int(assignee_user["id"]),
                    requirement_id=int(requirement_id) if requirement_id else None,
                    case_id=int(case_id) if case_id else None,
                    environment=environment,
                    description=description,
                    expected_result=expected_result,
                    actual_result=actual_result,
                    resolution_note="",
                )
                assignee_name = assignee_user["name"]
                saved_names = save_bug_attachments(db, bug_id, attachments)
                saved_inline_names = save_bug_attachments(db, bug_id, inline_images, inline_image_sources)
                detail = f"{g.current_user['name']} 创建缺陷并指派给 {assignee_name}"
                if saved_names:
                    detail += f"；上传附件 {len(saved_names)} 个"
                if saved_inline_names:
                    detail += f"；插入正文图片 {len(saved_inline_names)} 张"
                add_history(
                    db,
                    bug_id,
                    "创建缺陷",
                    detail,
                    g.current_user["name"],
                    environment_snapshot=environment,
                    status_snapshot="open",
                    assignee_snapshot=assignee_name,
                )
                db.commit()
                bump_bug_sync_token()
                group_notify_sent, group_notify_message = maybe_send_new_bug_group_notification(
                    bug_id=bug_id,
                    operator_name=str(g.current_user["name"] or ""),
                )
                should_show_group_notify_result = (
                    group_notify_sent
                    or group_notify_message != "当前项目未开启新建 Bug 群通知。"
                )
                if wants_json_response():
                    response_message = "创建成功"
                    if should_show_group_notify_result:
                        response_message = (
                            response_message + f"，{group_notify_message}"
                            if group_notify_sent
                            else response_message + f"，但新建Bug群通知未发送：{group_notify_message}"
                        )
                    return jsonify(
                        {
                            "ok": True,
                            "message": response_message,
                            "bug_id": bug_id,
                            "redirect_url": url_for("bug_detail", bug_id=bug_id, next=back_url),
                        }
                    )
                flash("Bug 已创建。", "success")
                if should_show_group_notify_result:
                    flash(
                        f"新建Bug群通知已发送，{group_notify_message}"
                        if group_notify_sent
                        else f"新建Bug群通知未发送：{group_notify_message}",
                        "success" if group_notify_sent else "error",
                    )
                return redirect(url_for("bug_detail", bug_id=bug_id, next=back_url))
        return render_template(
            "bug_form.html",
            users=users,
            requirements=requirements,
            cases=cases,
            back_url=back_url,
            bug_form_values=locals().get("bug_form", build_bug_form_prefill_from_request(prefill_case)),
        )

    @app.route("/bugs/<int:bug_id>")
    def bug_detail(bug_id: int) -> str | Response:
        bug = fetch_bug(bug_id)
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        back_url = local_back_url(url_for("bug_list"))
        active_tab = request.args.get("tab", "detail").strip() or "detail"
        if active_tab not in {"detail", "process", "history"}:
            active_tab = "detail"
        history = fetch_bug_history(bug_id)
        comments = fetch_bug_comments(bug_id)
        attachments = fetch_bug_attachments(bug_id)
        attachments_by_field, general_attachments = group_bug_attachments(attachments)
        unread_comment_count = count_unread_bug_comment_notifications(int(g.current_user["id"]), bug_id)
        mention_users = fetch_users()
        comment_mention_states = fetch_comment_mention_states(bug_id)
        return render_template(
            "bug_detail.html",
            bug=bug,
            history=history,
            comments=comments,
            comment_threads=build_bug_comment_threads(comments, mention_users, comment_mention_states),
            mention_users=mention_users,
            related_mention_users=build_bug_mention_users(bug, comments),
            unread_comment_count=unread_comment_count,
            users=fetch_bug_assignee_users(),
            requirements=fetch_requirements(),
            cases=fetch_cases_for_project(),
            attachments=attachments,
            attachments_by_field=attachments_by_field,
            general_attachments=general_attachments,
            active_tab=active_tab,
            back_url=back_url,
        )

    @app.route("/bugs/<int:bug_id>/comments/read", methods=["POST"])
    def mark_bug_comments_read(bug_id: int) -> Response:
        bug = fetch_bug(bug_id)
        if bug is None:
            return jsonify({"ok": False, "message": "未找到对应的 Bug。"}), 404
        marked_count = mark_bug_comment_notifications_read(int(g.current_user["id"]), bug_id)
        return jsonify(
            {
                "ok": True,
                "marked_count": marked_count,
                "unread_count": count_user_notifications(int(g.current_user["id"]), unread_only=True),
            }
        )

    @app.route("/bugs/<int:bug_id>/comments", methods=["POST"])
    def add_bug_comment(bug_id: int) -> Response:
        bug = fetch_bug(bug_id)
        default_target = url_for("bug_detail", bug_id=bug_id, tab="detail") + "#bug-comments"
        redirect_target = local_redirect_target(request.form.get("redirect_to", ""), default_target)
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        content = request.form.get("content", "").strip()
        parent_id_text = request.form.get("parent_id", "").strip()
        parent_id = int(parent_id_text) if parent_id_text.isdigit() else None
        if not content:
            flash("评论内容不能为空。", "error")
            return redirect(redirect_target)
        db = get_db()
        if parent_id is not None:
            parent_comment = db.execute(
                "SELECT id FROM bug_comments WHERE id = ? AND bug_id = ?",
                (parent_id, bug_id),
            ).fetchone()
            if parent_comment is None:
                parent_id = None
        now = current_time()
        cursor = db.execute(
            """
            INSERT INTO bug_comments (bug_id, user_id, parent_id, author_name, content, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                bug_id,
                int(g.current_user["id"]),
                parent_id,
                str(g.current_user["name"] or ""),
                content,
                now,
                now,
            ),
        )
        db.commit()
        notify_bug_comment_recipients(bug, int(cursor.lastrowid), content)
        flash("回复已发布。" if parent_id else "评论已发布。", "success")
        return redirect(redirect_target)

    @app.route("/bugs/<int:bug_id>/comments/<int:comment_id>/delete", methods=["POST"])
    def delete_bug_comment(bug_id: int, comment_id: int) -> Response:
        bug = fetch_bug(bug_id)
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        comment = fetch_bug_comment(comment_id, bug_id)
        if comment is None:
            flash("评论不存在或已删除。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail"))
        if not can_manage_bug_comment(comment):
            flash("仅评论人本人或管理员可删除评论。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail"))

        comment_ids = collect_comment_branch_ids(bug_id, comment_id)
        placeholders = ",".join("?" for _ in comment_ids)
        get_db().execute(f"DELETE FROM notifications WHERE comment_id IN ({placeholders})", comment_ids)
        get_db().execute(f"DELETE FROM bug_comments WHERE id IN ({placeholders})", comment_ids)
        get_db().commit()
        flash("评论及回复已删除。" if len(comment_ids) > 1 else "评论已删除。", "success")
        return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail") + "#bug-comments")

    @app.route("/bugs/<int:bug_id>/edit", methods=["POST"])
    def edit_bug(bug_id: int) -> Response:
        db = get_db()
        bug = fetch_bug(bug_id)
        back_url = local_back_url(url_for("bug_list"))
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        if not can_manage_bug(bug):
            flash("仅管理员或创建人可编辑该 Bug。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id, next=back_url))

        bug_form = normalize_bug_form(request.form, request.files)
        title = bug_form["title"]
        version = bug_form["version"]
        module = bug_form["module"]
        platform = bug_form["platform"]
        severity = bug_form["severity"]
        priority = bug_form["priority"]
        assignee_id = bug_form["assignee_id"]
        requirement_id = bug_form["requirement_id"]
        case_id = bug_form["case_id"]
        environment = bug_form["environment"]
        description = bug_form["description"]
        expected_result = bug_form["expected_result"]
        actual_result = bug_form["actual_result"]
        attachments = bug_form["attachments"]
        inline_images = bug_form["inline_images"]
        inline_image_sources = bug_form["inline_image_sources"]
        assignee_user = fetch_bug_assignee_user(assignee_id)

        if not all([title, version, module, platform, severity, assignee_id, description]):
            flash("请完整填写必填项。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail", edit="1", next=back_url))
        if assignee_user is None:
            flash("缺陷处理人不能选择 admin 账号。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail", edit="1", next=back_url))

        saved_names = save_bug_attachments(db, bug_id, attachments)
        saved_inline_names = save_bug_attachments(db, bug_id, inline_images, inline_image_sources)
        previous_severity = str(bug["severity"] or "")
        previous_assignee_id = int(bug["assignee_id"] or 0)
        next_assignee_id = int(assignee_user["id"])
        next_previous_assignee_id = derive_previous_assignee_id_for_bug(bug, next_assignee_id)
        db.execute(
            """
            UPDATE bugs
            SET title = ?, version = ?, module = ?, platform = ?, severity = ?, priority = ?, assignee_id = ?, previous_assignee_id = ?,
                requirement_id = ?, case_id = ?, environment = ?, description = ?, expected_result = ?,
                actual_result = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                title,
                version,
                module,
                platform,
                severity,
                priority,
                next_assignee_id,
                next_previous_assignee_id,
                int(requirement_id) if requirement_id else None,
                int(case_id) if case_id else None,
                environment,
                description,
                expected_result,
                actual_result,
                current_time(),
                bug_id,
            ),
        )
        detail = f"{g.current_user['name']} 编辑了缺陷信息"
        if saved_names:
            detail += f"；新增附件 {len(saved_names)} 个"
        if saved_inline_names:
            detail += f"；插入正文图片 {len(saved_inline_names)} 张"
        add_history(
            db,
            bug_id,
            "编辑缺陷",
            detail,
            g.current_user["name"],
            environment_snapshot=environment,
            status_snapshot=bug["status"],
            assignee_snapshot=assignee_user["name"],
        )
        db.commit()
        bump_bug_sync_token()
        flash("Bug 信息已更新。", "success")
        if (
            severity == MAIL_NOTIFY_SEVERITY
            and bug["status"] in TODO_STATUS_CODES
            and (previous_severity != MAIL_NOTIFY_SEVERITY or previous_assignee_id != next_assignee_id)
        ):
            notify_sent, notify_message = maybe_send_severe_bug_assignment_notification(
                bug_id=bug_id,
                assignee_user_id=next_assignee_id,
                trigger_reason="编辑后进入待办",
                operator_name=str(g.current_user["name"] or ""),
            )
            flash(
                f"严重Bug通知已发送，{notify_message}" if notify_sent else f"严重Bug通知未发送：{notify_message}",
                "success" if notify_sent else "error",
            )
        return redirect(url_for("bug_detail", bug_id=bug_id, tab="detail", next=back_url))

    @app.route("/bugs/<int:bug_id>/delete", methods=["POST"])
    def delete_bug(bug_id: int) -> Response:
        db = get_db()
        bug = fetch_bug(bug_id)
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        if not can_manage_bug(bug):
            flash("仅管理员或创建人可删除该 Bug。", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id))
        db.execute("DELETE FROM notifications WHERE bug_id = ?", (bug_id,))
        db.execute("DELETE FROM bug_attachments WHERE bug_id = ?", (bug_id,))
        db.execute("DELETE FROM bug_history WHERE bug_id = ?", (bug_id,))
        db.execute("DELETE FROM bug_comments WHERE bug_id = ?", (bug_id,))
        db.execute("DELETE FROM bugs WHERE id = ?", (bug_id,))
        db.commit()
        bump_bug_sync_token()
        flash("Bug 已删除。", "success")
        return redirect(url_for("bug_list"))

    @app.route("/attachments/<int:attachment_id>")
    def view_attachment(attachment_id: int) -> str | Response:
        attachment = fetch_attachment(attachment_id)
        if attachment is None:
            flash("附件不存在。", "error")
            return redirect(url_for("bug_list"))
        file_path = Path(attachment["file_path"])
        if not file_path.exists():
            flash("附件文件不存在。", "error")
            return redirect(url_for("bug_list"))
        preview_mode = request.args.get("preview", "").strip() == "1"
        if preview_mode:
            return render_template("attachment_preview.html", attachment=attachment)
        return send_file(file_path, mimetype=attachment["content_type"] or "application/octet-stream", as_attachment=False, download_name=attachment["filename"])

    @app.route("/bugs/<int:bug_id>/update", methods=["POST"])
    def update_bug(bug_id: int) -> Response:
        db = get_db()
        bug = fetch_bug(bug_id)
        if bug is None:
            flash("未找到对应的 Bug。", "error")
            return redirect(url_for("bug_list"))
        action = request.form.get("action", "").strip()
        redirect_target = local_redirect_target(
            request.form.get("redirect_to", ""),
            url_for("bug_detail", bug_id=bug_id),
        )
        if action == "change_platform":
            if not can_edit_bug_platform(bug):
                flash("仅当前处理人、提Bug人或管理员可修改端。", "error")
                return redirect(redirect_target)
            selected_platform = request.form.get("platform", "").strip()
            if selected_platform not in BUG_PLATFORM_OPTIONS:
                flash("请选择有效的端。", "error")
                return redirect(redirect_target)
            if selected_platform == (bug["platform"] or ""):
                flash("端未发生变化。", "success")
                return redirect(redirect_target)
            selected_module = bug_notify_key_for_platform(selected_platform)
            db.execute(
                "UPDATE bugs SET platform = ?, module = ?, updated_at = ? WHERE id = ?",
                (selected_platform, selected_module, current_time(), bug_id),
            )
            add_history(
                db,
                bug_id,
                "更新端",
                f"{g.current_user['name']} 将端更新为 {selected_platform}",
                g.current_user["name"],
                environment_snapshot=bug["environment"] or "",
                status_snapshot=bug["status"],
                assignee_snapshot=bug["assignee_name"] or "",
            )
            db.commit()
            bump_bug_sync_token()
            flash("端信息已更新。", "success")
            return redirect(redirect_target)
        note = request.form.get("resolution_note", "").strip()
        assignee_target = int(request.form.get("assignee_id", "0") or 0)
        try:
            new_status, new_assignee_id, new_previous_assignee_id, action_label, detail = apply_bug_action(
                db=db,
                bug=bug,
                action=action,
                operator_name=g.current_user["name"],
                note=note,
                assignee_id=assignee_target or None,
            )
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(redirect_target)
        db.execute(
            """
            UPDATE bugs
            SET status = ?, assignee_id = ?, previous_assignee_id = ?, resolution_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (new_status, new_assignee_id, new_previous_assignee_id, note or bug["resolution_note"], current_time(), bug_id),
        )
        target_user = fetch_user(new_assignee_id)
        add_history(
            db,
            bug_id,
            action_label,
            detail,
            g.current_user["name"],
            environment_snapshot=bug["environment"] or "",
            status_snapshot=new_status,
            assignee_snapshot=target_user["name"] if target_user else bug["assignee_name"],
        )
        db.commit()
        bump_bug_sync_token()
        flash("Bug 流转已更新。", "success")
        should_notify_assignee = (
            str(bug["severity"] or "") == MAIL_NOTIFY_SEVERITY
            and new_status in TODO_STATUS_CODES
            and (
                new_assignee_id != int(bug["assignee_id"] or 0)
                or action in {"resolve", "reject", "reassign"}
                or (action == "change_status" and new_status == "pending_verification")
            )
        )
        if should_notify_assignee:
            notify_sent, notify_message = maybe_send_severe_bug_assignment_notification(
                bug_id=bug_id,
                assignee_user_id=new_assignee_id,
                trigger_reason=action_label,
                operator_name=str(g.current_user["name"] or ""),
            )
            flash(
                f"严重Bug通知已发送，{notify_message}" if notify_sent else f"严重Bug通知未发送：{notify_message}",
                "success" if notify_sent else "error",
            )
        todo_redirect_path = url_for("my_todo_page")
        returns_to_todo_page = (
            redirect_target == todo_redirect_path
            or redirect_target.startswith(f"{todo_redirect_path}?")
            or redirect_target.startswith(f"{todo_redirect_path}#")
        )
        bug_list_path = url_for("bug_list")
        returns_to_bug_list = (
            redirect_target == bug_list_path
            or redirect_target.startswith(f"{bug_list_path}?")
            or redirect_target.startswith(f"{bug_list_path}#")
        )
        if new_status == "closed" and not returns_to_todo_page:
            return redirect(redirect_target if returns_to_bug_list else bug_list_path)
        return redirect(redirect_target)

    @app.route("/reports/testing")
    def testing_report() -> str:
        version = request.args.get("version", "").strip()
        page = max(1, int(request.args.get("page", "1") or 1))
        return render_template("report.html", report=fetch_report_data(version, page))

    @app.route("/reports/testing/chart.png")
    def testing_report_chart() -> Response:
        version = request.args.get("version", "").strip()
        image = build_case_chart_bytes(version=version)
        return Response(image, mimetype="image/svg+xml")

    @app.route("/reports/testing/export")
    def export_testing_report() -> Response:
        version = request.args.get("version", "").strip()
        report = fetch_report_data(version)
        html = render_template("report_export.html", report=report)
        filename = f"testing-report-{app_now().strftime('%Y%m%d-%H%M')}.html"
        return Response(html, mimetype="text/html; charset=utf-8", headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'})

    @app.route("/profile", methods=["GET", "POST"])
    def profile_page() -> str | Response:
        if g.current_user is None:
            return redirect(url_for("login"))
        db = get_db()
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            email = request.form.get("email", "").strip()
            password = request.form.get("password", "").strip()
            if not name:
                flash("姓名不能为空。", "error")
            else:
                if password:
                    db.execute(
                        "UPDATE users SET name = ?, email = ?, password = ? WHERE id = ?",
                        (name, email, password, g.current_user["id"]),
                    )
                else:
                    db.execute(
                        "UPDATE users SET name = ?, email = ? WHERE id = ?",
                        (name, email, g.current_user["id"]),
                    )
                db.commit()
                if password:
                    session.clear()
                    flash("密码已修改，请使用新密码重新登录。", "success")
                    return redirect(url_for("login"))
                flash("个人信息已更新。", "success")
                return redirect(url_for("profile_page"))
        return render_template("profile.html", profile_user=fetch_user(int(g.current_user["id"])))

    @app.route("/admin", methods=["GET", "POST"])
    def admin_center() -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        db = get_db()
        if request.method == "POST":
            redirect_target = admin_redirect_target()
            entity = request.form.get("entity", "").strip()
            action = request.form.get("action", "").strip()
            if entity == "project":
                project_id = int(request.form.get("project_id", "0") or 0)
                existing_project = fetch_project(project_id) if action == "update" else None
                name = request.form.get("name", "").strip()
                description = request.form.get("description", "").strip()
                bug_notify_enabled = 1 if request.form.get("bug_notify_enabled") == "1" else 0
                bug_notify_webhook = request.form.get("bug_notify_webhook", "").strip()
                bug_notify_secret = request.form.get("bug_notify_secret", "").strip()
                bug_notify_base_url = request.form.get("bug_notify_base_url", "").strip()
                bug_notify_rules: list[dict[str, object]] = []
                preserve_notify_rules = request.form.get("preserve_notify_rules") == "1"
                # SPA 不回显敏感配置，字段缺省时沿用现有值。
                if existing_project is not None:
                    if "bug_notify_webhook" not in request.form:
                        bug_notify_webhook = str(existing_project["bug_notify_webhook"] or "")
                    if "bug_notify_secret" not in request.form:
                        bug_notify_secret = str(existing_project["bug_notify_secret"] or "")
                if action in {"create", "update"}:
                    try:
                        validate_project_bug_notify_settings(
                            bug_notify_enabled,
                            bug_notify_webhook,
                            bug_notify_base_url,
                            label="默认新建 Bug 群通知",
                        )
                        if not preserve_notify_rules:
                            bug_notify_rules = parse_project_bug_notify_rule_form(request.form)
                    except ValueError as exc:
                        flash(str(exc), "error")
                        return redirect(redirect_target)
                if action == "create":
                    if not name:
                        flash("项目名称不能为空。", "error")
                    elif db.execute("SELECT 1 FROM projects WHERE name = ?", (name,)).fetchone():
                        flash("项目名称已存在。", "error")
                    else:
                        cursor = db.execute(
                            """
                            INSERT INTO projects (
                                name, description, bug_notify_enabled, bug_notify_webhook,
                                bug_notify_secret, bug_notify_base_url, created_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                name,
                                description,
                                bug_notify_enabled,
                                bug_notify_webhook,
                                bug_notify_secret,
                                bug_notify_base_url,
                                current_time(),
                            ),
                        )
                        save_project_bug_notify_rules(int(cursor.lastrowid), bug_notify_rules)
                        db.commit()
                        flash("项目已创建。", "success")
                elif action == "update":
                    project = existing_project
                    if project is None:
                        flash("未找到对应项目。", "error")
                    elif not name:
                        flash("项目名称不能为空。", "error")
                    elif db.execute("SELECT 1 FROM projects WHERE name = ? AND id != ?", (name, project_id)).fetchone():
                        flash("项目名称已存在。", "error")
                    else:
                        db.execute(
                            """
                            UPDATE projects
                            SET name = ?, description = ?, bug_notify_enabled = ?,
                                bug_notify_webhook = ?, bug_notify_secret = ?, bug_notify_base_url = ?
                            WHERE id = ?
                            """,
                            (
                                name,
                                description,
                                bug_notify_enabled,
                                bug_notify_webhook,
                                bug_notify_secret,
                                bug_notify_base_url,
                                project_id,
                            ),
                        )
                        if not preserve_notify_rules:
                            save_project_bug_notify_rules(project_id, bug_notify_rules)
                        db.commit()
                        if current_project_id() == project_id:
                            set_current_project(project_id)
                        flash("项目已更新。", "success")
                elif action == "delete":
                    project = fetch_project(project_id)
                    if project is None:
                        flash("未找到对应项目。", "error")
                    else:
                        deleted_summary = delete_project_with_related_data(project_id)
                        if current_project_id() == project_id:
                            remaining = fetch_projects()
                            if remaining:
                                set_current_project(int(remaining[0]["id"]))
                            else:
                                session.pop("project_id", None)
                        flash(
                            f"项目已删除，并清理 Bug {deleted_summary['bugs']} 条、需求 {deleted_summary['requirements']} 条、用例 {deleted_summary['cases']} 条。",
                            "success",
                        )
            elif entity == "user":
                user_id = int(request.form.get("user_id", "0") or 0)
                name = request.form.get("name", "").strip()
                account_type, role_code, role = resolve_user_role_from_form(request.form)
                username = request.form.get("username", "").strip()
                password = request.form.get("password", "").strip()
                email = request.form.get("email", "").strip()
                valid_role_codes = set(ROLE_LABELS) | {ADMIN_ROLE_CODE}
                if action == "create":
                    if account_type not in {"member", "admin"}:
                        flash("请选择有效的账号类型。", "error")
                    elif role_code not in valid_role_codes or not role:
                        flash("请选择有效的成员角色。", "error")
                    elif account_type == "member" and role_code == ADMIN_ROLE_CODE:
                        flash("普通成员不能使用管理员角色。", "error")
                    elif not all([name, username, password, email]):
                        flash("请完整填写账号信息，邮箱为必填。", "error")
                    elif db.execute("SELECT 1 FROM users WHERE name = ?", (name,)).fetchone():
                        flash("姓名已存在。", "error")
                    elif db.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone():
                        flash("账号已存在。", "error")
                    else:
                        db.execute(
                            """
                            INSERT INTO users (name, role, role_code, account_type, username, password, email, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (name, role, role_code, account_type, username, password, email, current_time()),
                        )
                        db.commit()
                        flash("账号已创建。", "success")
                elif action == "update":
                    target_user = fetch_user(user_id)
                    if target_user is None:
                        flash("未找到对应账号。", "error")
                    elif role_code not in valid_role_codes or not role:
                        flash("请选择有效的成员角色。", "error")
                    elif account_type == "member" and role_code == ADMIN_ROLE_CODE:
                        flash("普通成员不能使用管理员角色。", "error")
                    elif not all([name, username, email]):
                        flash("请完整填写账号信息，邮箱为必填。", "error")
                    elif db.execute("SELECT 1 FROM users WHERE name = ? AND id != ?", (name, user_id)).fetchone():
                        flash("姓名已存在。", "error")
                    elif db.execute("SELECT 1 FROM users WHERE username = ? AND id != ?", (username, user_id)).fetchone():
                        flash("账号已存在。", "error")
                    elif account_type not in {"member", "admin"}:
                        flash("请选择有效的账号类型。", "error")
                    else:
                        if password:
                            db.execute(
                                """
                                UPDATE users
                                SET name = ?, role = ?, role_code = ?, account_type = ?, username = ?, password = ?, email = ?
                                WHERE id = ?
                                """,
                                (name, role, role_code, account_type, username, password, email, user_id),
                            )
                        else:
                            db.execute(
                                """
                                UPDATE users
                                SET name = ?, role = ?, role_code = ?, account_type = ?, username = ?, email = ?
                                WHERE id = ?
                                """,
                                (name, role, role_code, account_type, username, email, user_id),
                            )
                        db.commit()
                        flash("账号已更新。", "success")
                elif action == "delete":
                    target_user = fetch_user(user_id)
                    if target_user is None:
                        flash("未找到对应账号。", "error")
                    elif g.current_user is not None and int(g.current_user["id"]) == user_id:
                        flash("当前登录账号不可删除。", "error")
                    elif user_usage_count(user_id) > 0:
                        flash("该账号仍被 Bug 流转使用，暂不可删除。", "error")
                    else:
                        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
                        db.commit()
                        flash("账号已删除。", "success")
            elif entity == "mail":
                flash("邮件发送已取消，请使用项目新建 Bug 群通知。", "error")
            elif entity == "report_notify":
                if action == "update":
                    try:
                        update_group_report_settings(request.form)
                    except ValueError as exc:
                        flash(str(exc), "error")
                    else:
                        flash("群测试报告通知设置已保存。", "success")
                elif action == "send_test":
                    try:
                        manual_note = request.form.get("manual_note", "").strip()
                        group_report_settings = fetch_group_report_settings()
                        effective_manual_note = manual_note or group_report_settings["manual_note"].strip()
                        effective_tracking_progress = group_report_settings["tracking_progress"].strip()
                        message_format = group_report_settings["message_format"].strip() or DEFAULT_GROUP_REPORT_SETTINGS["message_format"]
                        project_name, version_name, _sent_at = send_testing_report_to_group(
                            force=True,
                            mark_daily_sent=False,
                            manual_note=manual_note,
                        )
                    except Exception as exc:
                        flash(f"群测试报告发送失败：{exc}", "error")
                    else:
                        note_suffix = "，已附带补充信息" if (effective_manual_note or effective_tracking_progress) else ""
                        format_label = GROUP_REPORT_MESSAGE_FORMATS.get(message_format, GROUP_REPORT_MESSAGE_FORMATS["card"])
                        flash(f"测试发送成功，已发送 {project_name} / {version_name} {format_label}测试报告到群{note_suffix}。", "success")
            return redirect(redirect_target)
        return render_template("admin.html", admin_cards=admin_dashboard_cards())

    @app.route("/admin/projects")
    def admin_projects_page() -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        return render_template("admin_projects.html", projects=fetch_projects())

    @app.route("/admin/users")
    def admin_users_page() -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        return render_template("admin_users.html", users=fetch_users())

    @app.route("/admin/mail")
    def admin_mail_page() -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        flash("邮件发送已取消，请使用项目新建 Bug 群通知。", "error")
        return redirect(url_for("admin_projects_page"))

    @app.route("/admin/report-notify")
    def admin_report_notify_page() -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        settings = fetch_group_report_settings()
        selected_project_id = int(settings["project_id"]) if settings["project_id"].isdigit() else 0
        return render_template(
            "admin_report_notify.html",
            report_notify_settings=settings,
            report_projects=fetch_projects(),
            report_versions=fetch_report_versions(project_id=selected_project_id or None),
        )

    @app.route("/admin/projects/<int:project_id>")
    def admin_project_detail(project_id: int) -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        project = fetch_project(project_id)
        if project is None:
            flash("未找到对应项目。", "error")
            return redirect(url_for("admin_projects_page"))
        return render_template(
            "admin_project_detail.html",
            project=project,
            usage_count=project_usage_count(project_id),
            bug_notify_rules=fetch_project_bug_notify_rule_options(project_id),
        )

    @app.route("/admin/users/<int:user_id>")
    def admin_user_detail(user_id: int) -> str | Response:
        denied = require_admin_access()
        if denied is not None:
            return denied
        target_user = fetch_user(user_id)
        if target_user is None:
            flash("未找到对应账号。", "error")
            return redirect(url_for("admin_users_page"))
        return render_template(
            "admin_user_detail.html",
            target_user=target_user,
            usage_count=user_usage_count(user_id),
        )

    def api_json_value(value: Any) -> Any:
        """将数据库行和嵌套集合转换为稳定的 JSON 数据。"""
        if isinstance(value, sqlite3.Row):
            return {key: api_json_value(value[key]) for key in value.keys()}
        if isinstance(value, dict):
            return {str(key): api_json_value(item) for key, item in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [api_json_value(item) for item in value]
        if isinstance(value, Path):
            return str(value)
        return value

    def api_user_value(user: sqlite3.Row | None) -> dict[str, Any] | None:
        if user is None:
            return None
        return {
            "id": int(user["id"]),
            "name": user["name"] or "",
            "role": user["role"] or "",
            "role_code": user["role_code"] or "",
            "account_type": user["account_type"] or "member",
            "username": user["username"] or "",
            "email": user["email"] or "",
            "created_at": user["created_at"] or "",
        }

    def api_project_value(project: sqlite3.Row | None, include_settings: bool = False) -> dict[str, Any] | None:
        if project is None:
            return None
        payload = {
            "id": int(project["id"]),
            "name": project["name"] or "",
            "description": project["description"] or "",
            "bug_notify_enabled": bool(project["bug_notify_enabled"]),
            "created_at": project["created_at"] or "",
        }
        if include_settings:
            payload.update(
                {
                    "bug_notify_webhook_configured": bool(project["bug_notify_webhook"]),
                    "bug_notify_secret_configured": bool(project["bug_notify_secret"]),
                    "bug_notify_base_url": project["bug_notify_base_url"] or "",
                    "bug_notify_last_sent_at": project["bug_notify_last_sent_at"] or "",
                    "bug_notify_last_result": project["bug_notify_last_result"] or "",
                }
            )
        return payload

    def api_attachment_value(attachment: sqlite3.Row) -> dict[str, Any]:
        return {
            "id": int(attachment["id"]),
            "filename": attachment["filename"] or "",
            "content_type": attachment["content_type"] or "",
            "source_field": attachment["source_field"] or "attachments",
            "is_image": is_image_attachment(attachment),
            "url": url_for("view_attachment", attachment_id=int(attachment["id"])),
        }

    def api_legacy_action_result(result: Any) -> Response | tuple[Response, int]:
        """兼容旧业务处理函数，并把表单跳转转换为前端可消费的 JSON。"""
        status_code = 200
        response = result
        if isinstance(result, tuple):
            response = result[0]
            status_code = int(result[1])
        if isinstance(response, Response) and response.is_json:
            return response, status_code
        messages = get_flashed_messages(with_categories=True)
        has_error = any(category == "error" for category, _message in messages)
        payload = {
            "ok": not has_error,
            "message": messages[-1][1] if messages else ("操作成功。" if not has_error else "操作失败。"),
            "messages": [{"type": category, "text": message} for category, message in messages],
            "redirect_url": response.location if isinstance(response, Response) and response.location else "",
        }
        return jsonify(payload), 400 if has_error else status_code

    @app.get("/api/v1/health")
    def api_health() -> Response:
        return jsonify({"ok": True, "service": "ewll-bug", "version": "v1"})

    @app.post("/api/v1/auth/login")
    def api_login() -> tuple[Response, int] | Response:
        payload = request.get_json(silent=True) if request.is_json else request.form
        username = str(payload.get("username", "") or "").strip()
        password = str(payload.get("password", "") or "").strip()
        user = fetch_user_by_credentials(username, password)
        if user is None:
            return jsonify({"ok": False, "message": "账号或密码错误。"}), 401
        session["user_id"] = int(user["id"])
        default_project_id = default_project_id_for_user(int(user["id"]))
        if default_project_id is not None:
            session["project_id"] = default_project_id
        return jsonify({"ok": True, "message": "登录成功。", "data": {"user": api_user_value(user)}})

    @app.post("/api/v1/auth/logout")
    def api_logout() -> Response:
        session.clear()
        return jsonify({"ok": True, "message": "已退出登录。"})

    @app.get("/api/v1/bootstrap")
    def api_bootstrap() -> Response:
        return jsonify(
            {
                "ok": True,
                "data": {
                    "user": api_user_value(g.current_user),
                    "current_project": api_project_value(g.current_project),
                    "projects": [api_project_value(item) for item in g.projects],
                    "summary": api_json_value(fetch_summary()),
                    "is_admin": is_admin(),
                    "options": {
                        "statuses": [{"value": code, "label": label} for code, label in STATUS_OPTIONS],
                        "severities": BUG_SEVERITY_OPTIONS,
                        "priorities": BUG_PRIORITY_OPTIONS,
                        "platforms": BUG_PLATFORM_OPTIONS,
                        "requirement_statuses": [
                            {"value": code, "label": label} for code, label in REQUIREMENT_STATUS_OPTIONS
                        ],
                        "roles": [{"value": code, "label": label} for code, label in ROLE_OPTIONS],
                    },
                },
            }
        )

    @app.get("/api/v1/summary")
    def api_summary() -> Response:
        # 顶部导航定时刷新时只返回角标所需摘要，避免重复加载完整初始化数据。
        return jsonify({"ok": True, "data": {"summary": api_json_value(fetch_summary())}})

    @app.post("/api/v1/projects/current")
    def api_switch_project() -> tuple[Response, int] | Response:
        payload = request.get_json(silent=True) if request.is_json else request.form
        project_id = int(payload.get("project_id", 0) or 0)
        project = fetch_project(project_id)
        if project is None:
            return jsonify({"ok": False, "message": "项目不存在。"}), 404
        set_current_project(project_id)
        return jsonify({"ok": True, "data": {"current_project": api_project_value(project)}})

    @app.get("/api/v1/bugs")
    def api_bug_list() -> Response:
        filters = fetch_filters()
        page = request_page()
        bug_page = fetch_bug_page(filters, page)
        return jsonify(
            {
                "ok": True,
                "data": {
                    "page": api_json_value(bug_page),
                    "filters": filters,
                    "summary": fetch_bug_summary(filters),
                    "versions": fetch_bug_versions(),
                    "users": [api_user_value(item) for item in fetch_bug_assignee_users()],
                    "requirements": api_json_value(fetch_requirements()),
                    "cases": api_json_value(fetch_cases_for_project()),
                },
            }
        )

    @app.post("/api/v1/bugs")
    def api_create_bug() -> Response | tuple[Response, int]:
        return api_legacy_action_result(create_bug())

    @app.get("/api/v1/bugs/<int:bug_id>")
    def api_bug_detail(bug_id: int) -> tuple[Response, int] | Response:
        bug = fetch_bug(bug_id)
        if bug is None:
            return jsonify({"ok": False, "message": "未找到对应的 Bug。"}), 404
        history = fetch_bug_history(bug_id)
        comments = fetch_bug_comments(bug_id)
        attachments = fetch_bug_attachments(bug_id)
        return jsonify(
            {
                "ok": True,
                "data": {
                    "bug": api_json_value(bug),
                    "history": api_json_value(history),
                    "comments": api_json_value(build_bug_comment_threads(comments, fetch_users(), fetch_comment_mention_states(bug_id))),
                    "mention_users": [api_user_value(item) for item in fetch_users()],
                    "attachments": [api_attachment_value(item) for item in attachments],
                    "users": [api_user_value(item) for item in fetch_bug_assignee_users()],
                    "requirements": api_json_value(fetch_requirements()),
                    "cases": api_json_value(fetch_cases_for_project()),
                    "permissions": {
                        "can_manage": can_manage_bug(bug),
                        "can_edit_platform": can_edit_bug_platform(bug),
                    },
                    "allowed_statuses": allowed_status_transitions(str(bug["status"] or "")),
                },
            }
        )

    @app.post("/api/v1/bugs/<int:bug_id>/edit")
    def api_edit_bug(bug_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(edit_bug(bug_id))

    @app.post("/api/v1/bugs/<int:bug_id>/actions")
    def api_update_bug(bug_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(update_bug(bug_id))

    @app.post("/api/v1/bugs/<int:bug_id>/delete")
    def api_delete_bug(bug_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(delete_bug(bug_id))

    @app.post("/api/v1/bugs/<int:bug_id>/comments")
    def api_add_bug_comment(bug_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(add_bug_comment(bug_id))

    @app.post("/api/v1/bugs/<int:bug_id>/comments/<int:comment_id>/delete")
    def api_delete_bug_comment(bug_id: int, comment_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(delete_bug_comment(bug_id, comment_id))

    @app.get("/api/v1/todos")
    def api_todos() -> Response:
        todos = fetch_my_todos()
        return jsonify({"ok": True, "data": {"items": api_json_value(todos), "count": len(todos)}})

    @app.get("/api/v1/notifications")
    def api_notifications() -> Response:
        state = request.args.get("state", "").strip()
        state = state if state in {"", "unread"} else ""
        user_id = int(g.current_user["id"])
        return jsonify(
            {
                "ok": True,
                "data": {
                    "items": api_json_value(fetch_user_notifications(user_id, state=state)),
                    "state": state,
                    "unread_count": count_user_notifications(user_id, unread_only=True),
                    "total_count": count_user_notifications(user_id),
                },
            }
        )

    @app.post("/api/v1/notifications/read-all")
    def api_notifications_read_all() -> Response:
        count = mark_all_notifications_read(int(g.current_user["id"]))
        return jsonify({"ok": True, "message": f"已将 {count} 条消息标为已读。", "data": {"count": count}})

    @app.post("/api/v1/notifications/<int:notification_id>/read")
    def api_notification_read(notification_id: int) -> tuple[Response, int] | Response:
        user_id = int(g.current_user["id"])
        notification = fetch_notification(notification_id, user_id)
        if notification is None:
            return jsonify({"ok": False, "message": "消息不存在或已删除。"}), 404
        mark_notification_read(notification_id, user_id)
        return jsonify({"ok": True, "data": {"link_path": notification["link_path"] or ""}})

    @app.get("/api/v1/cases")
    def api_cases() -> Response:
        version = request.args.get("version", "").strip()
        documents = fetch_case_documents(version)
        return jsonify(
            {
                "ok": True,
                "data": {
                    "documents": api_json_value(documents),
                    "tree": api_json_value(build_case_tree(documents)),
                    "versions": fetch_case_versions(),
                    "distribution": api_json_value(execution_distribution(version=version)),
                },
            }
        )

    @app.get("/api/v1/cases/<int:document_id>")
    def api_case_document(document_id: int) -> tuple[Response, int] | Response:
        bundle = fetch_case_document_bundle(document_id)
        if bundle is None:
            return jsonify({"ok": False, "message": "未找到对应的在线文档。"}), 404
        return jsonify({"ok": True, "data": api_json_value(bundle)})

    @app.post("/api/v1/cases/<int:document_id>/autosave")
    def api_case_autosave(document_id: int) -> Response | tuple[Response, int]:
        return autosave_case_document(document_id)

    @app.post("/api/v1/cases/manage")
    def api_manage_cases() -> Response | tuple[Response, int]:
        return api_legacy_action_result(manage_case_library())

    @app.post("/api/v1/cases/upload")
    def api_upload_cases() -> Response | tuple[Response, int]:
        return api_legacy_action_result(upload_cases())

    @app.get("/api/v1/requirements")
    def api_requirements() -> Response:
        filters = {
            "keyword": request.args.get("keyword", "").strip(),
            "version": request.args.get("version", "").strip(),
        }
        return jsonify(
            {
                "ok": True,
                "data": {
                    "page": api_json_value(fetch_requirement_page(filters, request_page())),
                    "filters": filters,
                    "versions": fetch_requirement_versions(),
                    "summary": fetch_requirement_summary(),
                },
            }
        )

    @app.post("/api/v1/requirements")
    def api_create_requirement() -> Response | tuple[Response, int]:
        return api_legacy_action_result(create_requirement())

    @app.get("/api/v1/requirements/<int:requirement_id>")
    def api_requirement_detail(requirement_id: int) -> tuple[Response, int] | Response:
        requirement = fetch_requirement(requirement_id)
        if requirement is None:
            return jsonify({"ok": False, "message": "未找到对应需求。"}), 404
        return jsonify(
            {
                "ok": True,
                "data": {
                    "requirement": api_json_value(requirement),
                    "bugs": api_json_value(fetch_requirement_bugs(requirement_id)),
                    "can_manage": can_manage_requirement(requirement),
                },
            }
        )

    @app.post("/api/v1/requirements/<int:requirement_id>/edit")
    def api_edit_requirement(requirement_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(update_requirement(requirement_id))

    @app.post("/api/v1/requirements/<int:requirement_id>/delete")
    def api_delete_requirement(requirement_id: int) -> Response | tuple[Response, int]:
        return api_legacy_action_result(delete_requirement(requirement_id))

    @app.get("/api/v1/reports/testing")
    def api_testing_report() -> Response:
        version = request.args.get("version", "").strip()
        return jsonify({"ok": True, "data": api_json_value(fetch_report_data(version, request_page()))})

    @app.get("/api/v1/profile")
    def api_profile() -> Response:
        return jsonify({"ok": True, "data": {"user": api_user_value(fetch_user(int(g.current_user["id"])))}})

    @app.post("/api/v1/profile")
    def api_update_profile() -> Response | tuple[Response, int]:
        return api_legacy_action_result(profile_page())

    @app.get("/api/v1/admin")
    def api_admin() -> tuple[Response, int] | Response:
        if not is_admin():
            return jsonify({"ok": False, "message": "仅管理员可访问。"}), 403
        projects = fetch_projects()
        users = fetch_users()
        report_notify_settings = fetch_group_report_settings()
        report_notify_public = {
            key: value
            for key, value in report_notify_settings.items()
            if "secret" not in key and "webhook" not in key and "app_secret" not in key
        }
        report_notify_public.update(
            {
                "webhook_configured": bool(report_notify_settings["webhook_url"]),
                "secret_configured": bool(report_notify_settings["secret"]),
                "lark_app_secret_configured": bool(report_notify_settings["lark_app_secret"]),
            }
        )
        return jsonify(
            {
                "ok": True,
                "data": {
                    "cards": api_json_value(admin_dashboard_cards()),
                    "projects": [api_project_value(item, include_settings=True) for item in projects],
                    "users": [api_user_value(item) for item in users],
                    "project_usage": {str(item["id"]): project_usage_count(int(item["id"])) for item in projects},
                    "user_usage": {str(item["id"]): user_usage_count(int(item["id"])) for item in users},
                    "report_notify": report_notify_public,
                },
            }
        )

    @app.post("/api/v1/admin/actions")
    def api_admin_action() -> Response | tuple[Response, int]:
        if not is_admin():
            return jsonify({"ok": False, "message": "仅管理员可访问。"}), 403
        return api_legacy_action_result(admin_center())

    with app.app_context():
        init_db()
        run_migrations()
        seed_data()
        sync_case_execute_statuses()
        repaired_case_count = repair_misaligned_excel_cases(get_db())
        if repaired_case_count > 0:
            get_db().commit()

    if app.config.get("START_SCHEDULER"):
        start_mail_scheduler()

    return app


if __name__ == "__main__":
    app = create_app()
    app.run(
        debug=False,
        use_reloader=False,
        host=os.environ.get("HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5050")),
    )
